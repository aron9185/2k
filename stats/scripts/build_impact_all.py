from __future__ import annotations

import re
import statistics
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook

from build_cal_lane import (
    CalUniverseRow,
    apply_cal_normalization,
    build_source_index,
    canonical_id,
    compute_capped_z_score,
    detect_current_season,
    enrich_universe_rows,
    load_cal_universe,
    load_universe_csv,
    match_metric_row,
    normalize_name,
    parse_float,
    read_history_csv,
    resolve_details_csv_path,
    write_csv,
)
from build_finishing_layup import (
    ComponentSlot,
    build_player_contexts,
    interpolate_rating,
    percentile_inc,
)
from build_finishing_standing_dunk import standardize_rows, write_matrix_csv
from pull_nba_stats import EXPORT_DIR, HISTORY_DIR, MANUAL_DIR, ROOT


PROJECT_ROOT = ROOT.parent
AGGREGATE_TEAM_VALUES = {"2TM", "TOT", "Total"}


OVERALL_IMPACT_CURVE: Sequence[Tuple[float, float]] = (
    (0.06, 30.0),
    (0.12, 40.0),
    (0.20, 40.0),
    (0.28, 50.0),
    (0.36, 57.0),
    (0.44, 60.0),
    (0.52, 70.0),
    (0.60, 70.0),
    (0.68, 80.0),
    (0.76, 82.0),
    (0.84, 90.0),
    (0.90, 96.0),
    (0.96, 98.0),
    (1.00, 100.0),
)

OFFENSIVE_IMPACT_CURVE: Sequence[Tuple[float, float]] = (
    (0.05, 32.0),
    (0.12, 40.0),
    (0.20, 50.0),
    (0.30, 55.0),
    (0.42, 65.0),
    (0.54, 70.0),
    (0.66, 79.0),
    (0.76, 85.0),
    (0.85, 90.0),
    (0.92, 95.0),
    (0.97, 98.0),
    (1.00, 100.0),
)

DEFENSIVE_IMPACT_CURVE: Sequence[Tuple[float, float]] = (
    (0.05, 40.0),
    (0.10, 45.0),
    (0.16, 50.0),
    (0.24, 54.0),
    (0.32, 57.0),
    (0.42, 60.0),
    (0.52, 62.0),
    (0.62, 69.0),
    (0.70, 72.0),
    (0.78, 76.0),
    (0.86, 81.0),
    (0.92, 87.0),
    (0.97, 93.0),
    (1.00, 100.0),
)


OVERALL_COMPONENTS: Sequence[ComponentSlot] = (
    ComponentSlot("LEBRON", "LEBRON"),
    ComponentSlot("BPM", "BPM"),
    ComponentSlot("DPM", "DPM"),
    ComponentSlot("LEBRON WAR", "LEBRON WAR"),
    ComponentSlot("LEBRON Box Impact", "LEBRON Box Impact"),
    ComponentSlot("LEBRON Vs Role Average", "LEBRON Vs Role Average"),
    ComponentSlot("Multi-Year LEBRON", "Multi-Year LEBRON"),
    ComponentSlot("Stable On-Court Net Rating", "Stable On-Court Net Rating"),
    ComponentSlot("VORP", "VORP"),
    ComponentSlot("RAPTOR", "RAPTOR"),
    ComponentSlot("BPM LA-RAPM", "BPM LA-RAPM"),
    ComponentSlot("RAPTOR WAR", "RAPTOR WAR"),
    ComponentSlot("BPM RAPM", "BPM RAPM"),
    ComponentSlot("EBRON", "EBRON"),
    ComponentSlot("Luck-Adjusted RAPM", "Luck-Adjusted RAPM"),
    ComponentSlot("MAMBA", "MAMBA"),
    ComponentSlot("SPI", "SPI"),
    ComponentSlot("RPM", "RPM"),
    ComponentSlot("RPM WAR", "RPM WAR"),
    ComponentSlot("WS", "WS"),
    ComponentSlot("WS/48", "WS/48"),
    ComponentSlot("EPM", "EPM"),
)

OFFENSIVE_COMPONENTS: Sequence[ComponentSlot] = (
    ComponentSlot("O-LEBRON", "O-LEBRON"),
    ComponentSlot("O-LEBRON Box Impact", "O-LEBRON Box Impact"),
    ComponentSlot("O-BPM", "O-BPM"),
    ComponentSlot("O-DPM", "O-DPM"),
    ComponentSlot("O-DRIP", "O-DRIP"),
    ComponentSlot("LEBRON Offensive Points Added", "LEBRON Offensive Points Added"),
    ComponentSlot("O-LEBRON Vs Role Average", "O-LEBRON Vs Role Average"),
    ComponentSlot("Multi-Year O-LEBRON", "Multi-Year O-LEBRON"),
    ComponentSlot("O-BPM LA-RAPM", "O-BPM LA-RAPM"),
    ComponentSlot("O-EBRON", "O-EBRON"),
    ComponentSlot("Predictive O-LEBRON", "Predictive O-LEBRON"),
    ComponentSlot("O-EPM", "O-EPM"),
    ComponentSlot("OWS", "OWS"),
    ComponentSlot("O-MAMBA", "O-MAMBA"),
    ComponentSlot("OBPM RAPM", "OBPM RAPM"),
    ComponentSlot("O-RPM", "O-RPM"),
    ComponentSlot("O-RAPTOR", "O-RAPTOR"),
    ComponentSlot("LA-ORAPM", "LA-ORAPM"),
    ComponentSlot("O-SPI", "O-SPI"),
    ComponentSlot("Time Decay O-RAPM", "Time Decay O-RAPM"),
    ComponentSlot("Time Decay LA-ORAPM", "Time Decay LA-ORAPM"),
    ComponentSlot("Stable On-Court ORtg", "Stable On-Court ORtg"),
    ComponentSlot("Offensive Portability", "Offensive Portability"),
    ComponentSlot("Offensive Impact Luck", "Offensive Impact Luck"),
)

DEFENSIVE_COMPONENTS: Sequence[ComponentSlot] = (
    ComponentSlot("D-LEBRON", "D-LEBRON"),
    ComponentSlot("D-LEBRON Box Impact", "D-LEBRON Box Impact"),
    ComponentSlot("LEBRON Defensive Points Saved", "LEBRON Defensive Points Saved"),
    ComponentSlot("D-DRIP", "D-DRIP"),
    ComponentSlot("D-LEBRON Vs Role Average", "D-LEBRON Vs Role Average"),
    ComponentSlot("Multi-Year D-LEBRON", "Multi-Year D-LEBRON"),
    ComponentSlot("D-BPM LA-RAPM", "D-BPM LA-RAPM"),
    ComponentSlot("D-DPM", "D-DPM"),
    ComponentSlot("D-EBRON", "D-EBRON"),
    ComponentSlot("Defensive Portability", "Defensive Portability"),
    ComponentSlot("Predictive D-LEBRON", "Predictive D-LEBRON"),
    ComponentSlot("Time Decay D-RAPM", "Time Decay D-RAPM"),
    ComponentSlot("D-BPM", "D-BPM"),
    ComponentSlot("Defensive RAPTOR", "Defensive RAPTOR"),
    ComponentSlot("Defense Impact on Opponent Shot Quality", "Defense Impact on Opponent Shot Quality"),
    ComponentSlot("Defensive eFG% Impact", "Defensive eFG% Impact"),
    ComponentSlot("Defense Impact on Possession Quality", "Defense Impact on Possession Quality"),
    ComponentSlot("Defensive BPM 2.0", "Defensive BPM 2.0"),
    ComponentSlot("D-RPM", "D-RPM"),
    ComponentSlot("Luck-Adjusted DRAPM", "Luck-Adjusted DRAPM"),
    ComponentSlot("D-SPI", "D-SPI"),
    ComponentSlot("D-EPM", "D-EPM"),
    ComponentSlot("DWS", "DWS"),
    ComponentSlot("D-MAMBA", "D-MAMBA"),
    ComponentSlot("INV_Stable On-Court DRtg", "INV_Stable On-Court DRtg"),
)


WORKBOOK_COLUMNS = [
    "NBA ID",
    "Season",
    "Player",
    "Team(s)",
    "OVERALL IMPACT RATING",
    "",
    "OVERALL IMPACT",
    *[slot.header for slot in OVERALL_COMPONENTS],
    "OFFENSIVE IMPACT RATING",
    "",
    "OFFENSIVE IMPACT",
    *[slot.header for slot in OFFENSIVE_COMPONENTS],
    "DEFENSIVE IMPACT RATING",
    "",
    "DEFENSIVE IMPACT",
    *[slot.header for slot in DEFENSIVE_COMPONENTS],
]


@dataclass(frozen=True)
class SectionSpec:
    name: str
    rating_header: str
    rating_only_header: str
    component_slots: Sequence[ComponentSlot]
    weighted_groups: Sequence[Tuple[Sequence[str], float]]
    fallback_aliases: Sequence[str]
    curve: Sequence[Tuple[float, float]]


@dataclass
class MetricResult:
    raw_values: List[Optional[float]]
    normalized_values: List[Optional[float]]
    matched_by: List[str]
    mean_value: Optional[float]
    stdev_value: Optional[float]
    source_note: str


OVERALL_SECTION = SectionSpec(
    name="Overall Impact",
    rating_header="OVERALL IMPACT RATING",
    rating_only_header="Overall Impact Rating",
    component_slots=OVERALL_COMPONENTS,
    weighted_groups=((tuple(slot.metric_alias for slot in OVERALL_COMPONENTS), 1.0),),
    fallback_aliases=(),
    curve=OVERALL_IMPACT_CURVE,
)

OFFENSIVE_SECTION = SectionSpec(
    name="Offensive Impact",
    rating_header="OFFENSIVE IMPACT RATING",
    rating_only_header="Offensive Impact Rating",
    component_slots=OFFENSIVE_COMPONENTS,
    weighted_groups=(
        (
            (
                "O-LEBRON",
                "O-LEBRON Box Impact",
                "O-BPM",
                "O-DPM",
                "O-DRIP",
                "LEBRON Offensive Points Added",
                "O-LEBRON Vs Role Average",
                "Multi-Year O-LEBRON",
                "O-BPM LA-RAPM",
                "O-EBRON",
                "Predictive O-LEBRON",
                "O-EPM",
                "OWS",
                "O-MAMBA",
                "OBPM RAPM",
                "O-RPM",
                "O-RAPTOR",
                "LA-ORAPM",
                "O-SPI",
            ),
            0.75,
        ),
        (
            (
                "Time Decay O-RAPM",
                "Time Decay LA-ORAPM",
                "Stable On-Court ORtg",
                "Offensive Portability",
            ),
            0.20,
        ),
        (("Offensive Impact Luck",), 0.05),
    ),
    fallback_aliases=(),
    curve=OFFENSIVE_IMPACT_CURVE,
)

DEFENSIVE_SECTION = SectionSpec(
    name="Defensive Impact",
    rating_header="DEFENSIVE IMPACT RATING",
    rating_only_header="Defensive Impact Rating",
    component_slots=DEFENSIVE_COMPONENTS,
    weighted_groups=(
        (
            (
                "D-LEBRON",
                "D-LEBRON Box Impact",
                "LEBRON Defensive Points Saved",
                "D-DRIP",
                "D-LEBRON Vs Role Average",
                "Multi-Year D-LEBRON",
                "D-BPM LA-RAPM",
                "D-DPM",
                "D-EBRON",
                "Defensive Portability",
                "Predictive D-LEBRON",
                "Time Decay D-RAPM",
                "D-BPM",
                "Defensive RAPTOR",
                "Defense Impact on Opponent Shot Quality",
                "Defensive eFG% Impact",
                "Defense Impact on Possession Quality",
                "Defensive BPM 2.0",
                "D-RPM",
                "Luck-Adjusted DRAPM",
                "D-SPI",
                "D-EPM",
                "DWS",
                "D-MAMBA",
            ),
            0.85,
        ),
        (("INV_Stable On-Court DRtg",), 0.15),
    ),
    fallback_aliases=(),
    curve=DEFENSIVE_IMPACT_CURVE,
)

SECTIONS: Sequence[SectionSpec] = (
    OVERALL_SECTION,
    OFFENSIVE_SECTION,
    DEFENSIVE_SECTION,
)

RATING_ONLY_HEADERS = [
    "NBA_ID",
    "Season",
    "Player",
    "Team(s)",
    "Overall Impact Rating",
    "Offensive Impact Rating",
    "Defensive Impact Rating",
]

METRIC_SOURCE_NOTES: Dict[str, str] = {
    "LEBRON": "bball_index_impact.csv -> LEBRON, else lebron.csv fallback",
    "BPM": "stats/manual/bballref_advanced.xlsx -> BPM",
    "D-LEBRON": "bball_index_impact.csv -> D-LEBRON, else lebron.csv fallback",
    "VORP": "stats/manual/bballref_advanced.xlsx -> VORP",
    "WS": "stats/manual/bballref_advanced.xlsx -> WS",
    "WS/48": "stats/manual/bballref_advanced.xlsx -> WS/48",
    "O-LEBRON": "bball_index_impact.csv -> O-LEBRON, else lebron.csv fallback",
    "OWS": "stats/manual/bballref_advanced.xlsx -> OWS",
    "DWS": "stats/manual/bballref_advanced.xlsx -> DWS",
    "O-BPM": "stats/manual/bballref_advanced.xlsx -> OBPM",
    "D-BPM": "stats/manual/bballref_advanced.xlsx -> DBPM",
    "LEBRON WAR": "bball_index_impact.csv -> LEBRON WAR",
    "LEBRON Box Impact": "bball_index_impact.csv -> LEBRON Box Impact",
    "LEBRON Vs Role Average": "bball_index_impact.csv -> LEBRON Vs Role Average",
    "Multi-Year LEBRON": "bball_index_impact.csv -> Multi-Year LEBRON",
    "O-LEBRON Box Impact": "bball_index_impact.csv -> O-LEBRON Box Impact",
    "LEBRON Offensive Points Added": "bball_index_impact.csv -> LEBRON Offensive Points Added",
    "O-LEBRON Vs Role Average": "bball_index_impact.csv -> O-LEBRON Vs Role Average",
    "Multi-Year O-LEBRON": "bball_index_impact.csv -> Multi-Year O-LEBRON",
    "Predictive O-LEBRON": "bball_index_impact.csv -> Predictive O-LEBRON",
    "D-LEBRON Box Impact": "bball_index_impact.csv -> D-LEBRON Box Impact",
    "LEBRON Defensive Points Saved": "bball_index_impact.csv -> LEBRON Defensive Points Saved",
    "D-LEBRON Vs Role Average": "bball_index_impact.csv -> D-LEBRON Vs Role Average",
    "Multi-Year D-LEBRON": "bball_index_impact.csv -> Multi-Year D-LEBRON",
    "Predictive D-LEBRON": "bball_index_impact.csv -> Predictive D-LEBRON",
    "EPM": "dunksandthrees_epm.csv current-season overlay, else C:/2k/epm.xlsx -> EPM",
    "O-EPM": "dunksandthrees_epm.csv current-season overlay, else C:/2k/epm.xlsx -> OFF",
    "D-EPM": "dunksandthrees_epm.csv current-season overlay, else C:/2k/epm.xlsx -> DEF",
    "MAMBA": "mamba.csv -> MAMBA",
    "O-MAMBA": "mamba.csv -> O-MAMBA",
    "D-MAMBA": "mamba.csv -> D-MAMBA",
    "Stable On-Court Net Rating": "general_advanced.csv -> NET_RATING proxy",
    "Stable On-Court ORtg": "general_advanced.csv -> OFF_RATING proxy",
    "INV_Stable On-Court DRtg": "general_advanced.csv -> inverse DEF_RATING proxy",
    "Defensive eFG% Impact": "tracking_defensive_impact.csv -> inverse DEF_RIM_FG_PCT proxy",
}


def parse_args():
    import argparse

    parser = argparse.ArgumentParser(
        description="Build the combined Impact ratings export."
    )
    parser.add_argument(
        "--universe-csv",
        default=str(MANUAL_DIR / "playerlist.csv"),
        help="Season/player universe CSV. Defaults to stats/manual/playerlist.csv.",
    )
    parser.add_argument(
        "--workbook",
        default=str(MANUAL_DIR / "2k26_Temp_for_codex.xlsx"),
        help="Workbook fallback used for Cal role/minutes.",
    )
    parser.add_argument("--sheet", default="Cal", help="Workbook sheet used for role/minute fallback.")
    parser.add_argument(
        "--details-csv",
        default="",
        help=(
            "Optional role/minutes detail CSV. "
            "Pass stats/manual/player_universe.csv explicitly if you want that enrichment."
        ),
    )
    parser.add_argument(
        "--minutes-source",
        default=str(HISTORY_DIR / "general_traditional.csv"),
        help="History CSV used to refresh MIN and GP in real time.",
    )
    parser.add_argument("--minutes-column", default="MIN")
    parser.add_argument("--minutes-games-column", default="GP")
    parser.add_argument("--current-season", default="")
    parser.add_argument("--current-season-min-threshold", type=float, default=200.0)
    parser.add_argument("--standard-min-threshold", type=float, default=1000.0)
    parser.add_argument("--allow-id-fallback", action="store_true")
    parser.add_argument(
        "--bballref-source",
        default=str(MANUAL_DIR / "bballref_advanced.xlsx"),
    )
    parser.add_argument(
        "--epm-source",
        default=str(PROJECT_ROOT / "epm.xlsx"),
    )
    parser.add_argument("--mamba-source", default=str(HISTORY_DIR / "mamba.csv"))
    parser.add_argument("--lebron-source", default=str(HISTORY_DIR / "lebron.csv"))
    parser.add_argument(
        "--bball-index-impact-source",
        default=str(HISTORY_DIR / "bball_index_impact.csv"),
        help="Optional bball-index impact export used as the primary LEBRON-family source.",
    )
    parser.add_argument("--advanced-source", default=str(HISTORY_DIR / "general_advanced.csv"))
    parser.add_argument(
        "--dunks-epm-source",
        default=str(HISTORY_DIR / "dunksandthrees_epm.csv"),
        help="Optional current-season Dunks & Threes EPM CSV overlay.",
    )
    parser.add_argument(
        "--tracking-defensive-impact-source",
        default=str(HISTORY_DIR / "tracking_defensive_impact.csv"),
    )
    parser.add_argument("--output-prefix", default="impact_all")
    return parser.parse_args()


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() == "nan":
        return ""
    return text


def parse_metric_float(value: object) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            return float(value)
        except Exception:
            return None

    text = clean_text(value)
    if not text:
        return None

    text = (
        text.replace("\u2212", "-")
        .replace("−", "-")
        .replace("–", "-")
        .replace("—", "-")
        .replace("%", "")
        .replace(",", "")
        .replace("\u00a0", " ")
    )
    if text.startswith("+"):
        text = text[1:]
    try:
        return float(text)
    except Exception:
        return None


def canonical_season(value: object) -> str:
    text = clean_text(value)
    if not text:
        return ""

    match = re.match(r"^(\d{4})-(\d{2})$", text)
    if match:
        return f"{match.group(1)}-{match.group(2)}"

    match = re.match(r"^(\d{4})-(\d{4})$", text)
    if match:
        return f"{match.group(1)}-{match.group(2)[-2:]}"

    numeric = parse_metric_float(text)
    if numeric is None:
        return text
    year = int(numeric)
    return f"{year - 1}-{year % 100:02d}"


def load_excel_rows(path: Path, sheet_name: str = "") -> List[Dict[str, object]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
        iterator = worksheet.iter_rows(values_only=True)
        header_row = next(iterator, None)
        if header_row is None:
            return []
        headers = [clean_text(value) for value in header_row]

        rows: List[Dict[str, object]] = []
        for values in iterator:
            if values is None:
                continue
            row: Dict[str, object] = {}
            has_any_value = False
            for index, header in enumerate(headers):
                if not header:
                    continue
                value = values[index] if index < len(values) else None
                if value is not None and clean_text(value):
                    has_any_value = True
                row[header] = value
            if has_any_value:
                rows.append(row)
        return rows
    finally:
        workbook.close()


def load_bballref_advanced_rows(path: Path) -> List[Dict[str, object]]:
    raw_rows = load_excel_rows(path)
    grouped: Dict[Tuple[str, str], List[Dict[str, object]]] = {}

    for row in raw_rows:
        season = canonical_season(row.get("Season", ""))
        player = clean_text(row.get("Player", ""))
        if not season or not player:
            continue
        grouped.setdefault((season, normalize_name(player)), []).append(row)

    standardized: List[Dict[str, object]] = []
    for (_, _), rows in grouped.items():
        teams: List[str] = []
        aggregate_row: Optional[Dict[str, object]] = None
        aggregate_minutes = -1.0
        best_row: Optional[Dict[str, object]] = None
        best_minutes = -1.0

        for row in rows:
            team = clean_text(row.get("Team", ""))
            minutes = parse_metric_float(row.get("MP", "")) or 0.0

            if team and team not in AGGREGATE_TEAM_VALUES and team not in teams:
                teams.append(team)

            if team in AGGREGATE_TEAM_VALUES and minutes >= aggregate_minutes:
                aggregate_row = row
                aggregate_minutes = minutes

            if minutes >= best_minutes:
                best_row = row
                best_minutes = minutes

        chosen = aggregate_row or best_row
        if chosen is None:
            continue

        team_value = "/".join(teams) if teams else clean_text(chosen.get("Team", ""))
        standardized.append(
            {
                "Season": canonical_season(chosen.get("Season", "")),
                "PLAYER_ID": "",
                "PLAYER_NAME": clean_text(chosen.get("Player", "")),
                "TEAM_ABBREVIATION": team_value,
                "BPM": parse_metric_float(chosen.get("BPM", "")),
                "VORP": parse_metric_float(chosen.get("VORP", "")),
                "WS": parse_metric_float(chosen.get("WS", "")),
                "WS/48": parse_metric_float(chosen.get("WS/48", "")),
                "OWS": parse_metric_float(chosen.get("OWS", "")),
                "DWS": parse_metric_float(chosen.get("DWS", "")),
                "OBPM": parse_metric_float(chosen.get("OBPM", "")),
                "DBPM": parse_metric_float(chosen.get("DBPM", "")),
            }
        )

    return standardized


def load_epm_rows(path: Path) -> List[Dict[str, object]]:
    raw_rows = load_excel_rows(path)
    grouped: Dict[Tuple[str, str], Dict[str, object]] = {}

    for row in raw_rows:
        season = canonical_season(row.get("SEASON", ""))
        player = clean_text(row.get("NAME", ""))
        if not season or not player:
            continue

        key = (season, normalize_name(player))
        mpg = parse_metric_float(row.get("MPG", "")) or 0.0
        previous = grouped.get(key)
        previous_mpg = parse_metric_float(previous.get("MPG", "")) if previous else None
        if previous is None or mpg >= (previous_mpg or 0.0):
            grouped[key] = row

    standardized: List[Dict[str, object]] = []
    for row in grouped.values():
        standardized.append(
            {
                "Season": canonical_season(row.get("SEASON", "")),
                "PLAYER_ID": "",
                "PLAYER_NAME": clean_text(row.get("NAME", "")),
                "TEAM_ABBREVIATION": "",
                "OFF": parse_metric_float(row.get("OFF", "")),
                "DEF": parse_metric_float(row.get("DEF", "")),
                "EPM": parse_metric_float(row.get("EPM", "")),
            }
        )

    return standardized


def load_mamba_rows(path: Path) -> List[Dict[str, object]]:
    raw_rows = read_history_csv(path)
    grouped: Dict[Tuple[str, str], Dict[str, object]] = {}

    for row in raw_rows:
        season = canonical_season(row.get("year", ""))
        player = clean_text(row.get("player_name", ""))
        if not season or not player:
            continue

        key = (season, normalize_name(player))
        minutes = parse_metric_float(row.get("Minutes", "")) or 0.0
        previous = grouped.get(key)
        previous_minutes = parse_metric_float(previous.get("Minutes", "")) if previous else None
        if previous is None or minutes >= (previous_minutes or 0.0):
            grouped[key] = row

    standardized: List[Dict[str, object]] = []
    for row in grouped.values():
        standardized.append(
            {
                "Season": canonical_season(row.get("year", "")),
                "PLAYER_ID": canonical_id(row.get("nba_id", "")),
                "PLAYER_NAME": clean_text(row.get("player_name", "")),
                "TEAM_ABBREVIATION": "",
                "MAMBA": parse_metric_float(row.get("MAMBA", "")),
                "O-MAMBA": parse_metric_float(row.get("O-MAMBA", "")),
                "D-MAMBA": parse_metric_float(row.get("D-MAMBA", "")),
            }
        )

    return standardized


def load_lebron_rows(path: Path) -> List[Dict[str, object]]:
    raw_rows = read_history_csv(path)
    standardized: List[Dict[str, object]] = []

    for row in raw_rows:
        season = canonical_season(row.get("Season", "")) or canonical_season(row.get("year", ""))
        player = clean_text(row.get("Player", "")) or clean_text(row.get("player_name", ""))
        if not season or not player:
            continue

        standardized.append(
            {
                "Season": season,
                "PLAYER_ID": canonical_id(row.get("NBA_ID", row.get("nba_id", ""))),
                "PLAYER_NAME": player,
                "TEAM_ABBREVIATION": "",
                "LEBRON": parse_metric_float(row.get("LEBRON", "")),
                "O-LEBRON": parse_metric_float(row.get("O-LEBRON", "")),
                "D-LEBRON": parse_metric_float(row.get("D-LEBRON", "")),
            }
        )

    return standardized


def load_bball_index_impact_rows(path: Path) -> List[Dict[str, object]]:
    raw_rows = standardize_rows(
        read_history_csv(path),
        season_column="Season",
        player_column="Player",
    )
    standardized: List[Dict[str, object]] = []

    for row in raw_rows:
        season = clean_text(row.get("Season", ""))
        player = clean_text(row.get("PLAYER_NAME", ""))
        if not season or not player:
            continue

        standardized.append(
            {
                "Season": season,
                "PLAYER_ID": canonical_id(row.get("NBA_ID", row.get("NBA ID", row.get("PLAYER_ID", "")))),
                "PLAYER_NAME": player,
                "TEAM_ABBREVIATION": clean_text(row.get("Team(s)", row.get("Team", ""))),
                "LEBRON": parse_metric_float(row.get("LEBRON", "")),
                "D-LEBRON": parse_metric_float(row.get("D-LEBRON", "")),
                "O-LEBRON": parse_metric_float(row.get("O-LEBRON", "")),
                "LEBRON WAR": parse_metric_float(row.get("LEBRON WAR", "")),
                "LEBRON Box Impact": parse_metric_float(row.get("LEBRON Box Impact", "")),
                "LEBRON Vs Role Average": parse_metric_float(row.get("LEBRON Vs Role Average", "")),
                "Multi-Year LEBRON": parse_metric_float(row.get("Multi-Year LEBRON", "")),
                "O-LEBRON Box Impact": parse_metric_float(row.get("O-LEBRON Box Impact", "")),
                "LEBRON Offensive Points Added": parse_metric_float(row.get("LEBRON Offensive Points Added", "")),
                "O-LEBRON Vs Role Average": parse_metric_float(row.get("O-LEBRON Vs Role Average", "")),
                "Multi-Year O-LEBRON": parse_metric_float(row.get("Multi-Year O-LEBRON", "")),
                "Predictive O-LEBRON": parse_metric_float(row.get("Predictive O-LEBRON", "")),
                "D-LEBRON Box Impact": parse_metric_float(row.get("D-LEBRON Box Impact", "")),
                "LEBRON Defensive Points Saved": parse_metric_float(row.get("LEBRON Defensive Points Saved", "")),
                "D-LEBRON Vs Role Average": parse_metric_float(row.get("D-LEBRON Vs Role Average", "")),
                "Multi-Year D-LEBRON": parse_metric_float(row.get("Multi-Year D-LEBRON", "")),
                "Predictive D-LEBRON": parse_metric_float(row.get("Predictive D-LEBRON", "")),
            }
        )

    return standardized


def overlay_rows(
    base_rows: Sequence[Dict[str, object]],
    override_rows: Sequence[Dict[str, object]],
) -> List[Dict[str, object]]:
    combined: Dict[Tuple[str, str], Dict[str, object]] = {}

    for row in base_rows:
        season = str(row.get("Season", "")).strip()
        player = normalize_name(row.get("PLAYER_NAME", ""))
        if season and player:
            combined[(season, player)] = dict(row)

    for row in override_rows:
        season = str(row.get("Season", "")).strip()
        player = normalize_name(row.get("PLAYER_NAME", ""))
        if season and player:
            combined[(season, player)] = dict(row)

    return list(combined.values())


def get_row_key(row: CalUniverseRow) -> Tuple[str, str]:
    return row.season, normalize_name(row.player)


def set_base_value(base_row: Dict[str, object], alias: str, value: object, matched_by: str) -> None:
    if value is None:
        return
    text = str(value).strip()
    if not text:
        return
    base_row[alias] = value
    base_row[f"__matched__{alias}"] = matched_by


def merge_metric_columns(
    universe: Sequence[CalUniverseRow],
    base_rows_by_key: Dict[Tuple[str, str], Dict[str, object]],
    source_rows: Sequence[Dict[str, object]],
    column_map: Dict[str, str],
    allow_id_fallback: bool,
) -> None:
    if not source_rows:
        return

    by_id, by_name = build_source_index(list(source_rows))
    for universe_row in universe:
        source_row, matched_by = match_metric_row(
            universe_row,
            by_id,
            by_name,
            allow_id_fallback=allow_id_fallback,
        )
        if source_row is None:
            continue

        base_row = base_rows_by_key[get_row_key(universe_row)]
        if not str(base_row.get("TEAM_ABBREVIATION", "")).strip():
            team_value = str(source_row.get("TEAM_ABBREVIATION", "")).strip()
            if team_value:
                base_row["TEAM_ABBREVIATION"] = team_value

        for alias, source_column in column_map.items():
            set_base_value(base_row, alias, source_row.get(source_column, None), matched_by)


def derive_impact_metrics(base_rows: Iterable[Dict[str, object]]) -> None:
    for base_row in base_rows:
        defensive_rating = parse_metric_float(base_row.get("DEF_RATING", ""))
        if defensive_rating is not None:
            base_row["INV_Stable On-Court DRtg"] = -defensive_rating
            base_row["__matched__INV_Stable On-Court DRtg"] = base_row.get(
                "__matched__DEF_RATING",
                "",
            )

        defensive_rim_fg_pct = parse_metric_float(base_row.get("DEF_RIM_FG_PCT", ""))
        if defensive_rim_fg_pct is not None:
            base_row["Defensive eFG% Impact"] = -defensive_rim_fg_pct
            base_row["__matched__Defensive eFG% Impact"] = base_row.get(
                "__matched__DEF_RIM_FG_PCT",
                "",
            )


def build_metric_result(
    contexts,
    base_rows_by_key: Dict[Tuple[str, str], Dict[str, object]],
    metric_alias: str,
    current_season: str,
    current_season_min_threshold: float,
    standard_min_threshold: float,
) -> MetricResult:
    raw_values: List[Optional[float]] = []
    normalized_values: List[Optional[float]] = []
    matched_by_values: List[str] = []
    matched_numeric_values: List[float] = []

    for context in contexts:
        base_row = base_rows_by_key[get_row_key(context.universe_row)]
        raw_value = parse_float(base_row.get(metric_alias, ""))
        matched_by = str(base_row.get(f"__matched__{metric_alias}", "")).strip()
        raw_values.append(raw_value)
        matched_by_values.append(matched_by)
        if raw_value is not None:
            matched_numeric_values.append(raw_value)

    mean_value = None if not matched_numeric_values else statistics.mean(matched_numeric_values)
    stdev_value = (
        statistics.stdev(matched_numeric_values)
        if len(matched_numeric_values) >= 2
        else None
    )

    for index, context in enumerate(contexts):
        raw_value = raw_values[index]
        if raw_value is None or mean_value is None or stdev_value in (None, 0):
            normalized_values.append(None)
            continue

        raw_z = compute_capped_z_score(raw_value, mean_value, stdev_value)
        normalized_values.append(
            apply_cal_normalization(
                season=context.universe_row.season,
                rotation_role=context.universe_row.rotation_role,
                minutes=context.effective_minutes,
                raw_z=raw_z,
                current_season=current_season,
                current_season_min_threshold=current_season_min_threshold,
                standard_min_threshold=standard_min_threshold,
            )
        )

    return MetricResult(
        raw_values=raw_values,
        normalized_values=normalized_values,
        matched_by=matched_by_values,
        mean_value=mean_value,
        stdev_value=stdev_value,
        source_note=METRIC_SOURCE_NOTES.get(metric_alias, "no mapped local source yet"),
    )


def average_numeric(values: Sequence[Optional[float]]) -> Optional[float]:
    numeric = [value for value in values if value is not None]
    if not numeric:
        return None
    return statistics.mean(numeric)


def compute_section_score(
    section: SectionSpec,
    component_values_by_alias: Dict[str, Optional[float]],
) -> float:
    weighted_values: List[float] = []
    total_weight = 0.0
    for aliases, weight in section.weighted_groups:
        group_average = average_numeric([component_values_by_alias.get(alias) for alias in aliases])
        if group_average is None:
            if section.fallback_aliases:
                fallback_value = average_numeric(
                    [component_values_by_alias.get(alias) for alias in section.fallback_aliases]
                )
                if fallback_value is not None:
                    return fallback_value
            continue
        weighted_values.append(group_average * weight)
        total_weight += weight

    if not weighted_values or total_weight <= 0:
        return -1.0
    return sum(weighted_values) / total_weight


def compute_piecewise_rating(
    value: float,
    population: Sequence[float],
    curve: Sequence[Tuple[float, float]],
) -> float:
    minimum = min(population)
    low_x = minimum
    low_y = 25.0

    for percentile, high_y in curve:
        high_x = percentile_inc(population, percentile)
        if value <= high_x or percentile >= 1.0:
            return interpolate_rating(value, low_x, high_x, low_y, high_y)
        low_x = high_x
        low_y = high_y

    return curve[-1][1]


def build_section_outputs(
    section: SectionSpec,
    contexts,
    metric_results: Dict[str, MetricResult],
) -> Tuple[List[float], List[float], List[float], List[List[Optional[float]]]]:
    scores: List[float] = []
    component_rows: List[List[Optional[float]]] = []

    for index in range(len(contexts)):
        component_values_by_alias: Dict[str, Optional[float]] = {}
        component_row: List[Optional[float]] = []
        for slot in section.component_slots:
            value = metric_results[slot.metric_alias].normalized_values[index]
            component_values_by_alias[slot.metric_alias] = value
            component_row.append(value)

        scores.append(compute_section_score(section, component_values_by_alias))
        component_rows.append(component_row)

    if len(scores) < 2:
        raise SystemExit(f"Not enough {section.name} scores to compute aggregate z-scores.")

    median_value = statistics.median(scores)
    stdev_value = statistics.stdev(scores)
    if stdev_value == 0:
        raise SystemExit(f"{section.name} scores have zero variance; cannot compute ratings.")

    aggregate_z_scores = [(value - median_value) / stdev_value for value in scores]
    ratings = [
        compute_piecewise_rating(value, aggregate_z_scores, section.curve)
        for value in aggregate_z_scores
    ]
    return scores, aggregate_z_scores, ratings, component_rows


def main() -> None:
    args = parse_args()

    workbook_path = Path(args.workbook)
    universe_path = Path(args.universe_csv)
    details_path = resolve_details_csv_path(args.details_csv, universe_path)
    minutes_source_path = Path(args.minutes_source)
    bballref_source_path = Path(args.bballref_source)
    epm_source_path = Path(args.epm_source)
    mamba_source_path = Path(args.mamba_source)
    lebron_source_path = Path(args.lebron_source)
    bball_index_impact_source_path = Path(args.bball_index_impact_source)
    advanced_source_path = Path(args.advanced_source)
    dunks_epm_source_path = Path(args.dunks_epm_source)
    tracking_defensive_impact_source_path = Path(args.tracking_defensive_impact_source)

    required_paths = (
        minutes_source_path,
        bballref_source_path,
        epm_source_path,
        mamba_source_path,
        advanced_source_path,
        tracking_defensive_impact_source_path,
    )
    for path in required_paths:
        if not path.exists():
            raise SystemExit(f"Source file not found: {path}")
    if details_path and not details_path.exists():
        raise SystemExit(f"Details CSV not found: {details_path}")

    bballref_rows = load_bballref_advanced_rows(bballref_source_path)
    epm_rows = load_epm_rows(epm_source_path)
    mamba_rows = load_mamba_rows(mamba_source_path)
    lebron_rows = load_lebron_rows(lebron_source_path) if lebron_source_path.exists() else []
    bball_index_impact_rows = (
        load_bball_index_impact_rows(bball_index_impact_source_path)
        if bball_index_impact_source_path.exists()
        else []
    )
    if dunks_epm_source_path.exists():
        epm_rows = overlay_rows(epm_rows, read_history_csv(dunks_epm_source_path))
    advanced_rows = read_history_csv(advanced_source_path)
    tracking_defensive_impact_rows = standardize_rows(
        read_history_csv(tracking_defensive_impact_source_path),
        season_column="Season",
        player_column="PLAYER_NAME",
    )
    minutes_rows = read_history_csv(minutes_source_path)

    workbook_universe: List[CalUniverseRow] = []
    if workbook_path.exists():
        workbook_universe = load_cal_universe(workbook_path, sheet_name=args.sheet)

    if universe_path.exists():
        universe = load_universe_csv(universe_path)
    elif workbook_universe:
        universe = workbook_universe
    else:
        raise SystemExit(
            f"Universe CSV not found: {universe_path} and workbook not found: {workbook_path}"
        )

    if details_path:
        universe = enrich_universe_rows(universe, load_universe_csv(details_path))
    if workbook_universe:
        universe = enrich_universe_rows(universe, workbook_universe)

    current_season = detect_current_season(args.current_season, universe, minutes_rows)
    contexts = build_player_contexts(
        universe=universe,
        minutes_rows=minutes_rows,
        minutes_column=args.minutes_column,
        minutes_games_column=args.minutes_games_column,
        allow_id_fallback=args.allow_id_fallback,
    )

    base_rows_by_key: Dict[Tuple[str, str], Dict[str, object]] = {}
    for universe_row in universe:
        base_rows_by_key[get_row_key(universe_row)] = {
            "Season": universe_row.season,
            "PLAYER_NAME": universe_row.player,
            "PLAYER_ID": universe_row.nba_id,
            "NBA_ID": universe_row.nba_id,
            "TEAM_ABBREVIATION": "",
        }

    merge_metric_columns(
        universe,
        base_rows_by_key,
        bballref_rows,
        {
            "BPM": "BPM",
            "VORP": "VORP",
            "WS": "WS",
            "WS/48": "WS/48",
            "OWS": "OWS",
            "DWS": "DWS",
            "O-BPM": "OBPM",
            "D-BPM": "DBPM",
        },
        allow_id_fallback=args.allow_id_fallback,
    )
    merge_metric_columns(
        universe,
        base_rows_by_key,
        epm_rows,
        {
            "EPM": "EPM",
            "O-EPM": "OFF",
            "D-EPM": "DEF",
        },
        allow_id_fallback=args.allow_id_fallback,
    )
    merge_metric_columns(
        universe,
        base_rows_by_key,
        mamba_rows,
        {
            "MAMBA": "MAMBA",
            "O-MAMBA": "O-MAMBA",
            "D-MAMBA": "D-MAMBA",
        },
        allow_id_fallback=args.allow_id_fallback,
    )
    merge_metric_columns(
        universe,
        base_rows_by_key,
        lebron_rows,
        {
            "LEBRON": "LEBRON",
            "O-LEBRON": "O-LEBRON",
            "D-LEBRON": "D-LEBRON",
        },
        allow_id_fallback=args.allow_id_fallback,
    )
    merge_metric_columns(
        universe,
        base_rows_by_key,
        bball_index_impact_rows,
        {
            "LEBRON": "LEBRON",
            "D-LEBRON": "D-LEBRON",
            "O-LEBRON": "O-LEBRON",
            "LEBRON WAR": "LEBRON WAR",
            "LEBRON Box Impact": "LEBRON Box Impact",
            "LEBRON Vs Role Average": "LEBRON Vs Role Average",
            "Multi-Year LEBRON": "Multi-Year LEBRON",
            "O-LEBRON Box Impact": "O-LEBRON Box Impact",
            "LEBRON Offensive Points Added": "LEBRON Offensive Points Added",
            "O-LEBRON Vs Role Average": "O-LEBRON Vs Role Average",
            "Multi-Year O-LEBRON": "Multi-Year O-LEBRON",
            "Predictive O-LEBRON": "Predictive O-LEBRON",
            "D-LEBRON Box Impact": "D-LEBRON Box Impact",
            "LEBRON Defensive Points Saved": "LEBRON Defensive Points Saved",
            "D-LEBRON Vs Role Average": "D-LEBRON Vs Role Average",
            "Multi-Year D-LEBRON": "Multi-Year D-LEBRON",
            "Predictive D-LEBRON": "Predictive D-LEBRON",
        },
        allow_id_fallback=args.allow_id_fallback,
    )
    merge_metric_columns(
        universe,
        base_rows_by_key,
        advanced_rows,
        {
            "Stable On-Court Net Rating": "NET_RATING",
            "Stable On-Court ORtg": "OFF_RATING",
            "DEF_RATING": "DEF_RATING",
        },
        allow_id_fallback=args.allow_id_fallback,
    )
    merge_metric_columns(
        universe,
        base_rows_by_key,
        tracking_defensive_impact_rows,
        {"DEF_RIM_FG_PCT": "DEF_RIM_FG_PCT"},
        allow_id_fallback=args.allow_id_fallback,
    )

    derive_impact_metrics(base_rows_by_key.values())

    all_metric_aliases = {
        slot.metric_alias
        for section in SECTIONS
        for slot in section.component_slots
    }
    metric_results: Dict[str, MetricResult] = {}
    for metric_alias in sorted(all_metric_aliases):
        metric_results[metric_alias] = build_metric_result(
            contexts=contexts,
            base_rows_by_key=base_rows_by_key,
            metric_alias=metric_alias,
            current_season=current_season,
            current_season_min_threshold=args.current_season_min_threshold,
            standard_min_threshold=args.standard_min_threshold,
        )

    overall_scores, overall_z, overall_ratings, overall_rows = build_section_outputs(
        section=OVERALL_SECTION,
        contexts=contexts,
        metric_results=metric_results,
    )
    offensive_scores, offensive_z, offensive_ratings, offensive_rows = build_section_outputs(
        section=OFFENSIVE_SECTION,
        contexts=contexts,
        metric_results=metric_results,
    )
    defensive_scores, defensive_z, defensive_ratings, defensive_rows = build_section_outputs(
        section=DEFENSIVE_SECTION,
        contexts=contexts,
        metric_results=metric_results,
    )

    sheet_rows: List[List[object]] = []
    rating_only_rows: List[Dict[str, object]] = []
    audit_rows: List[Dict[str, object]] = []
    unmatched_rows: List[Dict[str, object]] = []

    for index, context in enumerate(contexts):
        base_row = base_rows_by_key[get_row_key(context.universe_row)]
        team_value = str(base_row.get("TEAM_ABBREVIATION", "")).strip()

        sheet_rows.append(
            [
                context.universe_row.nba_id,
                context.universe_row.season,
                context.universe_row.player,
                team_value,
                overall_ratings[index],
                overall_z[index],
                overall_scores[index],
                *overall_rows[index],
                offensive_ratings[index],
                offensive_z[index],
                offensive_scores[index],
                *offensive_rows[index],
                defensive_ratings[index],
                defensive_z[index],
                defensive_scores[index],
                *defensive_rows[index],
            ]
        )

        rating_only_rows.append(
            {
                "NBA_ID": context.universe_row.nba_id,
                "Season": context.universe_row.season,
                "Player": context.universe_row.player,
                "Team(s)": team_value,
                "Overall Impact Rating": overall_ratings[index],
                "Offensive Impact Rating": offensive_ratings[index],
                "Defensive Impact Rating": defensive_ratings[index],
            }
        )

        audit_row: Dict[str, object] = {
            "NBA_ID": context.universe_row.nba_id,
            "Season": context.universe_row.season,
            "Player": context.universe_row.player,
            "Team(s)": team_value,
            "RotationRole": context.universe_row.rotation_role,
            "MIN": context.effective_minutes,
            "WorkbookMIN": context.workbook_minutes,
            "LiveMIN": context.live_minutes,
            "LiveMINPerGame": context.live_minutes_per_game,
            "LiveGP": context.live_gp,
            "MinutesMatchedBy": context.minutes_matched_by,
            "CurrentSeason": current_season,
            "Overall Impact Score": overall_scores[index],
            "Overall Impact AggregateZ": overall_z[index],
            "Overall Impact Rating": overall_ratings[index],
            "Offensive Impact Score": offensive_scores[index],
            "Offensive Impact AggregateZ": offensive_z[index],
            "Offensive Impact Rating": offensive_ratings[index],
            "Defensive Impact Score": defensive_scores[index],
            "Defensive Impact AggregateZ": defensive_z[index],
            "Defensive Impact Rating": defensive_ratings[index],
        }

        missing_labels: List[str] = []
        for section in SECTIONS:
            missing_aliases = sorted(
                {
                    slot.metric_alias
                    for slot in section.component_slots
                    if metric_results[slot.metric_alias].normalized_values[index] is None
                }
            )
            if missing_aliases:
                missing_labels.extend(f"{section.name}::{alias}" for alias in missing_aliases)
            audit_row[f"{section.name} MissingCount"] = len(missing_aliases)
            audit_row[f"{section.name} MissingMetrics"] = " | ".join(missing_aliases)

        for metric_alias in sorted(metric_results):
            result = metric_results[metric_alias]
            audit_row[f"{metric_alias} Raw"] = result.raw_values[index]
            audit_row[f"{metric_alias} Z"] = result.normalized_values[index]
            audit_row[f"{metric_alias} MatchedBy"] = result.matched_by[index]
            audit_row[f"{metric_alias} Mean"] = result.mean_value
            audit_row[f"{metric_alias} Stdev"] = result.stdev_value
            audit_row[f"{metric_alias} Source"] = result.source_note

        audit_rows.append(audit_row)

        if missing_labels:
            unmatched_rows.append(
                {
                    "NBA_ID": context.universe_row.nba_id,
                    "Season": context.universe_row.season,
                    "Player": context.universe_row.player,
                    "Team(s)": team_value,
                    "RotationRole": context.universe_row.rotation_role,
                    "MIN": context.effective_minutes,
                    "MissingCount": len(missing_labels),
                    "MissingMetrics": " | ".join(missing_labels),
                }
            )

    output_prefix = args.output_prefix.strip() or "impact_all"
    sheet_path = EXPORT_DIR / f"{output_prefix}_sheet.csv"
    rating_only_path = EXPORT_DIR / f"{output_prefix}_ratings.csv"
    audit_path = EXPORT_DIR / f"{output_prefix}_audit.csv"
    unmatched_path = EXPORT_DIR / f"{output_prefix}_unmatched.csv"

    write_matrix_csv(sheet_path, WORKBOOK_COLUMNS, sheet_rows)
    write_csv(rating_only_path, RATING_ONLY_HEADERS, rating_only_rows)
    write_csv(audit_path, list(audit_rows[0].keys()) if audit_rows else [], audit_rows)
    write_csv(
        unmatched_path,
        ["NBA_ID", "Season", "Player", "Team(s)", "RotationRole", "MIN", "MissingCount", "MissingMetrics"],
        unmatched_rows,
    )

    print(f"[OK] Built Impact export for {len(sheet_rows)} player-season rows")
    print(
        "[INFO] Exact local sources: bballref_advanced.xlsx for WS/BPM/VORP/OWS/DWS/OBPM/DBPM, "
        "epm sources for EPM/O-EPM/D-EPM, mamba.csv for MAMBA, and bball_index_impact.csv "
        "as the primary LEBRON-family source with lebron.csv fallback for uncovered rows."
    )
    print(
        "[INFO] Proxies in use: NET_RATING/OFF_RATING/inverse DEF_RATING for on-court impact slots, "
        "and inverse DEF_RIM_FG_PCT for Defensive eFG% Impact."
    )
    print(
        "[INFO] Remaining RAPTOR, RPM, RAPM, DPM, DRIP, and SPI-family workbook slots are currently left unmatched "
        "so the audit shows the real coverage gap."
    )
    print(f"[OUT] Sheet -> {sheet_path}")
    print(f"[OUT] Ratings -> {rating_only_path}")
    print(f"[OUT] Audit -> {audit_path}")
    print(f"[OUT] Unmatched -> {unmatched_path}")


if __name__ == "__main__":
    main()
