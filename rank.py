import json
import re
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from datetime import datetime

# Paths for input and output files
GAME_STAT = True
input_file = './rank/rank.txt'
excel_file = '14pt'

if GAME_STAT:
    excel_file = f'./rank/game_{excel_file}.xlsx'
else:
    excel_file = f'./rank/{excel_file}.xlsx'

# Load previous Excel data into a dictionary
existing_data = defaultdict(lambda: {"rank": 0, "votes": 0, "entries": 0, "rankMomentum": 0, "votesMomentum": 0})

# Function to extract vote count from "subLabel"
def extract_votes(sub_label):
    match = re.search(r"(\d+) vote", sub_label)
    return int(match.group(1)) if match else 0

try:
    # Load the existing workbook and sheet, or create a new workbook if it doesn't exist
    workbook = load_workbook(excel_file)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        display_name = row[0]
        try:
            existing_data[display_name] = {
                "rank": float(row[1]),
                "votes": float(row[2]),
                "entries": int(row[3]),
                "rankMomentum": float(row[4]),
                "votesMomentum": float(row[5])
            }
        except:
            existing_data[display_name] = {
                "rank": float(row[1]),
                "votes": float(row[2]),
                "entries": int(row[3]),
                "rankMomentum": 0,
                "votesMomentum": 0
            }
            print("No momentum yet")
except FileNotFoundError:
    # If the Excel file doesn't exist yet, create a new workbook
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["displayName", "rank", "votes", "entries", "rankMomentum", "votesMomentum"])  # Column headers

# Read JSON data from the .txt file
with open(input_file, 'r', encoding='utf-8') as file:
    data = json.load(file)

# Update data with new JSON entries and momentum calculation
for result in data["results"]:
    display_name = result["label"]
    new_rank = int(result["rank"])
    new_votes = extract_votes(result["subLabel"])

    if display_name in existing_data:
        existing_entry = existing_data[display_name]
        rank_change = new_rank - existing_entry["rank"]
        votes_change = new_votes - existing_entry["votes"]

        # Update with new values and momentum
        existing_entry["rank"] = new_rank * 0.3 + existing_entry["rank"] * 0.7
        existing_entry["votes"] = new_votes * 0.3 + existing_entry["votes"] * 0.7
        existing_entry["entries"] += 1
        existing_entry["rankMomentum"] = rank_change
        existing_entry["votesMomentum"] = votes_change
    else:
        # New entry with no previous data
        existing_data[display_name] = {
            "rank": new_rank,
            "votes": new_votes,
            "entries": 1,
            "rankMomentum": 0,
            "votesMomentum": 0
        }

# Write updated data back to Excel file
sheet.delete_rows(2, sheet.max_row)  # Clear existing data
for display_name, stats in existing_data.items():
    sheet.append([
        display_name,
        stats["rank"],
        stats["votes"],
        stats["entries"],
        stats["rankMomentum"],
        stats["votesMomentum"]
    ])

# Save the workbook
workbook.save(excel_file)
print(f"Data has been updated with momentum in {excel_file}")
