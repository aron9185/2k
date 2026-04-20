from __future__ import annotations

import argparse
import csv
import statistics
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from build_cal_lane import (
    CalUniverseRow,
    apply_cal_normalization,
    build_source_index,
    canonical_id,
    compute_capped_z_score,
    detect_current_season,
    enrich_universe_rows,
    load_cal_universe,
    load_universe_csv,
    match_metric_row,
    normalize_name,
    parse_float,
    read_history_csv,
    resolve_details_csv_path,
    write_csv,
)
from build_finishing_layup import build_player_contexts
from build_finishing_standing_dunk import write_matrix_csv
from pull_nba_stats import EXPORT_DIR, HISTORY_DIR, MANUAL_DIR


OPTIONAL_BASE_COLUMNS = {"AV"}


@dataclass(frozen=True)
class SourceFileSpec:
    key: str
    path: Path
    season_column: str
    player_column: str
    id_column: str = ""


@dataclass(frozen=True)
class ColumnSourceSpec:
    columns: Sequence[str]
    source_key: str
    source_column: str
    source_note: str
    default_zero_when_missing: bool = False
    default_zero_requires_season_nonzero_coverage: bool = False


@dataclass
class MetricResult:
    raw_values: List[Optional[float]]
    normalized_values: List[Optional[float]]
    matched_by: List[str]
    mean_value: Optional[float]
    stdev_value: Optional[float]
    source_note: str


SOURCE_FILES: Sequence[SourceFileSpec] = (
    SourceFileSpec(
        "finishing_all_ratings",
        EXPORT_DIR / "finishing_all_ratings.csv",
        "Season",
        "Player",
        "NBA_ID",
    ),
    SourceFileSpec(
        "shooting_all_ratings",
        EXPORT_DIR / "shooting_all_ratings.csv",
        "Season",
        "Player",
        "NBA_ID",
    ),
    SourceFileSpec(
        "playmaking_all_ratings",
        EXPORT_DIR / "playmaking_all_ratings.csv",
        "Season",
        "Player",
        "NBA_ID",
    ),
    SourceFileSpec(
        "defense_all_ratings",
        EXPORT_DIR / "defense_all_ratings.csv",
        "Season",
        "Player",
        "NBA_ID",
    ),
    SourceFileSpec(
        "rebounding_all_ratings",
        EXPORT_DIR / "rebounding_all_ratings.csv",
        "Season",
        "Player",
        "NBA_ID",
    ),
    SourceFileSpec(
        "physical_all_ratings",
        EXPORT_DIR / "physical_all_ratings.csv",
        "Season",
        "Player",
        "NBA_ID",
    ),
    SourceFileSpec(
        "finishing_standing_dunk_sheet",
        EXPORT_DIR / "finishing_standing_dunk_sheet.csv",
        "Season",
        "Player",
        "NBA ID",
    ),
    SourceFileSpec(
        "shooting_three_point_sheet",
        EXPORT_DIR / "shooting_three_point_sheet.csv",
        "Season",
        "Player",
        "NBA ID",
    ),
    SourceFileSpec(
        "shooting_mid_range_sheet",
        EXPORT_DIR / "shooting_mid_range_sheet.csv",
        "Season",
        "Player",
        "NBA ID",
    ),
    SourceFileSpec(
        "shooting_close_shot_sheet",
        EXPORT_DIR / "shooting_close_shot_sheet.csv",
        "Season",
        "Player",
        "NBA ID",
    ),
    SourceFileSpec(
        "playmaking_all_sheet",
        EXPORT_DIR / "playmaking_all_sheet.csv",
        "Season",
        "Player",
        "NBA ID",
    ),
    SourceFileSpec(
        "defense_all_sheet",
        EXPORT_DIR / "defense_all_sheet.csv",
        "Season",
        "Player",
        "NBA ID",
    ),
    SourceFileSpec(
        "rebounding_all_sheet",
        EXPORT_DIR / "rebounding_all_sheet.csv",
        "Season",
        "Player",
        "NBA ID",
    ),
    SourceFileSpec(
        "impact_all_sheet",
        EXPORT_DIR / "impact_all_sheet.csv",
        "Season",
        "Player",
        "NBA ID",
    ),
    SourceFileSpec(
        "physical_all_sheet",
        EXPORT_DIR / "physical_all_sheet.csv",
        "Season",
        "Player",
        "NBA ID",
    ),
    SourceFileSpec(
        "general_traditional",
        HISTORY_DIR / "general_traditional.csv",
        "Season",
        "PLAYER_NAME",
        "PLAYER_ID",
    ),
    SourceFileSpec(
        "closest_defender",
        HISTORY_DIR / "closest_defender.csv",
        "Season",
        "PLAYER_NAME",
        "PLAYER_ID",
    ),
    SourceFileSpec(
        "very_tight",
        HISTORY_DIR / "very_tight.csv",
        "Season",
        "PLAYER_NAME",
        "PLAYER_ID",
    ),
    SourceFileSpec(
        "tracking_c&s",
        HISTORY_DIR / "tracking_c&s.csv",
        "Season",
        "PLAYER_NAME",
        "PLAYER_ID",
    ),
    SourceFileSpec(
        "pullup",
        HISTORY_DIR / "pullup.csv",
        "Season",
        "PLAYER_NAME",
        "PLAYER_ID",
    ),
    SourceFileSpec(
        "spotup",
        HISTORY_DIR / "spotup.csv",
        "Season",
        "PLAYER_NAME",
        "PLAYER_ID",
    ),
    SourceFileSpec(
        "shot_locations_by_distance_8ft",
        HISTORY_DIR / "shot_locations_by_distance_8ft.csv",
        "Season",
        "PLAYER_NAME",
        "PLAYER_ID",
    ),
    SourceFileSpec(
        "nbarapm",
        HISTORY_DIR / "nbarapm.csv",
        "Season",
        "Player",
        "NBA_ID",
    ),
    SourceFileSpec(
        "bball_index_badges",
        HISTORY_DIR / "bball_index_badges.csv",
        "Season",
        "Player",
        "",
    ),
    SourceFileSpec(
        "bios",
        HISTORY_DIR / "bios.csv",
        "Season",
        "PLAYER_NAME",
        "PLAYER_ID",
    ),
)


RATING_SPECS: Sequence[ColumnSourceSpec] = (
    ColumnSourceSpec(
        ("M", "T", "AD", "AO", "CB"),
        "shooting_all_ratings",
        "Mid-Range Shot Rating",
        "shooting_all_ratings.csv -> Mid-Range Shot Rating",
    ),
    ColumnSourceSpec(
        ("N", "R", "U", "AE", "AP"),
        "shooting_all_ratings",
        "3-Point Shot Rating",
        "shooting_all_ratings.csv -> 3-Point Shot Rating",
    ),
    ColumnSourceSpec(
        ("AQ", "CL", "DH", "DR", "DU"),
        "playmaking_all_ratings",
        "Ball Handle Rating",
        "playmaking_all_ratings.csv -> Ball Handle Rating",
    ),
    ColumnSourceSpec(
        ("CN", "CU", "DA", "DZ"),
        "playmaking_all_ratings",
        "Pass Accuracy Rating",
        "playmaking_all_ratings.csv -> Pass Accuracy Rating",
    ),
    ColumnSourceSpec(
        ("AY", "BI", "BP", "BT"),
        "finishing_all_ratings",
        "Layup Rating",
        "finishing_all_ratings.csv -> Layup Rating",
    ),
    ColumnSourceSpec(
        ("AZ", "BW", "CI"),
        "finishing_all_ratings",
        "Standing Dunk Rating",
        "finishing_all_ratings.csv -> Standing Dunk Rating",
    ),
    ColumnSourceSpec(
        ("BA", "BX"),
        "finishing_all_ratings",
        "Driving Dunk Rating",
        "finishing_all_ratings.csv -> Driving Dunk Rating",
    ),
    ColumnSourceSpec(
        ("BB", "BY", "CJ", "EQ", "GW"),
        "physical_all_ratings",
        "Vertical Rating",
        "physical_all_ratings.csv -> Vertical Rating",
    ),
    ColumnSourceSpec(
        ("BH", "BM", "BR"),
        "shooting_all_ratings",
        "Close Shot Rating",
        "shooting_all_ratings.csv -> Close Shot Rating",
    ),
    ColumnSourceSpec(
        ("BL",),
        "finishing_all_ratings",
        "Post Hook Rating",
        "finishing_all_ratings.csv -> Post Hook Rating",
    ),
    ColumnSourceSpec(
        ("BN", "CD", "CG", "DV"),
        "finishing_all_ratings",
        "Post Control Rating",
        "finishing_all_ratings.csv -> Post Control Rating",
    ),
    ColumnSourceSpec(
        ("BU", "CE", "DS", "EX", "GH", "GL", "GU"),
        "physical_all_ratings",
        "Strength Rating",
        "physical_all_ratings.csv -> Strength Rating",
    ),
    ColumnSourceSpec(
        ("CA",),
        "finishing_all_ratings",
        "Post Fade Rating",
        "finishing_all_ratings.csv -> Post Fade Rating",
    ),
    ColumnSourceSpec(
        ("DP",),
        "playmaking_all_ratings",
        "Speed with Ball Rating",
        "playmaking_all_ratings.csv -> Speed with Ball Rating",
    ),
    ColumnSourceSpec(
        ("EF", "EY", "FN", "FT", "GB"),
        "defense_all_ratings",
        "Perimeter Defense Rating",
        "defense_all_ratings.csv -> Perimeter Defense Rating",
    ),
    ColumnSourceSpec(
        ("EM", "FF"),
        "defense_all_ratings",
        "Steal Rating",
        "defense_all_ratings.csv -> Steal Rating",
    ),
    ColumnSourceSpec(
        ("EO", "FX", "GX"),
        "defense_all_ratings",
        "Block Rating",
        "defense_all_ratings.csv -> Block Rating",
    ),
    ColumnSourceSpec(
        ("EP", "HF"),
        "physical_all_ratings",
        "Speed Rating",
        "physical_all_ratings.csv -> Speed Rating",
    ),
    ColumnSourceSpec(
        ("FM", "FW", "GG"),
        "defense_all_ratings",
        "Interior Defense Rating",
        "defense_all_ratings.csv -> Interior Defense Rating",
    ),
    ColumnSourceSpec(
        ("FU", "GC", "HG"),
        "physical_all_ratings",
        "Agility Rating",
        "physical_all_ratings.csv -> Agility Rating",
    ),
    ColumnSourceSpec(
        ("GM", "GP", "GY"),
        "rebounding_all_ratings",
        "Offensive Rebound Rating",
        "rebounding_all_ratings.csv -> Offensive Rebound Rating",
    ),
    ColumnSourceSpec(
        ("GN", "GQ", "GZ"),
        "rebounding_all_ratings",
        "Defensive Rebound Rating",
        "rebounding_all_ratings.csv -> Defensive Rebound Rating",
    ),
)


PRECOMPUTED_SPECS: Sequence[ColumnSourceSpec] = (
    ColumnSourceSpec(
        ("K",),
        "shooting_three_point_sheet",
        "3PT Shot Making",
        "shooting_three_point_sheet.csv -> 3PT Shot Making",
    ),
    ColumnSourceSpec(
        ("L",),
        "shooting_mid_range_sheet",
        "Midrange Shot Making",
        "shooting_mid_range_sheet.csv -> Midrange Shot Making",
    ),
    ColumnSourceSpec(
        ("AI",),
        "shooting_three_point_sheet",
        "3PT Pull Up Talent",
        "shooting_three_point_sheet.csv -> 3PT Pull Up Talent",
    ),
    ColumnSourceSpec(
        ("AJ",),
        "shooting_mid_range_sheet",
        "Midrange Pull Up Talent",
        "shooting_mid_range_sheet.csv -> Midrange Pull Up Talent",
    ),
    ColumnSourceSpec(
        ("AW",),
        "finishing_standing_dunk_sheet",
        "Putback Scoring Impact Per 75 Possessions",
        "finishing_standing_dunk_sheet.csv -> Putback Scoring Impact Per 75 Possessions",
    ),
    ColumnSourceSpec(
        ("BF",),
        "shooting_close_shot_sheet",
        "Floating_shot_FGM",
        "shooting_close_shot_sheet.csv -> Floating_shot_FGM",
    ),
    ColumnSourceSpec(
        ("BG",),
        "shooting_close_shot_sheet",
        "Floater Talent",
        "shooting_close_shot_sheet.csv -> Floater Talent",
    ),
    ColumnSourceSpec(
        ("CR",),
        "rebounding_all_sheet",
        "Defensive Rebound",
        "rebounding_all_sheet.csv -> Defensive Rebound",
    ),
    ColumnSourceSpec(
        ("CS",),
        "playmaking_all_sheet",
        "Pass Vision",
        "playmaking_all_sheet.csv -> Pass Vision",
    ),
    ColumnSourceSpec(
        ("CT",),
        "impact_all_sheet",
        "OFFENSIVE IMPACT",
        "impact_all_sheet.csv -> OFFENSIVE IMPACT",
    ),
    ColumnSourceSpec(
        ("CY",),
        "playmaking_all_sheet",
        "Passing Creation Quality",
        "playmaking_all_sheet.csv -> Passing Creation Quality",
    ),
    ColumnSourceSpec(
        ("CZ",),
        "playmaking_all_sheet",
        "High Value Assists Per 75 Possessions",
        "playmaking_all_sheet.csv -> High Value Assists Per 75 Possessions",
    ),
    ColumnSourceSpec(
        ("DE",),
        "playmaking_all_sheet",
        "Ball Handle",
        "playmaking_all_sheet.csv -> Ball Handle",
    ),
    ColumnSourceSpec(
        ("DF",),
        "playmaking_all_sheet",
        "AVG_DRIB_PER_TOUCH",
        "playmaking_all_sheet.csv -> AVG_DRIB_PER_TOUCH",
    ),
    ColumnSourceSpec(
        ("DG",),
        "physical_all_sheet",
        "STAM",
        "physical_all_sheet.csv -> STAM",
    ),
    ColumnSourceSpec(
        ("DL",),
        "playmaking_all_sheet",
        "Speed with Ball",
        "playmaking_all_sheet.csv -> Speed with Ball",
    ),
    ColumnSourceSpec(
        ("DM",),
        "physical_all_sheet",
        "SPEED",
        "physical_all_sheet.csv -> SPEED",
    ),
    ColumnSourceSpec(
        ("DN", "FL"),
        "physical_all_sheet",
        "Agility",
        "physical_all_sheet.csv -> Agility",
    ),
    ColumnSourceSpec(
        ("DO",),
        "playmaking_all_sheet",
        "DRIVES",
        "playmaking_all_sheet.csv -> DRIVES",
    ),
    ColumnSourceSpec(
        ("ED", "EW", "FR"),
        "defense_all_sheet",
        "Perimeter Defense",
        "defense_all_sheet.csv -> Perimeter Defense",
    ),
    ColumnSourceSpec(
        ("EE",),
        "defense_all_sheet",
        "CONTESTED_SHOTS_3PT",
        "defense_all_sheet.csv -> CONTESTED_SHOTS_3PT",
    ),
    ColumnSourceSpec(
        ("EJ",),
        "defense_all_sheet",
        "Pickpocket Rating",
        "defense_all_sheet.csv -> Pickpocket Rating",
    ),
    ColumnSourceSpec(
        ("EK",),
        "defense_all_sheet",
        "Steal",
        "defense_all_sheet.csv -> Steal",
    ),
    ColumnSourceSpec(
        ("EL",),
        "defense_all_sheet",
        "Stable Recovered Blocks%",
        "defense_all_sheet.csv -> Stable Recovered Blocks%",
    ),
    ColumnSourceSpec(
        ("EU",),
        "physical_all_sheet",
        "STR",
        "physical_all_sheet.csv -> STR",
    ),
    ColumnSourceSpec(
        ("EV",),
        "defense_all_sheet",
        "Interior Defense",
        "defense_all_sheet.csv -> Interior Defense",
    ),
    ColumnSourceSpec(
        ("FC",),
        "defense_all_sheet",
        "DEFLECTIONS",
        "defense_all_sheet.csv -> DEFLECTIONS",
    ),
    ColumnSourceSpec(
        ("FD",),
        "defense_all_sheet",
        "Passing Lane Defense",
        "defense_all_sheet.csv -> Passing Lane Defense",
    ),
    ColumnSourceSpec(
        ("FE",),
        "defense_all_sheet",
        "Pass Perception",
        "defense_all_sheet.csv -> Pass Perception",
    ),
    ColumnSourceSpec(
        ("FJ",),
        "defense_all_sheet",
        "Off-Ball Chaser Defense",
        "defense_all_sheet.csv -> Off-Ball Chaser Defense",
    ),
    ColumnSourceSpec(
        ("FK",),
        "defense_all_sheet",
        "Help Defense IQ",
        "defense_all_sheet.csv -> Help Defense IQ",
    ),
    ColumnSourceSpec(
        ("FS",),
        "defense_all_sheet",
        "Perimeter Isolation Defense",
        "defense_all_sheet.csv -> Perimeter Isolation Defense",
    ),
    ColumnSourceSpec(
        ("GA",),
        "defense_all_sheet",
        "Ball Screen Navigation",
        "defense_all_sheet.csv -> Ball Screen Navigation",
    ),
    ColumnSourceSpec(
        ("GF",),
        "defense_all_sheet",
        "Post Defense",
        "defense_all_sheet.csv -> Post Defense",
    ),
    ColumnSourceSpec(
        ("GK",),
        "physical_all_sheet",
        "BOX_OUTS",
        "physical_all_sheet.csv -> BOX_OUTS",
    ),
    ColumnSourceSpec(
        ("GT",),
        "physical_all_sheet",
        "SCREEN_ASSISTS",
        "physical_all_sheet.csv -> SCREEN_ASSISTS",
    ),
    ColumnSourceSpec(
        ("HD",),
        "shooting_three_point_sheet",
        "Off-Ball Gravity",
        "shooting_three_point_sheet.csv -> Off-Ball Gravity",
    ),
    ColumnSourceSpec(
        ("HE",),
        "physical_all_sheet",
        "OFFSCREEN_POSS",
        "physical_all_sheet.csv -> OFFSCREEN_POSS",
    ),
)


RAW_METRIC_SPECS: Sequence[ColumnSourceSpec] = (
    ColumnSourceSpec(
        ("G",),
        "tight_plus_very_tight",
        "FG3M",
        "closest_defender.csv + very_tight.csv -> combined FG3M total",
        True,
    ),
    ColumnSourceSpec(
        ("H",),
        "tight_plus_very_tight",
        "FG3_PCT",
        "closest_defender.csv + very_tight.csv -> combined FG3_PCT from total FG3M / FG3A",
        True,
    ),
    ColumnSourceSpec(
        ("I",),
        "closest_defender",
        "FG3M",
        "closest_defender.csv -> FG3M total",
        True,
    ),
    ColumnSourceSpec(
        ("J",),
        "closest_defender",
        "FG3_PCT",
        "closest_defender.csv -> FG3_PCT",
        True,
    ),
    ColumnSourceSpec(
        ("Q",),
        "shot_locations_by_distance_8ft",
        "24+ ft._FGM",
        "shot_locations_by_distance_8ft.csv -> 24+ ft._FGM",
        True,
    ),
    ColumnSourceSpec(
        ("Z",),
        "tracking_c&s",
        "CATCH_SHOOT_FGM",
        "tracking_c&s.csv -> CATCH_SHOOT_FGM",
        True,
    ),
    ColumnSourceSpec(
        ("AA",),
        "tracking_c&s",
        "CATCH_SHOOT_EFG_PCT",
        "tracking_c&s.csv -> CATCH_SHOOT_EFG_PCT",
        True,
    ),
    ColumnSourceSpec(
        ("AB",),
        "spotup",
        "FGM",
        "spotup.csv -> FGM",
        True,
    ),
    ColumnSourceSpec(
        ("AC",),
        "spotup",
        "PPP",
        "spotup.csv -> PPP",
        True,
    ),
    ColumnSourceSpec(
        ("AK",),
        "bball_index_badges",
        "Pull-Up Shooting Talent",
        "bball_index_badges.csv -> Pull-Up Shooting Talent",
    ),
    ColumnSourceSpec(
        ("AL",),
        "bball_index_badges",
        "Off-Ball Shot Making",
        "bball_index_badges.csv -> Off-Ball Shot Making",
    ),
    ColumnSourceSpec(
        ("AM",),
        "pullup",
        "PULL_UP_FGM",
        "pullup.csv -> PULL_UP_FGM",
        True,
    ),
    ColumnSourceSpec(
        ("AN",),
        "pullup",
        "PULL_UP_EFG_PCT",
        "pullup.csv -> PULL_UP_EFG_PCT",
        True,
    ),
    ColumnSourceSpec(
        ("AV",),
        "putback_frequency",
        "Putback Frequency%",
        "nbarapm.csv -> SelfORebPct * TeammateMissORebPerc / 100",
    ),
    ColumnSourceSpec(
        ("AU",),
        "shooting_splits_badges",
        "Putback_FGM",
        "shooting_splits.csv -> sum of Putback *_FGM columns",
        True,
        True,
    ),
    ColumnSourceSpec(
        ("AX",),
        "shooting_splits_badges",
        "Alley_Oop_FGM",
        "shooting_splits.csv -> sum of Alley Oop *_FGM columns",
        True,
        True,
    ),
    ColumnSourceSpec(
        ("DY",),
        "bball_index_badges",
        "Lob Passing Creation Rate",
        "bball_index_badges.csv -> Lob Passing Creation Rate",
    ),
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build the badges workbook-shaped export from the existing ratings and history sources."
    )
    parser.add_argument(
        "--universe-csv",
        default=str(MANUAL_DIR / "playerlist.csv"),
        help="Season/player universe CSV. Defaults to stats/manual/playerlist.csv.",
    )
    parser.add_argument(
        "--workbook",
        default=str(MANUAL_DIR / "2k26_badges.xlsx"),
        help="Workbook used to mirror the badge sheet layout and formulas.",
    )
    parser.add_argument(
        "--sheet",
        default="",
        help="Optional workbook sheet name. Defaults to the first sheet in the workbook.",
    )
    parser.add_argument(
        "--details-csv",
        default="",
        help=(
            "Optional role/minutes detail CSV. "
            "Leave blank to enrich RotationRole from bball_index_badges.csv only."
        ),
    )
    parser.add_argument(
        "--minutes-source",
        default=str(HISTORY_DIR / "general_traditional.csv"),
        help="History CSV used to refresh MIN and GP in real time.",
    )
    parser.add_argument(
        "--minutes-column",
        default="MIN",
        help="Column from the minutes source used for live minutes.",
    )
    parser.add_argument(
        "--minutes-games-column",
        default="GP",
        help="Optional column used to convert a per-game MIN column into total minutes.",
    )
    parser.add_argument(
        "--current-season",
        default="",
        help="Season string that should use the lower in-season minute threshold.",
    )
    parser.add_argument(
        "--current-season-min-threshold",
        type=float,
        default=200.0,
        help="Minute threshold for penalty rows in the current season.",
    )
    parser.add_argument(
        "--standard-min-threshold",
        type=float,
        default=1000.0,
        help="Minute threshold for penalty rows in completed seasons.",
    )
    parser.add_argument(
        "--allow-id-fallback",
        action="store_true",
        help="Allow season+NBA_ID fallback when normalized season+player matching fails.",
    )
    parser.add_argument(
        "--output-prefix",
        default="badges_all",
        help="Prefix used for CSV outputs inside stats/exports.",
    )
    return parser.parse_args()


def read_dict_rows(path: Path) -> List[Dict[str, str]]:
    with path.open("r", newline="", encoding="utf-8") as fh:
        return [dict(row) for row in csv.DictReader(fh)]


def standardize_rows(
    rows: Sequence[Dict[str, str]],
    season_column: str,
    player_column: str,
    id_column: str = "",
) -> List[Dict[str, str]]:
    standardized: List[Dict[str, str]] = []
    for row in rows:
        season = str(row.get(season_column, "")).strip()
        player = str(row.get(player_column, "")).strip()
        if not season or not player:
            continue
        merged = dict(row)
        merged["Season"] = season
        merged["PLAYER_NAME"] = player
        if id_column:
            merged["PLAYER_ID"] = canonical_id(row.get(id_column, ""))
        else:
            merged.setdefault("PLAYER_ID", "")
        standardized.append(merged)
    return standardized


def load_standardized_sources() -> Dict[str, List[Dict[str, str]]]:
    standardized: Dict[str, List[Dict[str, str]]] = {}
    for spec in SOURCE_FILES:
        rows = read_history_csv(spec.path)
        standardized[spec.key] = standardize_rows(
            rows,
            season_column=spec.season_column,
            player_column=spec.player_column,
            id_column=spec.id_column,
        )
    standardized["tight_plus_very_tight"] = build_tight_plus_very_tight_rows(
        standardized["closest_defender"],
        standardized["very_tight"],
    )
    standardized["shooting_splits_badges"] = build_shooting_split_badge_rows(
        read_dict_rows(HISTORY_DIR / "shooting_splits.csv")
    )
    standardized["putback_frequency"] = build_putback_frequency_rows(
        standardized["nbarapm"],
    )
    return standardized


def build_badges_sheet_structure(
    workbook_path: Path,
    sheet_name: str,
) -> Tuple[str, List[str], List[str]]:
    workbook = load_workbook(workbook_path, data_only=False)
    selected_sheet = sheet_name.strip() or workbook.sheetnames[0]
    if selected_sheet not in workbook.sheetnames:
        raise SystemExit(f"Sheet not found in workbook: {selected_sheet}")
    worksheet = workbook[selected_sheet]
    column_letters = [get_column_letter(index) for index in range(1, worksheet.max_column + 1)]
    headers = [str(worksheet.cell(1, index).value or "").strip() for index in range(1, worksheet.max_column + 1)]
    return selected_sheet, column_letters, headers


def build_role_detail_rows(rows: Sequence[Dict[str, str]]) -> List[CalUniverseRow]:
    detail_rows: List[CalUniverseRow] = []
    for row in rows:
        season = str(row.get("Season", "")).strip()
        player = str(row.get("PLAYER_NAME", "")).strip()
        if not season or not player:
            continue
        detail_rows.append(
            CalUniverseRow(
                nba_id="",
                season=season,
                player=player,
                rotation_role=str(row.get("Rotation Role", "")).strip(),
                minutes=None,
            )
        )
    return detail_rows


def build_tight_plus_very_tight_rows(
    tight_rows: Sequence[Dict[str, str]],
    very_tight_rows: Sequence[Dict[str, str]],
) -> List[Dict[str, str]]:
    merged_rows: Dict[Tuple[str, str], Dict[str, str]] = {}

    def build_key(row: Dict[str, str]) -> Tuple[str, str]:
        season = str(row.get("Season", "")).strip()
        player_id = canonical_id(row.get("PLAYER_ID", ""))
        if player_id:
            return season, f"id:{player_id}"
        return season, f"name:{normalize_name(row.get('PLAYER_NAME', ''))}"

    for row in tight_rows:
        key = build_key(row)
        merged_rows[key] = {
            "Season": str(row.get("Season", "")).strip(),
            "PLAYER_ID": canonical_id(row.get("PLAYER_ID", "")),
            "PLAYER_NAME": str(row.get("PLAYER_NAME", "")).strip(),
            "FG3M": row.get("FG3M", ""),
            "FG3A": row.get("FG3A", ""),
        }

    for row in very_tight_rows:
        key = build_key(row)
        current = merged_rows.setdefault(
            key,
            {
                "Season": str(row.get("Season", "")).strip(),
                "PLAYER_ID": canonical_id(row.get("PLAYER_ID", "")),
                "PLAYER_NAME": str(row.get("PLAYER_NAME", "")).strip(),
                "FG3M": "",
                "FG3A": "",
            },
        )
        current["PLAYER_ID"] = current.get("PLAYER_ID", "") or canonical_id(row.get("PLAYER_ID", ""))
        current["PLAYER_NAME"] = current.get("PLAYER_NAME", "") or str(row.get("PLAYER_NAME", "")).strip()

        tight_fgm = parse_float(current.get("FG3M", "")) or 0.0
        tight_fga = parse_float(current.get("FG3A", "")) or 0.0
        very_fgm = parse_float(row.get("FG3M", "")) or 0.0
        very_fga = parse_float(row.get("FG3A", "")) or 0.0
        current["FG3M"] = str(tight_fgm + very_fgm)
        current["FG3A"] = str(tight_fga + very_fga)

    output_rows: List[Dict[str, str]] = []
    for row in merged_rows.values():
        fg3m = parse_float(row.get("FG3M", ""))
        fg3a = parse_float(row.get("FG3A", ""))
        output_rows.append(
            {
                "Season": str(row.get("Season", "")).strip(),
                "PLAYER_ID": canonical_id(row.get("PLAYER_ID", "")),
                "PLAYER_NAME": str(row.get("PLAYER_NAME", "")).strip(),
                "FG3M": fg3m if fg3m is not None else "",
                "FG3A": fg3a if fg3a is not None else "",
                "FG3_PCT": (fg3m / fg3a) if fg3m is not None and fg3a not in (None, 0) else "",
            }
        )
    return output_rows


def build_shooting_split_badge_rows(
    rows: Sequence[Dict[str, str]],
) -> List[Dict[str, str]]:
    if not rows:
        return []

    fieldnames = list(rows[0].keys())
    putback_fgm_columns = [
        column for column in fieldnames if "Putback" in column and column.endswith("_FGM")
    ]
    putback_fga_columns = [
        column for column in fieldnames if "Putback" in column and column.endswith("_FGA")
    ]
    alley_oop_fgm_columns = [
        column for column in fieldnames if "Alley Oop" in column and column.endswith("_FGM")
    ]

    output_rows: List[Dict[str, str]] = []
    for row in rows:
        season = str(row.get("Season", "")).strip()
        player = str(row.get("Name", "")).strip()
        if not season or not player:
            continue

        putback_fga = sum(parse_float(row.get(column, "")) or 0.0 for column in putback_fga_columns)
        putback_fgm = sum(parse_float(row.get(column, "")) or 0.0 for column in putback_fgm_columns)
        alley_oop_fgm = sum(parse_float(row.get(column, "")) or 0.0 for column in alley_oop_fgm_columns)

        output_rows.append(
            {
                "Season": season,
                "PLAYER_ID": canonical_id(row.get("PlayerID", "")),
                "PLAYER_NAME": player,
                "Putback_FGA": putback_fga,
                "Putback_FGM": putback_fgm,
                "Alley_Oop_FGM": alley_oop_fgm,
            }
        )
    return output_rows


def build_putback_frequency_rows(
    nbarapm_rows: Sequence[Dict[str, str]],
) -> List[Dict[str, str]]:
    output_rows: List[Dict[str, str]] = []

    for nbarapm_row in nbarapm_rows:
        season = str(nbarapm_row.get("Season", "")).strip()
        player = str(nbarapm_row.get("PLAYER_NAME", "")).strip()
        if not season or not player:
            continue

        self_oreb_pct = parse_float(nbarapm_row.get("SelfORebPct", ""))
        teammate_miss_oreb_perc = parse_float(nbarapm_row.get("TeammateMissORebPerc", ""))
        frequency = (
            ""
            if self_oreb_pct is None or teammate_miss_oreb_perc is None
            else (self_oreb_pct * teammate_miss_oreb_perc) / 100.0
        )

        output_rows.append(
            {
                "Season": season,
                "PLAYER_ID": canonical_id(nbarapm_row.get("PLAYER_ID", "")),
                "PLAYER_NAME": player,
                "Putback Frequency%": frequency,
                "SelfORebPct": "" if self_oreb_pct is None else self_oreb_pct,
                "TeammateMissORebPerc": "" if teammate_miss_oreb_perc is None else teammate_miss_oreb_perc,
            }
        )

    return output_rows


def build_nonzero_season_coverage(
    rows: Sequence[Dict[str, str]],
    source_column: str,
) -> set[str]:
    seasons: set[str] = set()
    for row in rows:
        season = str(row.get("Season", "")).strip()
        value = parse_float(row.get(source_column, ""))
        if season and value is not None and value > 0:
            seasons.add(season)
    return seasons


def resolve_direct_values(
    contexts,
    rows: Sequence[Dict[str, str]],
    source_column: str,
    allow_id_fallback: bool,
    default_zero_when_missing: bool = False,
    default_zero_nonzero_seasons: Optional[set[str]] = None,
) -> Tuple[List[object], List[str]]:
    by_id, by_name = build_source_index(list(rows))
    values: List[object] = []
    matched_by_values: List[str] = []

    for context in contexts:
        source_row, matched_by = match_metric_row(
            context.universe_row,
            by_id,
            by_name,
            allow_id_fallback=allow_id_fallback,
        )
        value = None if source_row is None else parse_float(source_row.get(source_column, ""))
        if (
            value is None
            and default_zero_when_missing
            and (
                default_zero_nonzero_seasons is None
                or context.universe_row.season in default_zero_nonzero_seasons
            )
        ):
            value = 0.0
            matched_by_values.append("default-zero" if not matched_by else f"{matched_by}|default-zero")
        else:
            matched_by_values.append(matched_by)
        values.append("" if value is None else value)

    return values, matched_by_values


def build_metric_result(
    contexts,
    rows: Sequence[Dict[str, str]],
    source_column: str,
    current_season: str,
    current_season_min_threshold: float,
    standard_min_threshold: float,
    allow_id_fallback: bool,
    source_note: str,
    default_zero_when_missing: bool = False,
    default_zero_nonzero_seasons: Optional[set[str]] = None,
) -> MetricResult:
    by_id, by_name = build_source_index(list(rows))
    raw_values: List[Optional[float]] = []
    matched_by_values: List[str] = []
    matched_numeric_values: List[float] = []

    for context in contexts:
        source_row, matched_by = match_metric_row(
            context.universe_row,
            by_id,
            by_name,
            allow_id_fallback=allow_id_fallback,
        )
        raw_value = None if source_row is None else parse_float(source_row.get(source_column, ""))
        if (
            raw_value is None
            and default_zero_when_missing
            and (
                default_zero_nonzero_seasons is None
                or context.universe_row.season in default_zero_nonzero_seasons
            )
        ):
            raw_value = 0.0
            matched_by = "default-zero" if not matched_by else f"{matched_by}|default-zero"
        raw_values.append(raw_value)
        matched_by_values.append(matched_by)
        if raw_value is not None:
            matched_numeric_values.append(raw_value)

    mean_value = None if not matched_numeric_values else statistics.mean(matched_numeric_values)
    stdev_value = (
        statistics.stdev(matched_numeric_values)
        if len(matched_numeric_values) >= 2
        else None
    )

    normalized_values: List[Optional[float]] = []
    for index, context in enumerate(contexts):
        raw_value = raw_values[index]
        if raw_value is None or mean_value is None or stdev_value in (None, 0):
            normalized_values.append(None)
            continue
        raw_z = compute_capped_z_score(raw_value, mean_value, stdev_value)
        normalized_values.append(
            apply_cal_normalization(
                season=context.universe_row.season,
                rotation_role=context.universe_row.rotation_role,
                minutes=context.effective_minutes,
                raw_z=raw_z,
                current_season=current_season,
                current_season_min_threshold=current_season_min_threshold,
                standard_min_threshold=standard_min_threshold,
            )
        )

    return MetricResult(
        raw_values=raw_values,
        normalized_values=normalized_values,
        matched_by=matched_by_values,
        mean_value=mean_value,
        stdev_value=stdev_value,
        source_note=source_note,
    )


def numeric(value: object) -> Optional[float]:
    return parse_float(value)


def numeric_or_zero(value: object) -> float:
    parsed = numeric(value)
    return 0.0 if parsed is None else parsed


def is_blank(value: object) -> bool:
    return numeric(value) is None


def average_numeric(values: Iterable[object]) -> Optional[float]:
    numeric_values = [numeric(value) for value in values]
    present_values = [value for value in numeric_values if value is not None]
    if not present_values:
        return None
    return statistics.mean(present_values)


def max_numeric(values: Iterable[object]) -> float:
    numeric_values = [numeric(value) for value in values]
    present_values = [value for value in numeric_values if value is not None]
    if not present_values:
        return 0.0
    return max(present_values)


def percentrank_inc(population: Sequence[object], current_value: object) -> object:
    current = numeric(current_value)
    ordered = sorted(value for value in (numeric(item) for item in population) if value is not None)
    if current is None or not ordered:
        return ""

    if len(ordered) == 1:
        return 1.0 if current >= ordered[0] else 0.0

    tolerance = 1e-12
    matching_indexes = [
        index
        for index, value in enumerate(ordered)
        if abs(value - current) <= tolerance
    ]
    if matching_indexes:
        return statistics.mean(index / (len(ordered) - 1) for index in matching_indexes)

    if current <= ordered[0]:
        return 0.0
    if current >= ordered[-1]:
        return 1.0

    for index in range(1, len(ordered)):
        lower = ordered[index - 1]
        upper = ordered[index]
        if lower < current < upper:
            lower_rank = (index - 1) / (len(ordered) - 1)
            upper_rank = index / (len(ordered) - 1)
            return lower_rank + (current - lower) / (upper - lower) * (upper_rank - lower_rank)

    return 1.0


def ge(value: object, threshold: float) -> bool:
    parsed = numeric(value)
    return parsed is not None and parsed >= threshold


def le(value: object, threshold: float) -> bool:
    parsed = numeric(value)
    return parsed is not None and parsed <= threshold


def any_ge(values: Iterable[object], threshold: float) -> bool:
    return any(ge(value, threshold) for value in values)


def badge_tier(cases: Sequence[Tuple[bool, int]]) -> int:
    for condition, tier in cases:
        if condition:
            return tier
    return 0


def compute_formula_columns(values: Dict[str, List[object]]) -> None:
    row_count = len(values["A"])

    values["F"] = []
    for index in range(row_count):
        average_value = average_numeric([values["G"][index], values["H"][index]])
        values["F"].append(
            (-1.0 if average_value is None else average_value) * 0.4
            + numeric_or_zero(values["K"][index]) * 0.3
            + numeric_or_zero(values["L"][index]) * 0.3
        )
    values["E"] = [percentrank_inc(values["F"], values["F"][index]) for index in range(row_count)]

    values["Y"] = [
        (numeric_or_zero(values["Z"][index]) + numeric_or_zero(values["AB"][index])) * 0.35
        + (numeric_or_zero(values["AA"][index]) + numeric_or_zero(values["AC"][index])) * 0.15
        for index in range(row_count)
    ]
    values["X"] = [percentrank_inc(values["Y"], values["Y"][index]) for index in range(row_count)]

    values["AH"] = [
        max_numeric([values["AI"][index], values["AJ"][index], values["AK"][index]]) * 0.4
        + numeric_or_zero(values["AL"][index]) * 0.4
        + numeric_or_zero(values["AM"][index]) * 0.15
        + numeric_or_zero(values["AN"][index]) * 0.05
        for index in range(row_count)
    ]
    values["AG"] = [percentrank_inc(values["AH"], values["AH"][index]) for index in range(row_count)]

    values["AT"] = []
    values["AS"] = []
    for index in range(row_count):
        average_value = average_numeric([values["AU"][index], values["AV"][index], values["AW"][index]])
        total = (0.0 if average_value is None else average_value * 0.4) + numeric_or_zero(values["AX"][index]) * 0.6
        values["AT"].append(total)
    for index in range(row_count):
        if all(is_blank(values[column][index]) for column in ("AU", "AV", "AW", "AX")):
            values["AS"].append(-1.0)
        else:
            values["AS"].append(percentrank_inc(values["AT"], values["AT"][index]))

    values["BE"] = []
    for index in range(row_count):
        average_value = average_numeric([values["BF"][index], values["BG"][index]])
        values["BE"].append(-0.5 if average_value is None else average_value)
    values["BD"] = [percentrank_inc(values["BE"], values["BE"][index]) for index in range(row_count)]

    values["CQ"] = [
        numeric_or_zero(values["CR"][index]) * 0.4
        + numeric_or_zero(values["CS"][index]) * 0.4
        + numeric_or_zero(values["CT"][index]) * 0.2
        for index in range(row_count)
    ]
    values["CP"] = [percentrank_inc(values["CQ"], values["CQ"][index]) for index in range(row_count)]

    values["CX"] = [
        average_numeric([values["CY"][index], values["CZ"][index]]) or ""
        for index in range(row_count)
    ]
    values["CW"] = [percentrank_inc(values["CX"], values["CX"][index]) for index in range(row_count)]

    values["DD"] = []
    for index in range(row_count):
        average_value = average_numeric([values["DE"][index], values["DF"][index]])
        base_value = values["DE"][index] if average_value is None else average_value
        values["DD"].append(
            numeric_or_zero(base_value) * 0.85 + numeric_or_zero(values["DG"][index]) * 0.15
        )
    values["DC"] = [percentrank_inc(values["DD"], values["DD"][index]) for index in range(row_count)]

    values["DK"] = []
    for index in range(row_count):
        primary_average = average_numeric([values["DL"][index], values["DM"][index], values["DN"][index], values["DO"][index]])
        if primary_average is None:
            primary_average = average_numeric([values["DL"][index], values["DM"][index], values["DN"][index]])
        values["DK"].append("" if primary_average is None else primary_average)
    values["DJ"] = [percentrank_inc(values["DK"], values["DK"][index]) for index in range(row_count)]

    values["EC"] = [
        average_numeric([values["ED"][index], values["EE"][index]]) or ""
        for index in range(row_count)
    ]
    values["EB"] = [percentrank_inc(values["EC"], values["EC"][index]) for index in range(row_count)]

    values["EI"] = [
        numeric_or_zero(values["EJ"][index]) * 0.4
        + numeric_or_zero(values["EK"][index]) * 0.45
        + numeric_or_zero(values["EL"][index]) * 0.15
        for index in range(row_count)
    ]
    values["EH"] = [percentrank_inc(values["EI"], values["EI"][index]) for index in range(row_count)]

    values["ET"] = [
        numeric_or_zero(values["EU"][index]) * 0.6
        + max_numeric([values["EV"][index], values["EW"][index]]) * 0.4
        for index in range(row_count)
    ]
    values["ES"] = [percentrank_inc(values["ET"], values["ET"][index]) for index in range(row_count)]

    values["FB"] = []
    for index in range(row_count):
        average_value = average_numeric([values["FC"][index], values["FD"][index], values["FE"][index]])
        values["FB"].append(-1.0 if average_value is None else average_value)
    values["FA"] = [percentrank_inc(values["FB"], values["FB"][index]) for index in range(row_count)]

    values["FI"] = []
    for index in range(row_count):
        if ge(values["FJ"][index], 0):
            values["FI"].append(
                numeric_or_zero(values["FJ"][index]) * 0.7
                + numeric_or_zero(values["FK"][index]) * 0.2
                + numeric_or_zero(values["FL"][index]) * 0.1
            )
        else:
            fallback_average = average_numeric([values["FK"][index], values["FL"][index]])
            values["FI"].append(-1.0 if fallback_average is None else fallback_average)
    values["FH"] = [percentrank_inc(values["FI"], values["FI"][index]) for index in range(row_count)]

    values["FQ"] = [
        average_numeric([values["FR"][index], values["FS"][index]]) or ""
        for index in range(row_count)
    ]
    values["FP"] = [percentrank_inc(values["FQ"], values["FQ"][index]) for index in range(row_count)]

    values["FZ"] = [
        0.0 if is_blank(values["GA"][index]) else percentrank_inc(values["GA"], values["GA"][index])
        for index in range(row_count)
    ]
    values["GE"] = [
        0.0 if is_blank(values["GF"][index]) else numeric_or_zero(percentrank_inc(values["GF"], values["GF"][index])) * 100.0
        for index in range(row_count)
    ]
    values["GJ"] = [
        0.0 if is_blank(values["GK"][index]) else numeric_or_zero(percentrank_inc(values["GK"], values["GK"][index])) * 100.0
        for index in range(row_count)
    ]
    values["GS"] = [
        0.0 if is_blank(values["GT"][index]) else percentrank_inc(values["GT"], values["GT"][index])
        for index in range(row_count)
    ]

    values["HC"] = []
    for index in range(row_count):
        average_value = average_numeric([values["HD"][index], values["HE"][index]])
        values["HC"].append(-0.5 if average_value is None else average_value)
    values["HB"] = [
        0.0 if is_blank(values["HC"][index]) else percentrank_inc(values["HC"], values["HC"][index])
        for index in range(row_count)
    ]

    values["P"] = [percentrank_inc(values["Q"], values["Q"][index]) for index in range(row_count)]
    values["BK"] = [percentrank_inc(values["BL"], values["BL"][index]) for index in range(row_count)]
    values["DX"] = [percentrank_inc(values["DY"], values["DY"][index]) for index in range(row_count)]

    for column in (
        "D", "O", "S", "W", "AF", "AR", "BC", "BJ", "BO", "BQ", "BS", "BV", "BZ", "CC", "CF", "CH",
        "CK", "CM", "CO", "CV", "DB", "DI", "DQ", "DT", "DW", "EA", "EG", "EN", "ER", "EZ", "FG",
        "FO", "FV", "FY", "GD", "GI", "GO", "GR", "GV", "HA",
    ):
        values[column] = []

    for index in range(row_count):
        values["D"].append(
            badge_tier(
                (
                    (ge(values["E"][index], 0.99) and any_ge((values["M"][index], values["N"][index]), 99), 5),
                    (ge(values["E"][index], 0.95) and any_ge((values["M"][index], values["N"][index]), 95), 4),
                    (ge(values["E"][index], 0.92) and any_ge((values["M"][index], values["N"][index]), 92), 3),
                    (ge(values["E"][index], 0.85) and any_ge((values["M"][index], values["N"][index]), 85), 2),
                    (ge(values["E"][index], 0.73) and any_ge((values["M"][index], values["N"][index]), 73), 1),
                )
            )
        )
        values["O"].append(
            badge_tier(
                (
                    (ge(values["P"][index], 0.99) and ge(values["R"][index], 99), 5),
                    (ge(values["P"][index], 0.96) and ge(values["R"][index], 96), 4),
                    (ge(values["P"][index], 0.93) and ge(values["R"][index], 93), 3),
                    (ge(values["P"][index], 0.89) and ge(values["R"][index], 89), 2),
                    (ge(values["P"][index], 0.83) and ge(values["R"][index], 83), 1),
                )
            )
        )
        values["S"].append(
            0
            if not le(values["V"][index], 75)
            else badge_tier(
                (
                    (any_ge((values["U"][index], values["T"][index]), 99), 5),
                    (any_ge((values["U"][index], values["T"][index]), 97), 4),
                    (any_ge((values["U"][index], values["T"][index]), 94), 3),
                    (ge(values["U"][index], 82) or ge(values["T"][index], 85), 2),
                    (any_ge((values["U"][index], values["T"][index]), 71), 1),
                )
            )
        )
        values["W"].append(
            badge_tier(
                (
                    (ge(values["X"][index], 0.98) and any_ge((values["AE"][index], values["AD"][index]), 98), 5),
                    (ge(values["X"][index], 0.94) and (ge(values["AE"][index], 95) or ge(values["AD"][index], 93)), 4),
                    (ge(values["X"][index], 0.89) and any_ge((values["AE"][index], values["AD"][index]), 89), 3),
                    (ge(values["X"][index], 0.78) and any_ge((values["AE"][index], values["AD"][index]), 78), 2),
                    (ge(values["X"][index], 0.65) and any_ge((values["AE"][index], values["AD"][index]), 65), 1),
                )
            )
        )
        values["AF"].append(
            badge_tier(
                (
                    (ge(values["AG"][index], 0.99) and ge(values["AQ"][index], 82) and any_ge((values["AO"][index], values["AP"][index]), 99), 5),
                    (ge(values["AG"][index], 0.96) and ge(values["AQ"][index], 76) and any_ge((values["AO"][index], values["AP"][index]), 96), 4),
                    (ge(values["AG"][index], 0.91) and ge(values["AQ"][index], 72) and any_ge((values["AO"][index], values["AP"][index]), 91), 3),
                    (ge(values["AG"][index], 0.87) and ge(values["AQ"][index], 65) and any_ge((values["AO"][index], values["AP"][index]), 87), 2),
                    (ge(values["AG"][index], 0.76) and ge(values["AQ"][index], 60) and any_ge((values["AO"][index], values["AP"][index]), 76), 1),
                )
            )
        )
        values["AR"].append(
            badge_tier(
                (
                    (ge(values["AS"][index], 0.96) and ge(values["BB"][index], 84) and (ge(values["AY"][index], 95) or ge(values["AZ"][index], 98) or ge(values["BA"][index], 97)), 5),
                    (ge(values["AS"][index], 0.89) and ge(values["BB"][index], 80) and (ge(values["AY"][index], 87) or ge(values["AZ"][index], 92) or ge(values["BA"][index], 89)), 4),
                    (ge(values["AS"][index], 0.80) and ge(values["BB"][index], 72) and (ge(values["AY"][index], 79) or ge(values["AZ"][index], 84) or ge(values["BA"][index], 80)), 3),
                    (ge(values["AS"][index], 0.70) and ge(values["BB"][index], 67) and (ge(values["AY"][index], 69) or ge(values["AZ"][index], 75) or ge(values["BA"][index], 70)), 2),
                    (ge(values["AS"][index], 0.70) and ge(values["BB"][index], 58) and (ge(values["AY"][index], 57) or ge(values["AZ"][index], 60) or ge(values["BA"][index], 54)), 1),
                )
            )
        )
        values["BC"].append(
            badge_tier(
                (
                    (ge(values["BD"][index], 0.98) and any_ge((values["BH"][index], values["BI"][index]), 98), 5),
                    (ge(values["BD"][index], 0.935) and (ge(values["BH"][index], 92) or ge(values["BI"][index], 95)), 4),
                    (ge(values["BD"][index], 0.87) and (ge(values["BH"][index], 86) or ge(values["BI"][index], 88)), 3),
                    (ge(values["BD"][index], 0.78) and any_ge((values["BH"][index], values["BI"][index]), 78), 2),
                    (ge(values["BD"][index], 0.665) and (ge(values["BH"][index], 68) or ge(values["BI"][index], 65)), 1),
                )
            )
        )
        values["BJ"].append(
            badge_tier(
                (
                    (ge(values["BK"][index], 0.97) and ge(values["BM"][index], 99) and ge(values["BN"][index], 97), 5),
                    (ge(values["BK"][index], 0.90) and ge(values["BM"][index], 94) and ge(values["BN"][index], 90), 4),
                    (ge(values["BK"][index], 0.835) and ge(values["BM"][index], 87) and ge(values["BN"][index], 80), 3),
                    (ge(values["BK"][index], 0.75) and ge(values["BM"][index], 75) and ge(values["BN"][index], 65), 2),
                    (ge(values["BK"][index], 0.61) and ge(values["BM"][index], 60) and ge(values["BN"][index], 61), 1),
                )
            )
        )
        values["BO"].append(
            badge_tier(((ge(values["BP"][index], 99), 5), (ge(values["BP"][index], 97), 4), (ge(values["BP"][index], 93), 3), (ge(values["BP"][index], 85), 2), (ge(values["BP"][index], 79), 1)))
        )
        values["BQ"].append(
            badge_tier(((ge(values["BR"][index], 99), 5), (ge(values["BR"][index], 96), 4), (ge(values["BR"][index], 92), 3), (ge(values["BR"][index], 84), 2), (ge(values["BR"][index], 73), 1)))
        )
        values["BS"].append(
            badge_tier(
                (
                    (ge(values["BT"][index], 97) and ge(values["BU"][index], 97), 5),
                    (ge(values["BT"][index], 96) and ge(values["BU"][index], 83), 4),
                    (ge(values["BT"][index], 90) and ge(values["BU"][index], 75), 3),
                    (ge(values["BT"][index], 80) and ge(values["BU"][index], 67), 2),
                    (ge(values["BT"][index], 70) and ge(values["BU"][index], 60), 1),
                )
            )
        )
        values["BV"].append(
            badge_tier(
                (
                    ((ge(values["BW"][index], 99) or ge(values["BX"][index], 99)) and ge(values["BY"][index], 90), 5),
                    ((ge(values["BW"][index], 96) or ge(values["BX"][index], 96)) and ge(values["BY"][index], 85), 4),
                    ((ge(values["BW"][index], 93) or ge(values["BX"][index], 93)) and ge(values["BY"][index], 80), 3),
                    ((ge(values["BW"][index], 87) or ge(values["BX"][index], 87)) and ge(values["BY"][index], 75), 2),
                    ((ge(values["BW"][index], 73) or ge(values["BX"][index], 73)) and ge(values["BY"][index], 65), 1),
                )
            )
        )
        values["BZ"].append(
            badge_tier(
                (
                    (ge(values["CA"][index], 90) and ge(values["CB"][index], 94), 5),
                    (ge(values["CA"][index], 84) and ge(values["CB"][index], 90), 4),
                    (ge(values["CA"][index], 79) and ge(values["CB"][index], 80), 3),
                    (ge(values["CA"][index], 70) and ge(values["CB"][index], 71), 2),
                    (ge(values["CA"][index], 60) and ge(values["CB"][index], 61), 1),
                )
            )
        )
        values["CC"].append(
            badge_tier(
                (
                    (ge(values["CD"][index], 98) and ge(values["CE"][index], 96), 5),
                    (ge(values["CD"][index], 93) and ge(values["CE"][index], 95), 4),
                    (ge(values["CD"][index], 85) and ge(values["CE"][index], 86), 3),
                    (ge(values["CD"][index], 75) and ge(values["CE"][index], 79), 2),
                    (ge(values["CD"][index], 64) and ge(values["CE"][index], 70), 1),
                )
            )
        )
        values["CF"].append(
            badge_tier(((ge(values["CG"][index], 99), 5), (ge(values["CG"][index], 95), 4), (ge(values["CG"][index], 87), 3), (ge(values["CG"][index], 77), 2), (ge(values["CG"][index], 67), 1)))
        )
        values["CH"].append(
            badge_tier(
                (
                    (ge(values["CI"][index], 99) and ge(values["CJ"][index], 71), 5),
                    (ge(values["CI"][index], 95) and ge(values["CJ"][index], 60), 4),
                    (ge(values["CI"][index], 90) and ge(values["CJ"][index], 60), 3),
                    (ge(values["CI"][index], 81) and ge(values["CJ"][index], 60), 2),
                    (ge(values["CI"][index], 72) and ge(values["CJ"][index], 60), 1),
                )
            )
        )
        values["CK"].append(
            badge_tier(((ge(values["CL"][index], 98), 5), (ge(values["CL"][index], 95), 4), (ge(values["CL"][index], 93), 3), (ge(values["CL"][index], 86), 2), (ge(values["CL"][index], 75), 1)))
        )
        values["CM"].append(
            badge_tier(((ge(values["CN"][index], 99), 5), (ge(values["CN"][index], 96), 4), (ge(values["CN"][index], 94), 3), (ge(values["CN"][index], 91), 2), (ge(values["CN"][index], 85), 1)))
        )
        values["CO"].append(
            badge_tier(
                (
                    (ge(values["CP"][index], 0.98) and ge(values["CU"][index], 98), 5),
                    (ge(values["CP"][index], 0.93) and ge(values["CU"][index], 93), 4),
                    (ge(values["CP"][index], 0.87) and ge(values["CU"][index], 87), 3),
                    (ge(values["CP"][index], 0.75) and ge(values["CU"][index], 75), 2),
                    (ge(values["CP"][index], 0.65) and ge(values["CU"][index], 65), 1),
                )
            )
        )
        values["CV"].append(
            badge_tier(
                (
                    (ge(values["CW"][index], 0.98) and ge(values["DA"][index], 98), 5),
                    (ge(values["CW"][index], 0.92) and ge(values["DA"][index], 92), 4),
                    (ge(values["CW"][index], 0.82) and ge(values["DA"][index], 82), 3),
                    (ge(values["CW"][index], 0.71) and ge(values["DA"][index], 71), 2),
                    (ge(values["CW"][index], 0.55) and ge(values["DA"][index], 55), 1),
                )
            )
        )
        values["DB"].append(
            badge_tier(
                (
                    (ge(values["DC"][index], 0.97) and ge(values["DH"][index], 97), 5),
                    (ge(values["DC"][index], 0.94) and ge(values["DH"][index], 94), 4),
                    (ge(values["DC"][index], 0.90) and ge(values["DH"][index], 90), 3),
                    (ge(values["DC"][index], 0.81) and ge(values["DH"][index], 81), 2),
                    (ge(values["DC"][index], 0.71) and ge(values["DH"][index], 71), 1),
                )
            )
        )
        values["DI"].append(
            badge_tier(
                (
                    (ge(values["DJ"][index], 0.94) and ge(values["DP"][index], 94), 5),
                    (ge(values["DJ"][index], 0.91) and ge(values["DP"][index], 91), 4),
                    (ge(values["DJ"][index], 0.86) and ge(values["DP"][index], 86), 3),
                    (ge(values["DJ"][index], 0.75) and ge(values["DP"][index], 75), 2),
                    (ge(values["DJ"][index], 0.68) and ge(values["DP"][index], 68), 1),
                )
            )
        )
        values["DQ"].append(
            badge_tier(
                (
                    (ge(values["DR"][index], 80) and ge(values["DS"][index], 93), 5),
                    (ge(values["DR"][index], 77) and ge(values["DS"][index], 84), 4),
                    (ge(values["DR"][index], 73) and ge(values["DS"][index], 73), 3),
                    (ge(values["DR"][index], 67) and ge(values["DS"][index], 65), 2),
                    (ge(values["DR"][index], 60) and ge(values["DS"][index], 60), 1),
                )
            )
        )
        values["DT"].append(
            badge_tier(
                (
                    (ge(values["DU"][index], 99), 5),
                    (ge(values["DU"][index], 96), 4),
                    ((ge(values["DU"][index], 92) or ge(values["DV"][index], 96)), 3),
                    ((ge(values["DU"][index], 80) or ge(values["DV"][index], 86)), 2),
                    ((ge(values["DU"][index], 70) or ge(values["DV"][index], 75)), 1),
                )
            )
        )
        values["DW"].append(
            badge_tier(
                (
                    (ge(values["DX"][index], 0.99) and ge(values["DZ"][index], 99), 5),
                    ((ge(values["DX"][index], 0.97) or ge(values["DZ"][index], 97)) and ge(values["DX"][index], 0.89) and ge(values["DZ"][index], 89), 4),
                    ((ge(values["DX"][index], 0.95) or ge(values["DZ"][index], 95)) and ge(values["DX"][index], 0.89) and ge(values["DZ"][index], 89), 3),
                    ((ge(values["DX"][index], 0.89) or ge(values["DZ"][index], 89)) and ge(values["DX"][index], 0.78) and ge(values["DZ"][index], 78), 2),
                    (ge(values["DX"][index], 0.78) and ge(values["DZ"][index], 78), 1),
                )
            )
        )
        values["EA"].append(
            badge_tier(
                (
                    (ge(values["EB"][index], 0.99) and ge(values["EF"][index], 99), 5),
                    (ge(values["EB"][index], 0.95) and ge(values["EF"][index], 95), 4),
                    (ge(values["EB"][index], 0.92) and ge(values["EF"][index], 92), 3),
                    (ge(values["EB"][index], 0.82) and ge(values["EF"][index], 82), 2),
                    (ge(values["EB"][index], 0.71) and ge(values["EF"][index], 71), 1),
                )
            )
        )
        values["EG"].append(
            badge_tier(
                (
                    (ge(values["EH"][index], 0.99) and ge(values["EM"][index], 99), 5),
                    (ge(values["EH"][index], 0.96) and ge(values["EM"][index], 96), 4),
                    (ge(values["EH"][index], 0.91) and ge(values["EM"][index], 91), 3),
                    (ge(values["EH"][index], 0.79) and ge(values["EM"][index], 79), 2),
                    (ge(values["EH"][index], 0.75) and ge(values["EM"][index], 67), 1),
                )
            )
        )
        values["EN"].append(
            badge_tier(
                (
                    (ge(values["EO"][index], 99) and (ge(values["EP"][index], 75) or ge(values["EQ"][index], 85)), 5),
                    (ge(values["EO"][index], 92) and (ge(values["EP"][index], 71) or ge(values["EQ"][index], 83)), 4),
                    (ge(values["EO"][index], 88) and (ge(values["EP"][index], 66) or ge(values["EQ"][index], 80)), 3),
                    (ge(values["EO"][index], 78) and (ge(values["EP"][index], 60) or ge(values["EQ"][index], 74)), 2),
                    (ge(values["EO"][index], 68) and ge(values["EP"][index], 50) and ge(values["EQ"][index], 60), 1),
                )
            )
        )
        values["ER"].append(
            badge_tier(
                (
                    (ge(values["ES"][index], 0.99) and (ge(values["EX"][index], 98) or ge(values["EY"][index], 98)), 5),
                    (ge(values["ES"][index], 0.95) and (ge(values["EX"][index], 95) or ge(values["EY"][index], 95)), 4),
                    (ge(values["ES"][index], 0.92) and (ge(values["EX"][index], 90) or ge(values["EY"][index], 90)), 3),
                    (ge(values["ES"][index], 0.77) and ge(values["EX"][index], 72) and ge(values["EY"][index], 82), 2),
                    (ge(values["ES"][index], 0.67) and ge(values["EX"][index], 62) and ge(values["EY"][index], 71), 1),
                )
            )
        )
        values["EZ"].append(
            badge_tier(
                (
                    (ge(values["FA"][index], 0.98) and ge(values["FF"][index], 98), 5),
                    (ge(values["FA"][index], 0.94) and ge(values["FF"][index], 94), 4),
                    (ge(values["FA"][index], 0.85) and ge(values["FF"][index], 85), 3),
                    (ge(values["FA"][index], 0.85) and ge(values["FF"][index], 73), 2),
                    (ge(values["FA"][index], 0.73) and ge(values["FF"][index], 60), 1),
                )
            )
        )
        values["FG"].append(
            badge_tier(
                (
                    (ge(values["FH"][index], 0.975) and (ge(values["FM"][index], 97) or ge(values["FN"][index], 98)), 5),
                    (ge(values["FH"][index], 0.90) and (ge(values["FM"][index], 94) or ge(values["FN"][index], 87)), 4),
                    (ge(values["FH"][index], 0.83) and (ge(values["FM"][index], 85) or ge(values["FN"][index], 80)), 3),
                    (ge(values["FH"][index], 0.72) and (ge(values["FM"][index], 76) or ge(values["FN"][index], 68)), 2),
                    (ge(values["FH"][index], 0.63) and (ge(values["FM"][index], 69) or ge(values["FN"][index], 58)), 1),
                )
            )
        )
        values["FO"].append(
            badge_tier(
                (
                    (ge(values["FP"][index], 0.92) and ge(values["FT"][index], 99) and ge(values["FU"][index], 86), 5),
                    (ge(values["FP"][index], 0.90) and ge(values["FT"][index], 96) and ge(values["FU"][index], 84), 4),
                    (ge(values["FP"][index], 0.855) and ge(values["FT"][index], 91) and ge(values["FU"][index], 80), 3),
                    (ge(values["FP"][index], 0.805) and ge(values["FT"][index], 85) and ge(values["FU"][index], 76), 2),
                    (ge(values["FP"][index], 0.72) and ge(values["FT"][index], 74) and ge(values["FU"][index], 70), 1),
                )
            )
        )
        values["FV"].append(
            badge_tier(
                (
                    (ge(values["FW"][index], 89) and ge(values["FX"][index], 99), 5),
                    (ge(values["FW"][index], 84) and ge(values["FX"][index], 97), 4),
                    (ge(values["FW"][index], 77) and ge(values["FX"][index], 93), 3),
                    (ge(values["FW"][index], 70) and ge(values["FX"][index], 84), 2),
                    (ge(values["FW"][index], 60) and ge(values["FX"][index], 74), 1),
                )
            )
        )
        values["FY"].append(
            badge_tier(
                (
                    (ge(values["FZ"][index], 0.995) or (ge(values["GB"][index], 99) and ge(values["GC"][index], 92)), 5),
                    (ge(values["FZ"][index], 0.99) or (ge(values["GB"][index], 97) and ge(values["GC"][index], 85)), 4),
                    (ge(values["FZ"][index], 0.98) or (ge(values["GB"][index], 90) and ge(values["GC"][index], 79)), 3),
                    (ge(values["FZ"][index], 0.75) and ge(values["GB"][index], 83) and ge(values["GC"][index], 75), 2),
                    (ge(values["FZ"][index], 0.72) and ge(values["GB"][index], 73) and ge(values["GC"][index], 71), 1),
                )
            )
        )
        values["GD"].append(
            badge_tier(
                (
                    (ge(values["GE"][index], 97) and (ge(values["GG"][index], 99) or ge(values["GH"][index], 97)), 5),
                    (ge(values["GE"][index], 88) and ge(values["GG"][index], 93) and ge(values["GH"][index], 92), 4),
                    (ge(values["GE"][index], 93) or (ge(values["GG"][index], 88) and ge(values["GH"][index], 84)), 3),
                    (ge(values["GE"][index], 88) or (ge(values["GG"][index], 82) and ge(values["GH"][index], 78)), 2),
                    (ge(values["GE"][index], 78) or (ge(values["GG"][index], 74) and ge(values["GH"][index], 70)), 1),
                )
            )
        )
        values["GI"].append(
            badge_tier(
                (
                    (ge(values["GJ"][index], 96) and ge(values["GL"][index], 96) and (ge(values["GM"][index], 98) or ge(values["GN"][index], 98)), 5),
                    (ge(values["GJ"][index], 89) and ge(values["GL"][index], 89) and (ge(values["GM"][index], 94) or ge(values["GN"][index], 94)), 4),
                    (ge(values["GJ"][index], 83) and ge(values["GL"][index], 83) and (ge(values["GM"][index], 85) or ge(values["GN"][index], 85)), 3),
                    (ge(values["GJ"][index], 70) and ge(values["GL"][index], 71) and (ge(values["GM"][index], 70) or ge(values["GN"][index], 70)), 2),
                    (ge(values["GJ"][index], 55) and ge(values["GL"][index], 68) and (ge(values["GM"][index], 55) or ge(values["GN"][index], 55)), 1),
                )
            )
        )
        values["GO"].append(
            badge_tier(
                (
                    (ge(values["GP"][index], 99) or ge(values["GQ"][index], 99), 5),
                    (ge(values["GP"][index], 96) or ge(values["GQ"][index], 96), 4),
                    (ge(values["GP"][index], 92) or ge(values["GQ"][index], 92), 3),
                    (ge(values["GP"][index], 80) or ge(values["GQ"][index], 80), 2),
                    (ge(values["GP"][index], 60) and ge(values["GQ"][index], 60), 1),
                )
            )
        )
        values["GR"].append(
            badge_tier(
                (
                    (ge(values["GS"][index], 0.99) and ge(values["GU"][index], 99), 5),
                    (ge(values["GS"][index], 0.95) and ge(values["GU"][index], 95), 4),
                    (ge(values["GS"][index], 0.91) and ge(values["GU"][index], 91), 3),
                    (ge(values["GS"][index], 0.83) and ge(values["GU"][index], 83), 2),
                    (ge(values["GS"][index], 0.72) and ge(values["GU"][index], 72), 1),
                )
            )
        )
        values["GV"].append(
            badge_tier(
                (
                    (ge(values["GW"][index], 88) and (ge(values["GX"][index], 99) or ge(values["GY"][index], 99) or ge(values["GZ"][index], 99)), 5),
                    (ge(values["GW"][index], 83) and (ge(values["GX"][index], 95) or ge(values["GY"][index], 96) or ge(values["GZ"][index], 96)), 4),
                    (ge(values["GW"][index], 77) and (ge(values["GX"][index], 94) or ge(values["GY"][index], 91) or ge(values["GZ"][index], 91)), 3),
                    (ge(values["GW"][index], 70) and (ge(values["GX"][index], 86) or ge(values["GY"][index], 83) or ge(values["GZ"][index], 83)), 2),
                    (ge(values["GW"][index], 63) and (ge(values["GX"][index], 80) or ge(values["GY"][index], 80) or ge(values["GZ"][index], 80)), 1),
                )
            )
        )
        values["HA"].append(
            badge_tier(
                (
                    (ge(values["HB"][index], 1.0) or ge(values["HF"][index], 99) or ge(values["HG"][index], 99), 5),
                    (ge(values["HB"][index], 0.995) or (ge(values["HF"][index], 92) and ge(values["HG"][index], 88)), 4),
                    (ge(values["HB"][index], 0.81) and ge(values["HF"][index], 85) and ge(values["HG"][index], 77), 3),
                    (ge(values["HB"][index], 0.69) and ge(values["HF"][index], 73) and ge(values["HG"][index], 65), 2),
                    (ge(values["HB"][index], 0.57) and ge(values["HF"][index], 57) and ge(values["HG"][index], 57), 1),
                )
            )
        )


def populate_source_columns(
    column_values: Dict[str, List[object]],
    column_matched_by: Dict[str, List[str]],
    column_source_notes: Dict[str, str],
    contexts,
    source_rows_by_key: Dict[str, List[Dict[str, str]]],
    specs: Sequence[ColumnSourceSpec],
    allow_id_fallback: bool,
) -> None:
    for spec in specs:
        default_zero_nonzero_seasons = (
            build_nonzero_season_coverage(source_rows_by_key[spec.source_key], spec.source_column)
            if spec.default_zero_requires_season_nonzero_coverage
            else None
        )
        values, matched_by = resolve_direct_values(
            contexts=contexts,
            rows=source_rows_by_key[spec.source_key],
            source_column=spec.source_column,
            allow_id_fallback=allow_id_fallback,
            default_zero_when_missing=spec.default_zero_when_missing,
            default_zero_nonzero_seasons=default_zero_nonzero_seasons,
        )
        for column in spec.columns:
            column_values[column] = list(values)
            column_matched_by[column] = list(matched_by)
            column_source_notes[column] = spec.source_note


def populate_raw_metric_columns(
    column_values: Dict[str, List[object]],
    column_matched_by: Dict[str, List[str]],
    column_source_notes: Dict[str, str],
    contexts,
    source_rows_by_key: Dict[str, List[Dict[str, str]]],
    current_season: str,
    current_season_min_threshold: float,
    standard_min_threshold: float,
    allow_id_fallback: bool,
) -> Dict[str, MetricResult]:
    metric_results: Dict[str, MetricResult] = {}
    for spec in RAW_METRIC_SPECS:
        default_zero_nonzero_seasons = (
            build_nonzero_season_coverage(source_rows_by_key[spec.source_key], spec.source_column)
            if spec.default_zero_requires_season_nonzero_coverage
            else None
        )
        result = build_metric_result(
            contexts=contexts,
            rows=source_rows_by_key[spec.source_key],
            source_column=spec.source_column,
            current_season=current_season,
            current_season_min_threshold=current_season_min_threshold,
            standard_min_threshold=standard_min_threshold,
            allow_id_fallback=allow_id_fallback,
            source_note=spec.source_note,
            default_zero_when_missing=spec.default_zero_when_missing,
            default_zero_nonzero_seasons=default_zero_nonzero_seasons,
        )
        metric_results[f"{spec.source_key}:{spec.source_column}"] = result
        values = ["" if value is None else value for value in result.normalized_values]
        for column in spec.columns:
            column_values[column] = list(values)
            column_matched_by[column] = list(result.matched_by)
            column_source_notes[column] = spec.source_note
    return metric_results


def build_badge_only_rows(
    column_values: Dict[str, List[object]],
    contexts,
) -> Tuple[List[str], List[Dict[str, object]]]:
    badge_columns = [
        "Deadeye", "Limitless Range", "Mini Marksman", "Set Shot Specialist", "Shifty Shooter",
        "Aerial Wizard", "Float Game", "Hook Specialist", "Layup Mixmaster", "Paint Prodigy",
        "Physical Finisher", "Posterizer", "Post Fade Phenom", "Post Powerhouse", "Post Up Poet",
        "Rise Up", "Ankle Assassin", "Bail Out", "Break Starter", "Dimer", "Handles for Days",
        "Lightning Launch", "Strong Handle", "Unpluckable", "Versatile Visionary", "Challenger",
        "Glove", "High Flying Denier", "Immovable Enforcer", "Interceptor", "Off Ball Pest",
        "On-Ball Menace", "Paint Patroller", "Pick Dodger", "Post Lockdown", "Boxout Beast",
        "Rebound Chaser", "Brick Wall", "Pogo Stick", "Slippery Off Ball",
    ]
    badge_headers = ["NBA_ID", "Season", "Player", *badge_columns]
    rows: List[Dict[str, object]] = []
    header_to_column = {
        "Deadeye": "D", "Limitless Range": "O", "Mini Marksman": "S", "Set Shot Specialist": "W",
        "Shifty Shooter": "AF", "Aerial Wizard": "AR", "Float Game": "BC", "Hook Specialist": "BJ",
        "Layup Mixmaster": "BO", "Paint Prodigy": "BQ", "Physical Finisher": "BS", "Posterizer": "BV",
        "Post Fade Phenom": "BZ", "Post Powerhouse": "CC", "Post Up Poet": "CF", "Rise Up": "CH",
        "Ankle Assassin": "CK", "Bail Out": "CM", "Break Starter": "CO", "Dimer": "CV",
        "Handles for Days": "DB", "Lightning Launch": "DI", "Strong Handle": "DQ", "Unpluckable": "DT",
        "Versatile Visionary": "DW", "Challenger": "EA", "Glove": "EG", "High Flying Denier": "EN",
        "Immovable Enforcer": "ER", "Interceptor": "EZ", "Off Ball Pest": "FG", "On-Ball Menace": "FO",
        "Paint Patroller": "FV", "Pick Dodger": "FY", "Post Lockdown": "GD", "Boxout Beast": "GI",
        "Rebound Chaser": "GO", "Brick Wall": "GR", "Pogo Stick": "GV", "Slippery Off Ball": "HA",
    }

    for index, context in enumerate(contexts):
        row: Dict[str, object] = {
            "NBA_ID": context.universe_row.nba_id,
            "Season": context.universe_row.season,
            "Player": context.universe_row.player,
        }
        for header in badge_columns:
            row[header] = column_values[header_to_column[header]][index]
        rows.append(row)

    return badge_headers, rows


def main() -> None:
    args = parse_args()
    workbook_path = Path(args.workbook)
    sheet_name, workbook_column_letters, workbook_headers = build_badges_sheet_structure(
        workbook_path=workbook_path,
        sheet_name=args.sheet,
    )
    column_position_map = {column: index for index, column in enumerate(workbook_column_letters)}
    header_by_column = {
        column: workbook_headers[column_position_map[column]]
        for column in workbook_column_letters
    }

    universe_path = Path(args.universe_csv)
    if universe_path.suffix.lower() == ".csv":
        universe = load_universe_csv(universe_path)
    else:
        universe = load_cal_universe(universe_path, sheet_name=sheet_name)

    source_rows_by_key = load_standardized_sources()
    role_details = build_role_detail_rows(source_rows_by_key["bball_index_badges"])

    details_csv_path = resolve_details_csv_path(args.details_csv, universe_path)
    if details_csv_path is not None:
        if details_csv_path.suffix.lower() == ".csv":
            detail_rows = load_universe_csv(details_csv_path)
        else:
            detail_rows = load_cal_universe(details_csv_path, sheet_name=sheet_name)
        universe = enrich_universe_rows(universe, detail_rows)

    universe = enrich_universe_rows(universe, role_details)

    minutes_rows = source_rows_by_key["general_traditional"]
    current_season = detect_current_season(
        explicit_current_season=args.current_season,
        universe_rows=universe,
        source_rows=minutes_rows,
    )
    contexts = build_player_contexts(
        universe=universe,
        minutes_rows=minutes_rows,
        minutes_column=args.minutes_column,
        minutes_games_column=args.minutes_games_column,
        allow_id_fallback=args.allow_id_fallback,
    )

    row_count = len(contexts)
    column_values: Dict[str, List[object]] = {
        "A": [context.universe_row.nba_id for context in contexts],
        "B": [context.universe_row.season for context in contexts],
        "C": [context.universe_row.player for context in contexts],
    }
    column_matched_by: Dict[str, List[str]] = {
        "A": ["universe"] * row_count,
        "B": ["universe"] * row_count,
        "C": ["universe"] * row_count,
    }
    column_source_notes: Dict[str, str] = {
        "A": "playerlist.csv -> NBA_ID",
        "B": "playerlist.csv -> Season",
        "C": "playerlist.csv -> Player",
    }

    populate_source_columns(column_values, column_matched_by, column_source_notes, contexts, source_rows_by_key, RATING_SPECS, args.allow_id_fallback)
    populate_source_columns(column_values, column_matched_by, column_source_notes, contexts, source_rows_by_key, PRECOMPUTED_SPECS, args.allow_id_fallback)
    raw_metric_results = populate_raw_metric_columns(
        column_values=column_values,
        column_matched_by=column_matched_by,
        column_source_notes=column_source_notes,
        contexts=contexts,
        source_rows_by_key=source_rows_by_key,
        current_season=current_season,
        current_season_min_threshold=args.current_season_min_threshold,
        standard_min_threshold=args.standard_min_threshold,
        allow_id_fallback=args.allow_id_fallback,
    )

    direct_values, direct_matched_by = resolve_direct_values(
        contexts=contexts,
        rows=source_rows_by_key["bios"],
        source_column="PLAYER_HEIGHT_INCHES",
        allow_id_fallback=args.allow_id_fallback,
    )
    column_values["V"] = direct_values
    column_matched_by["V"] = direct_matched_by
    column_source_notes["V"] = "bios.csv -> PLAYER_HEIGHT_INCHES"

    compute_formula_columns(column_values)

    sheet_rows: List[List[object]] = []
    badge_headers, badge_only_rows = build_badge_only_rows(column_values=column_values, contexts=contexts)
    audit_rows: List[Dict[str, object]] = []
    unmatched_rows: List[Dict[str, object]] = []

    base_columns = sorted(
        set(column_matched_by.keys()) - {"A", "B", "C"},
        key=lambda column: column_position_map[column],
    )
    base_column_labels = {
        column: f"{column} {header_by_column[column] or '(blank)'}".strip()
        for column in base_columns
    }
    badge_column_letters = [
        "D", "O", "S", "W", "AF", "AR", "BC", "BJ", "BO", "BQ", "BS", "BV", "BZ", "CC", "CF", "CH",
        "CK", "CM", "CO", "CV", "DB", "DI", "DQ", "DT", "DW", "EA", "EG", "EN", "ER", "EZ", "FG",
        "FO", "FV", "FY", "GD", "GI", "GO", "GR", "GV", "HA",
    ]

    for index, context in enumerate(contexts):
        row_values_by_column = {column: column_values.get(column, [""] * row_count)[index] for column in workbook_column_letters}
        sheet_rows.append([row_values_by_column[column] for column in workbook_column_letters])

        missing_base_columns = [
            base_column_labels[column]
            for column in base_columns
            if column not in OPTIONAL_BASE_COLUMNS and is_blank(column_values[column][index])
        ]

        audit_row: Dict[str, object] = {
            "NBA_ID": context.universe_row.nba_id,
            "Season": context.universe_row.season,
            "Player": context.universe_row.player,
            "RotationRole": context.universe_row.rotation_role,
            "MIN": context.effective_minutes,
            "WorkbookMIN": context.workbook_minutes,
            "LiveMIN": context.live_minutes,
            "LiveMINPerGame": context.live_minutes_per_game,
            "LiveGP": context.live_gp,
            "MinutesMatchedBy": context.minutes_matched_by,
            "CurrentSeason": current_season,
            "MissingBaseCount": len(missing_base_columns),
            "MissingBaseColumns": " | ".join(missing_base_columns),
            "ActiveBadgeCount": sum(1 for column in badge_column_letters if numeric_or_zero(column_values[column][index]) > 0),
        }
        for column in badge_column_letters:
            audit_row[header_by_column[column]] = column_values[column][index]
        for column in base_columns:
            label = base_column_labels[column]
            audit_row[f"{label} Value"] = column_values[column][index]
            audit_row[f"{label} MatchedBy"] = column_matched_by.get(column, [""] * row_count)[index]
            audit_row[f"{label} Source"] = column_source_notes.get(column, "")
        audit_rows.append(audit_row)

        if missing_base_columns:
            unmatched_rows.append(
                {
                    "NBA_ID": context.universe_row.nba_id,
                    "Season": context.universe_row.season,
                    "Player": context.universe_row.player,
                    "RotationRole": context.universe_row.rotation_role,
                    "MIN": context.effective_minutes,
                    "MissingBaseCount": len(missing_base_columns),
                    "MissingBaseColumns": " | ".join(missing_base_columns),
                }
            )

    output_prefix = args.output_prefix.strip() or "badges_all"
    sheet_path = EXPORT_DIR / f"{output_prefix}_sheet.csv"
    badge_only_path = EXPORT_DIR / f"{output_prefix}_badges.csv"
    audit_path = EXPORT_DIR / f"{output_prefix}_audit.csv"
    unmatched_path = EXPORT_DIR / f"{output_prefix}_unmatched.csv"

    write_matrix_csv(sheet_path, workbook_headers, sheet_rows)
    write_csv(badge_only_path, badge_headers, badge_only_rows)
    write_csv(audit_path, list(audit_rows[0].keys()) if audit_rows else [], audit_rows)
    write_csv(unmatched_path, ["NBA_ID", "Season", "Player", "RotationRole", "MIN", "MissingBaseCount", "MissingBaseColumns"], unmatched_rows)

    print(f"[OK] Built Badges export for {len(sheet_rows)} player-season rows")
    print("[INFO] Universe comes from playerlist.csv, RotationRole is enriched from bball_index_badges.csv, and live MIN still comes from general_traditional.csv.")
    print("[INFO] Deadeye now uses tight + very tight combined totals under the workbook's FG3M_verytight / FG3_PCT_verytight lanes.")
    print("[INFO] Putback Frequency% now uses nbarapm.csv as SelfORebPct * TeammateMissORebPerc / 100, while Putback_FGM / Alley_Oop_FGM only default missing rows to zero when that season has shot-type coverage.")
    print(f"[INFO] Raw badge-only metric series normalized in this build: {len(raw_metric_results)}")
    print(f"[OUT] Sheet -> {sheet_path}")
    print(f"[OUT] Badges -> {badge_only_path}")
    print(f"[OUT] Audit -> {audit_path}")
    print(f"[OUT] Unmatched -> {unmatched_path}")


if __name__ == "__main__":
    main()
