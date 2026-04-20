from __future__ import annotations

import statistics
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string

from build_cal_lane import (
    CalUniverseRow,
    apply_cal_normalization,
    build_source_index,
    canonical_id,
    compute_capped_z_score,
    detect_current_season,
    enrich_universe_rows,
    load_cal_universe,
    load_universe_csv,
    match_metric_row,
    normalize_name,
    parse_float,
    read_history_csv,
    resolve_details_csv_path,
    write_csv,
)
from build_finishing_layup import (
    ComponentSlot,
    build_player_contexts,
    interpolate_rating,
    percentile_inc,
)
from build_finishing_standing_dunk import standardize_rows, write_matrix_csv
from pull_nba_stats import EXPORT_DIR, HISTORY_DIR, MANUAL_DIR


SPEED_CURVE: Sequence[Tuple[float, float]] = (
    (0.04, 36.0),
    (0.08, 45.0),
    (0.14, 53.0),
    (0.22, 60.0),
    (0.32, 65.0),
    (0.44, 70.0),
    (0.54, 73.0),
    (0.64, 76.0),
    (0.72, 79.0),
    (0.80, 82.0),
    (0.88, 85.0),
    (0.94, 88.0),
    (0.98, 94.0),
    (1.00, 100.0),
)

AGILITY_CURVE: Sequence[Tuple[float, float]] = (
    (0.04, 33.0),
    (0.08, 40.0),
    (0.14, 48.0),
    (0.22, 55.0),
    (0.32, 61.04),
    (0.44, 68.0),
    (0.54, 72.0),
    (0.64, 75.0),
    (0.72, 78.0),
    (0.80, 82.0),
    (0.88, 85.0),
    (0.94, 89.0),
    (0.98, 96.0),
    (1.00, 100.0),
)

STRENGTH_CURVE: Sequence[Tuple[float, float]] = (
    (0.05, 35.0),
    (0.10, 42.0),
    (0.16, 50.0),
    (0.24, 56.0),
    (0.32, 61.04),
    (0.42, 67.0),
    (0.52, 71.0),
    (0.62, 75.0),
    (0.70, 77.0),
    (0.78, 80.0),
    (0.86, 84.0),
    (0.92, 88.0),
    (0.97, 93.0),
    (1.00, 100.0),
)

VERTICAL_CURVE: Sequence[Tuple[float, float]] = (
    (0.04, 42.0),
    (0.08, 49.0),
    (0.14, 55.0),
    (0.22, 60.0),
    (0.32, 66.0),
    (0.42, 72.0),
    (0.52, 73.0),
    (0.62, 75.0),
    (0.70, 78.0),
    (0.78, 81.0),
    (0.86, 85.0),
    (0.92, 88.24),
    (0.97, 94.0),
    (1.00, 100.0),
)

STAMINA_CURVE: Sequence[Tuple[float, float]] = (
    (0.25, 85.0),
    (0.50, 87.0),
    (0.75, 90.0),
    (0.90, 99.0),
    (1.00, 100.0),
)

HUSTLE_CURVE: Sequence[Tuple[float, float]] = (
    (0.03, 49.0),
    (0.06, 56.82),
    (0.12, 65.0),
    (0.20, 69.0),
    (0.30, 74.0),
    (0.40, 76.0),
    (0.50, 80.0),
    (0.60, 84.0),
    (0.70, 85.0),
    (0.78, 90.0),
    (0.85, 93.0),
    (0.92, 97.0),
    (0.97, 99.0),
    (1.00, 100.0),
)


SPEED_COMPONENTS: Sequence[ComponentSlot] = (
    ComponentSlot("attributes_Speed", "attributes_Speed"),
    ComponentSlot("Transition Shot Creation", "Transition Shot Creation"),
    ComponentSlot("Avg Speed Offense", "Avg Speed Offense"),
    ComponentSlot("Movement Speed Rating", "Movement Speed Rating"),
    ComponentSlot(
        "Offensive Transition Frequency Impact",
        "Offensive Transition Frequency Impact",
    ),
    ComponentSlot("Transition Frequency Impact", "Transition Frequency Impact"),
    ComponentSlot("TRANSITION_POSS", "TRANSITION_POSS"),
    ComponentSlot("AVG_SPEED", "AVG_SPEED"),
    ComponentSlot("INV_WEIGHT_HEIGHT_RATIO", "INV_WEIGHT_HEIGHT_RATIO"),
    ComponentSlot("INV_THREE_QUARTER_SPRINT", "INV_THREE_QUARTER_SPRINT"),
    ComponentSlot("AVG_SPEED_OFF", "AVG_SPEED_OFF"),
    ComponentSlot("Avg Speed Defense", "Avg Speed Defense"),
    ComponentSlot("TRANSITION_POSS_PCT", "TRANSITION_POSS_PCT"),
    ComponentSlot("PCT_PTS_FB", "PCT_PTS_FB"),
)

AGILITY_COMPONENTS: Sequence[ComponentSlot] = (
    ComponentSlot("attributes_Agility", "attributes_Agility"),
    ComponentSlot("SPEED", "SPEED"),
    ComponentSlot("Perimeter Isolation Defense", "Perimeter Isolation Defense"),
    ComponentSlot("Transition Shot Creation", "Transition Shot Creation"),
    ComponentSlot("INV_WEIGHT_HEIGHT_RATIO", "INV_WEIGHT_HEIGHT_RATIO"),
    ComponentSlot("INV_THREE_QUARTER_SPRINT", "INV_THREE_QUARTER_SPRINT"),
    ComponentSlot("INV_LANE_AGILITY_TIME", "INV_LANE_AGILITY_TIME"),
    ComponentSlot("INV_MODIFIED_LANE_AGILITY_TIME", "INV_MODIFIED_LANE_AGILITY_TIME"),
    ComponentSlot("TRANSITION_POSS_PCT", "TRANSITION_POSS_PCT"),
    ComponentSlot("TRANSITION_POSS", "TRANSITION_POSS"),
    ComponentSlot("Movement Speed Rating", "Movement Speed Rating"),
    ComponentSlot("Avg Speed Offense", "Avg Speed Offense"),
    ComponentSlot("Avg Speed Defense", "Avg Speed Defense"),
    ComponentSlot("AVG_SPEED", "AVG_SPEED"),
    ComponentSlot("OFFSCREEN_POSS", "OFFSCREEN_POSS"),
    ComponentSlot("OFFSCREEN_POSS_PCT", "OFFSCREEN_POSS_PCT"),
    ComponentSlot("CUT_POSS", "CUT_POSS"),
    ComponentSlot("CUT_POSS_PCT", "CUT_POSS_PCT"),
)

STRENGTH_COMPONENTS: Sequence[ComponentSlot] = (
    ComponentSlot("attributes_Strength", "attributes_Strength"),
    ComponentSlot("Post Control", "Post Control"),
    ComponentSlot("PLAYER_WEIGHT", "PLAYER_WEIGHT"),
    ComponentSlot("WEIGHT_HEIGHT_RATIO", "WEIGHT_HEIGHT_RATIO"),
    ComponentSlot("BOX_OUTS", "BOX_OUTS"),
    ComponentSlot("SCREEN_ASSISTS", "SCREEN_ASSISTS"),
    ComponentSlot("Stable And 1s Per 75", "Stable And 1s Per 75"),
    ComponentSlot("Post Defense", "Post Defense"),
    ComponentSlot("Screen Assists Per 75 Possessions", "Screen Assists Per 75 Possessions"),
    ComponentSlot("Finishing Talent", "Finishing Talent"),
    ComponentSlot("Screening Talent", "Screening Talent"),
)

VERTICAL_COMPONENTS: Sequence[ComponentSlot] = (
    ComponentSlot("attributes_Vertical", "attributes_Vertical"),
    ComponentSlot("Driving Dunk", "Driving Dunk"),
    ComponentSlot("jumpSubscore_average", "jumpSubscore_average"),
    ComponentSlot("jumpSubscore_max", "jumpSubscore_max"),
    ComponentSlot("VERTICAL_LEAP", "VERTICAL_LEAP"),
    ComponentSlot("Standing Dunk", "Standing Dunk"),
    ComponentSlot("Offensive Rebound", "Offensive Rebound"),
    ComponentSlot("Block", "Block"),
    ComponentSlot("jumpSubscore_total", "jumpSubscore_total"),
    ComponentSlot("INV_PLAYER_HEIGHT_INCHES", "INV_PLAYER_HEIGHT_INCHES"),
    ComponentSlot("INV_PLAYER_WEIGHT", "INV_PLAYER_WEIGHT"),
)

STAMINA_COMPONENTS: Sequence[ComponentSlot] = (
    ComponentSlot("MIN", "MIN"),
    ComponentSlot("attributes_Stamina", "attributes_Stamina"),
    ComponentSlot("INV_AGE", "INV_AGE"),
    ComponentSlot("Usage Rate", "Usage Rate"),
    ComponentSlot("True Usage", "True Usage"),
    ComponentSlot("Stable True Usage%", "Stable True Usage%"),
    ComponentSlot("Games Played", "Games Played"),
    ComponentSlot("TOTAL MIN", "TOTAL MIN"),
)

HUSTLE_COMPONENTS: Sequence[ComponentSlot] = (
    ComponentSlot("Defensive Playmaking", "Defensive Playmaking"),
    ComponentSlot("DEFLECTIONS", "DEFLECTIONS"),
    ComponentSlot("LOOSE_BALLS_RECOVERED", "LOOSE_BALLS_RECOVERED"),
    ComponentSlot("CONTESTED_SHOTS", "CONTESTED_SHOTS"),
    ComponentSlot("BOX_OUTS", "BOX_OUTS"),
)


WORKBOOK_COLUMNS = [
    "NBA ID",
    "Season",
    "Player",
    "SPEED",
    "",
    "SPEED",
    *[slot.header for slot in SPEED_COMPONENTS],
    "Agility",
    "",
    "Agility",
    *[slot.header for slot in AGILITY_COMPONENTS],
    "STR",
    "",
    "STR",
    *[slot.header for slot in STRENGTH_COMPONENTS],
    "VERT",
    "",
    "VERT",
    *[slot.header for slot in VERTICAL_COMPONENTS],
    "STAM",
    "",
    "STAM",
    *[slot.header for slot in STAMINA_COMPONENTS],
    "HSTL",
    "",
    "HSTL",
    *[slot.header for slot in HUSTLE_COMPONENTS],
]


@dataclass(frozen=True)
class ScoreGroup:
    aliases: Sequence[str]
    weight: float
    optional_if_missing: bool = False


@dataclass(frozen=True)
class SectionSpec:
    name: str
    rating_header: str
    rating_only_header: str
    component_slots: Sequence[ComponentSlot]
    score_groups: Sequence[ScoreGroup]
    curve: Sequence[Tuple[float, float]]
    curve_start_rating: float = 25.0
    missing_score: float = -1.0


@dataclass
class MetricResult:
    raw_values: List[Optional[float]]
    normalized_values: List[Optional[float]]
    matched_by: List[str]
    mean_value: Optional[float]
    stdev_value: Optional[float]
    source_note: str


SPEED_SECTION = SectionSpec(
    name="Speed",
    rating_header="Speed Rating",
    rating_only_header="Speed Rating",
    component_slots=SPEED_COMPONENTS,
    score_groups=(
        ScoreGroup(tuple(slot.metric_alias for slot in SPEED_COMPONENTS[:11]), 0.85),
        ScoreGroup(tuple(slot.metric_alias for slot in SPEED_COMPONENTS[11:]), 0.15, True),
    ),
    curve=SPEED_CURVE,
)

AGILITY_SECTION = SectionSpec(
    name="Agility",
    rating_header="Agility Rating",
    rating_only_header="Agility Rating",
    component_slots=AGILITY_COMPONENTS,
    score_groups=(
        ScoreGroup(tuple(slot.metric_alias for slot in AGILITY_COMPONENTS[:4]), 0.60),
        ScoreGroup(tuple(slot.metric_alias for slot in AGILITY_COMPONENTS[4:14]), 0.35),
        ScoreGroup(tuple(slot.metric_alias for slot in AGILITY_COMPONENTS[14:]), 0.05, True),
    ),
    curve=AGILITY_CURVE,
)

STRENGTH_SECTION = SectionSpec(
    name="Strength",
    rating_header="Strength Rating",
    rating_only_header="Strength Rating",
    component_slots=STRENGTH_COMPONENTS,
    score_groups=(
        ScoreGroup(tuple(slot.metric_alias for slot in STRENGTH_COMPONENTS[:10]), 0.95),
        ScoreGroup((STRENGTH_COMPONENTS[10].metric_alias,), 0.05),
    ),
    curve=STRENGTH_CURVE,
)

VERTICAL_SECTION = SectionSpec(
    name="Vertical",
    rating_header="Vertical Rating",
    rating_only_header="Vertical Rating",
    component_slots=VERTICAL_COMPONENTS,
    score_groups=(
        ScoreGroup(tuple(slot.metric_alias for slot in VERTICAL_COMPONENTS[:5]), 0.70),
        ScoreGroup(tuple(slot.metric_alias for slot in VERTICAL_COMPONENTS[5:]), 0.30),
    ),
    curve=VERTICAL_CURVE,
)

STAMINA_SECTION = SectionSpec(
    name="Stamina",
    rating_header="Stamina Rating",
    rating_only_header="Stamina Rating",
    component_slots=STAMINA_COMPONENTS,
    score_groups=(
        ScoreGroup((STAMINA_COMPONENTS[0].metric_alias,), 0.70),
        ScoreGroup(tuple(slot.metric_alias for slot in STAMINA_COMPONENTS[1:6]), 0.25),
        ScoreGroup(tuple(slot.metric_alias for slot in STAMINA_COMPONENTS[6:]), 0.05),
    ),
    curve=STAMINA_CURVE,
    curve_start_rating=60.0,
)

HUSTLE_SECTION = SectionSpec(
    name="Hustle",
    rating_header="Hustle Rating",
    rating_only_header="Hustle Rating",
    component_slots=HUSTLE_COMPONENTS,
    score_groups=(ScoreGroup(tuple(slot.metric_alias for slot in HUSTLE_COMPONENTS), 1.0),),
    curve=HUSTLE_CURVE,
    missing_score=0.0,
)

SECTIONS: Sequence[SectionSpec] = (
    SPEED_SECTION,
    AGILITY_SECTION,
    STRENGTH_SECTION,
    VERTICAL_SECTION,
    STAMINA_SECTION,
    HUSTLE_SECTION,
)

RATING_ONLY_HEADERS = [
    "NBA_ID",
    "Season",
    "Player",
    "Speed Rating",
    "Agility Rating",
    "Strength Rating",
    "Vertical Rating",
    "Stamina Rating",
    "Hustle Rating",
]

B_BALL_SOURCE_NOTE = "bball_index_physical.csv"

DIRECT_SCORE_ALIASES = {
    "Post Control",
    "Driving Dunk",
    "Standing Dunk",
    "Offensive Rebound",
    "Block",
}

RAW_Z_ONLY_ALIASES = {
    "PLAYER_HEIGHT_INCHES",
    "PLAYER_WEIGHT",
    "AGE",
    "INV_PLAYER_HEIGHT_INCHES",
    "INV_PLAYER_WEIGHT",
    "INV_AGE",
    "WEIGHT_HEIGHT_RATIO",
    "INV_WEIGHT_HEIGHT_RATIO",
}

METRIC_SOURCE_NOTES: Dict[str, str] = {
    "attributes_Speed": "Upgrade/players.csv -> attributes_Speed",
    "attributes_Agility": "Upgrade/players.csv -> attributes_Agility",
    "attributes_Strength": "Upgrade/players.csv -> attributes_Strength",
    "attributes_Vertical": "Upgrade/players.csv -> attributes_Vertical",
    "attributes_Stamina": "Upgrade/players.csv -> attributes_Stamina",
    "attributes_Hustle": "Upgrade/players.csv -> attributes_Hustle",
    "Transition Shot Creation": B_BALL_SOURCE_NOTE,
    "Avg Speed Offense": f"{B_BALL_SOURCE_NOTE}, else tracking_speed.csv -> AVG_SPEED_OFF proxy",
    "Avg Speed Defense": f"{B_BALL_SOURCE_NOTE}, else tracking_speed.csv -> AVG_SPEED_DEF proxy",
    "Movement Speed Rating": B_BALL_SOURCE_NOTE,
    "Offensive Transition Frequency Impact": B_BALL_SOURCE_NOTE,
    "Transition Frequency Impact": B_BALL_SOURCE_NOTE,
    "TRANSITION_POSS": "playtype_transition.csv -> POSS",
    "TRANSITION_POSS_PCT": "playtype_transition.csv -> POSS_PCT",
    "AVG_SPEED": "tracking_speed.csv -> AVG_SPEED",
    "AVG_SPEED_OFF": "tracking_speed.csv -> AVG_SPEED_OFF",
    "INV_WEIGHT_HEIGHT_RATIO": "derived from bios.csv -> -(PLAYER_WEIGHT / PLAYER_HEIGHT_INCHES)",
    "WEIGHT_HEIGHT_RATIO": "derived from bios.csv -> PLAYER_WEIGHT / PLAYER_HEIGHT_INCHES",
    "INV_THREE_QUARTER_SPRINT": "draft.csv -> -THREE_QUARTER_SPRINT",
    "Perimeter Isolation Defense": B_BALL_SOURCE_NOTE,
    "INV_LANE_AGILITY_TIME": "draft.csv -> -LANE_AGILITY_TIME",
    "INV_MODIFIED_LANE_AGILITY_TIME": "draft.csv -> -MODIFIED_LANE_AGILITY_TIME",
    "OFFSCREEN_POSS": "playtype_offscreen.csv -> POSS",
    "OFFSCREEN_POSS_PCT": "playtype_offscreen.csv -> POSS_PCT",
    "CUT_POSS": "playtype_cut.csv -> POSS",
    "CUT_POSS_PCT": "playtype_cut.csv -> POSS_PCT",
    "PCT_PTS_FB": "general_scoring.csv -> PCT_PTS_FB",
    "Post Control": "finishing_post_control_audit.csv -> Post Control direct score",
    "PLAYER_WEIGHT": "bios.csv -> PLAYER_WEIGHT",
    "BOX_OUTS": "hustle.csv -> BOX_OUTS",
    "SCREEN_ASSISTS": "hustle.csv -> SCREEN_ASSISTS",
    "Stable And 1s Per 75": B_BALL_SOURCE_NOTE,
    "Post Defense": B_BALL_SOURCE_NOTE,
    "Screen Assists Per 75 Possessions": B_BALL_SOURCE_NOTE,
    "Finishing Talent": B_BALL_SOURCE_NOTE,
    "Screening Talent": B_BALL_SOURCE_NOTE,
    "Driving Dunk": "finishing_driving_dunk_audit.csv -> Driving Dunk direct score",
    "jumpSubscore_average": "overall_dunk_stats.csv -> jumpSubscore_average",
    "jumpSubscore_max": "overall_dunk_stats.csv -> jumpSubscore_max",
    "VERTICAL_LEAP": "draft.csv -> MAX_VERTICAL_LEAP, else STANDING_VERTICAL_LEAP fallback",
    "Standing Dunk": "finishing_standing_dunk_audit.csv -> Standing Dunk direct score",
    "Offensive Rebound": "2k26_rebound.xlsx -> Offensive Rebound score direct",
    "Block": "2k26_defense.xlsx -> Block score direct",
    "jumpSubscore_total": "overall_dunk_stats.csv -> jumpSubscore_total",
    "INV_PLAYER_HEIGHT_INCHES": "bios.csv -> -PLAYER_HEIGHT_INCHES",
    "INV_PLAYER_WEIGHT": "bios.csv -> -PLAYER_WEIGHT",
    "MIN": "general_traditional.csv -> MIN",
    "INV_AGE": "bios.csv -> -AGE",
    "Usage Rate": f"{B_BALL_SOURCE_NOTE}, else general_advanced.csv -> USG_PCT proxy",
    "True Usage": B_BALL_SOURCE_NOTE,
    "Stable True Usage%": B_BALL_SOURCE_NOTE,
    "Games Played": "general_traditional.csv -> GP",
    "TOTAL MIN": "derived from general_traditional.csv -> MIN * GP",
    "Defensive Playmaking": B_BALL_SOURCE_NOTE,
    "DEFLECTIONS": "hustle.csv -> DEFLECTIONS",
    "LOOSE_BALLS_RECOVERED": "hustle.csv -> LOOSE_BALLS_RECOVERED",
    "CONTESTED_SHOTS": "hustle.csv -> CONTESTED_SHOTS",
    "SPEED": "derived from the Physical Speed section score direct",
}


def parse_args():
    import argparse

    parser = argparse.ArgumentParser(description="Build the combined Physical ratings export.")
    parser.add_argument(
        "--universe-csv",
        default=str(MANUAL_DIR / "playerlist.csv"),
        help="Season/player universe CSV. Defaults to stats/manual/playerlist.csv.",
    )
    parser.add_argument(
        "--workbook",
        default=str(MANUAL_DIR / "2k26_Temp_for_codex.xlsx"),
        help="Workbook fallback used for Cal role/minutes.",
    )
    parser.add_argument("--sheet", default="Cal", help="Workbook sheet used for role/minute fallback.")
    parser.add_argument(
        "--details-csv",
        default="",
        help=(
            "Optional role/minutes detail CSV. "
            "Pass stats/manual/player_universe.csv explicitly if you want that enrichment."
        ),
    )
    parser.add_argument(
        "--minutes-source",
        default=str(HISTORY_DIR / "general_traditional.csv"),
        help="History CSV used to refresh MIN and GP in real time.",
    )
    parser.add_argument("--minutes-column", default="MIN")
    parser.add_argument("--minutes-games-column", default="GP")
    parser.add_argument("--current-season", default="")
    parser.add_argument("--current-season-min-threshold", type=float, default=200.0)
    parser.add_argument("--standard-min-threshold", type=float, default=1000.0)
    parser.add_argument("--allow-id-fallback", action="store_true")
    parser.add_argument(
        "--bball-physical-source",
        default=str(HISTORY_DIR / "bball_index_physical.csv"),
    )
    parser.add_argument("--general-scoring-source", default=str(HISTORY_DIR / "general_scoring.csv"))
    parser.add_argument("--general-advanced-source", default=str(HISTORY_DIR / "general_advanced.csv"))
    parser.add_argument("--bios-source", default=str(HISTORY_DIR / "bios.csv"))
    parser.add_argument("--tracking-speed-source", default=str(HISTORY_DIR / "tracking_speed.csv"))
    parser.add_argument(
        "--playtype-transition-source",
        default=str(HISTORY_DIR / "playtype_transition.csv"),
    )
    parser.add_argument(
        "--playtype-offscreen-source",
        default=str(HISTORY_DIR / "playtype_offscreen.csv"),
    )
    parser.add_argument("--playtype-cut-source", default=str(HISTORY_DIR / "playtype_cut.csv"))
    parser.add_argument("--hustle-source", default=str(HISTORY_DIR / "hustle.csv"))
    parser.add_argument("--draft-source", default=str(HISTORY_DIR / "draft.csv"))
    parser.add_argument(
        "--overall-dunk-source",
        default=str(HISTORY_DIR / "overall_dunk_stats.csv"),
    )
    parser.add_argument("--players-source", default=str(Path("Upgrade") / "players.csv"))
    parser.add_argument(
        "--post-control-audit-source",
        default=str(EXPORT_DIR / "finishing_post_control_audit.csv"),
    )
    parser.add_argument(
        "--driving-dunk-audit-source",
        default=str(EXPORT_DIR / "finishing_driving_dunk_audit.csv"),
    )
    parser.add_argument(
        "--standing-dunk-audit-source",
        default=str(EXPORT_DIR / "finishing_standing_dunk_audit.csv"),
    )
    parser.add_argument(
        "--rebound-workbook-source",
        default=str(MANUAL_DIR / "2k26_rebound.xlsx"),
    )
    parser.add_argument(
        "--defense-workbook-source",
        default=str(MANUAL_DIR / "2k26_defense.xlsx"),
    )
    parser.add_argument("--output-prefix", default="physical_all")
    return parser.parse_args()


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() == "nan":
        return ""
    return text


def load_bball_detail_rows(rows: Sequence[Dict[str, str]]) -> List[CalUniverseRow]:
    detail_rows: List[CalUniverseRow] = []
    for row in rows:
        season = clean_text(row.get("Season", ""))
        player = clean_text(row.get("Player", ""))
        if not season or not player:
            continue
        detail_rows.append(
            CalUniverseRow(
                nba_id=canonical_id(row.get("NBA_ID", row.get("NBA ID", ""))),
                season=season,
                player=player,
                rotation_role=clean_text(row.get("Rotation Role", "")),
                minutes=parse_float(row.get("Minutes", "")),
            )
        )
    return detail_rows


def load_bball_index_rows(path: Path) -> List[Dict[str, object]]:
    raw_rows = read_history_csv(path)
    metric_columns = {
        "Transition Shot Creation",
        "Avg Speed Offense",
        "Avg Speed Defense",
        "Movement Speed Rating",
        "Offensive Transition Frequency Impact",
        "Transition Frequency Impact",
        "Perimeter Isolation Defense",
        "Stable And 1s Per 75",
        "Post Defense",
        "Screen Assists Per 75 Possessions",
        "Finishing Talent",
        "Screening Talent",
        "Usage Rate",
        "True Usage",
        "Stable True Usage%",
        "Defensive Playmaking",
    }

    standardized: List[Dict[str, object]] = []
    for row in raw_rows:
        season = clean_text(row.get("Season", ""))
        player = clean_text(row.get("Player", ""))
        if not season or not player:
            continue

        merged: Dict[str, object] = {
            "Season": season,
            "PLAYER_ID": canonical_id(row.get("NBA_ID", row.get("NBA ID", ""))),
            "PLAYER_NAME": player,
            "TEAM_ABBREVIATION": clean_text(row.get("Team(s)", row.get("Team", ""))),
        }
        for metric in metric_columns:
            merged[metric] = parse_float(row.get(metric, ""))
        standardized.append(merged)

    return standardized


def load_static_players_rows(path: Path) -> List[Dict[str, object]]:
    raw_rows = read_history_csv(path)
    standardized: List[Dict[str, object]] = []

    for row in raw_rows:
        player_name = " ".join(
            part
            for part in [clean_text(row.get("firstName", "")), clean_text(row.get("lastName", ""))]
            if part
        ).strip()
        if not player_name:
            continue

        standardized.append(
            {
                "PLAYER_ID": canonical_id(row.get("player_id", "")),
                "PLAYER_NAME": player_name,
                "attributes_Speed": parse_float(row.get("attributes_Speed", "")),
                "attributes_Agility": parse_float(row.get("attributes_Agility", "")),
                "attributes_Strength": parse_float(row.get("attributes_Strength", "")),
                "attributes_Vertical": parse_float(row.get("attributes_Vertical", "")),
                "attributes_Stamina": parse_float(row.get("attributes_Stamina", "")),
                "attributes_Hustle": parse_float(row.get("attributes_Hustle", "")),
            }
        )

    return standardized


def load_static_draft_rows(path: Path) -> List[Dict[str, object]]:
    raw_rows = read_history_csv(path)
    standardized: List[Dict[str, object]] = []

    for row in raw_rows:
        player = clean_text(row.get("PLAYER_NAME", ""))
        if not player:
            continue

        max_vertical = parse_float(row.get("MAX_VERTICAL_LEAP", ""))
        standing_vertical = parse_float(row.get("STANDING_VERTICAL_LEAP", ""))
        vertical_leap = max_vertical if max_vertical is not None else standing_vertical

        standardized.append(
            {
                "PLAYER_ID": canonical_id(row.get("PLAYER_ID", "")),
                "PLAYER_NAME": player,
                "THREE_QUARTER_SPRINT": parse_float(row.get("THREE_QUARTER_SPRINT", "")),
                "LANE_AGILITY_TIME": parse_float(row.get("LANE_AGILITY_TIME", "")),
                "MODIFIED_LANE_AGILITY_TIME": parse_float(
                    row.get("MODIFIED_LANE_AGILITY_TIME", "")
                ),
                "VERTICAL_LEAP": vertical_leap,
            }
        )

    return standardized


def load_excel_score_rows(
    path: Path,
    score_column_letter: str,
    score_alias: str,
) -> List[Dict[str, object]]:
    score_index = column_index_from_string(score_column_letter) - 1
    workbook = load_workbook(path, data_only=True, read_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    standardized: List[Dict[str, object]] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        nba_id = canonical_id(row[0] if len(row) > 0 else "")
        season = clean_text(row[1] if len(row) > 1 else "")
        player = clean_text(row[2] if len(row) > 2 else "")
        if not season or not player:
            continue
        score_value = parse_float(row[score_index] if len(row) > score_index else "")
        standardized.append(
            {
                "Season": season,
                "PLAYER_ID": nba_id,
                "PLAYER_NAME": player,
                score_alias: score_value,
            }
        )

    return standardized


def load_score_csv_rows(path: Path, score_column: str, score_alias: str) -> List[Dict[str, object]]:
    raw_rows = read_history_csv(path)
    standardized: List[Dict[str, object]] = []
    for row in raw_rows:
        season = clean_text(row.get("Season", ""))
        player = clean_text(row.get("Player", row.get("PLAYER_NAME", "")))
        if not season or not player:
            continue
        standardized.append(
            {
                "Season": season,
                "PLAYER_ID": canonical_id(row.get("NBA_ID", row.get("PLAYER_ID", ""))),
                "PLAYER_NAME": player,
                score_alias: parse_float(row.get(score_column, "")),
            }
        )
    return standardized


def build_static_source_index(
    rows: Sequence[Dict[str, object]],
) -> Tuple[Dict[str, Dict[str, object]], Dict[str, Dict[str, object]]]:
    by_id: Dict[str, Dict[str, object]] = {}
    by_name: Dict[str, Dict[str, object]] = {}

    for row in rows:
        player_id = canonical_id(row.get("PLAYER_ID", ""))
        player_name = normalize_name(row.get("PLAYER_NAME", ""))
        if player_id:
            by_id[player_id] = row
        if player_name:
            by_name[player_name] = row

    return by_id, by_name


def get_row_key(row: CalUniverseRow) -> Tuple[str, str]:
    return row.season, normalize_name(row.player)


def set_base_value(
    base_row: Dict[str, object],
    alias: str,
    value: object,
    matched_by: str,
    *,
    overwrite: bool = False,
) -> None:
    if value is None:
        return
    text = str(value).strip()
    if not text:
        return
    if not overwrite and str(base_row.get(alias, "")).strip():
        return
    base_row[alias] = value
    base_row[f"__matched__{alias}"] = matched_by


def merge_metric_columns(
    universe: Sequence[CalUniverseRow],
    base_rows_by_key: Dict[Tuple[str, str], Dict[str, object]],
    source_rows: Sequence[Dict[str, object]],
    column_map: Dict[str, str],
    allow_id_fallback: bool,
    *,
    overwrite: bool = False,
) -> None:
    if not source_rows:
        return

    by_id, by_name = build_source_index(list(source_rows))
    for universe_row in universe:
        source_row, matched_by = match_metric_row(
            universe_row,
            by_id,
            by_name,
            allow_id_fallback=allow_id_fallback,
        )
        if source_row is None:
            continue

        base_row = base_rows_by_key[get_row_key(universe_row)]
        if not str(base_row.get("TEAM_ABBREVIATION", "")).strip():
            team_value = str(source_row.get("TEAM_ABBREVIATION", "")).strip()
            if team_value:
                base_row["TEAM_ABBREVIATION"] = team_value

        for alias, source_column in column_map.items():
            set_base_value(
                base_row,
                alias,
                source_row.get(source_column, None),
                matched_by,
                overwrite=overwrite,
            )


def merge_static_metric_columns(
    universe: Sequence[CalUniverseRow],
    base_rows_by_key: Dict[Tuple[str, str], Dict[str, object]],
    source_rows: Sequence[Dict[str, object]],
    column_map: Dict[str, str],
    *,
    overwrite: bool = False,
) -> None:
    if not source_rows:
        return

    by_id, by_name = build_static_source_index(source_rows)
    for universe_row in universe:
        source_row = None
        matched_by = ""

        if universe_row.nba_id and universe_row.nba_id in by_id:
            source_row = by_id[universe_row.nba_id]
            matched_by = "static-id"
        else:
            name_key = normalize_name(universe_row.player)
            if name_key in by_name:
                source_row = by_name[name_key]
                matched_by = "static-player"

        if source_row is None:
            continue

        base_row = base_rows_by_key[get_row_key(universe_row)]
        for alias, source_column in column_map.items():
            set_base_value(
                base_row,
                alias,
                source_row.get(source_column, None),
                matched_by,
                overwrite=overwrite,
            )


def derive_physical_metrics(base_rows: Iterable[Dict[str, object]]) -> None:
    for base_row in base_rows:
        height = parse_float(base_row.get("PLAYER_HEIGHT_INCHES", ""))
        weight = parse_float(base_row.get("PLAYER_WEIGHT", ""))
        age = parse_float(base_row.get("AGE", ""))
        minutes = parse_float(base_row.get("MIN", ""))
        games_played = parse_float(base_row.get("GP", "")) or parse_float(
            base_row.get("Games Played", "")
        )
        sprint = parse_float(base_row.get("THREE_QUARTER_SPRINT", ""))
        lane_agility = parse_float(base_row.get("LANE_AGILITY_TIME", ""))
        modified_lane_agility = parse_float(base_row.get("MODIFIED_LANE_AGILITY_TIME", ""))

        if height not in (None, 0) and weight is not None:
            ratio = weight / height
            base_row["WEIGHT_HEIGHT_RATIO"] = ratio
            base_row["__matched__WEIGHT_HEIGHT_RATIO"] = "derived"
            base_row["INV_WEIGHT_HEIGHT_RATIO"] = -ratio
            base_row["__matched__INV_WEIGHT_HEIGHT_RATIO"] = "derived"

        if height is not None:
            base_row["INV_PLAYER_HEIGHT_INCHES"] = -height
            base_row["__matched__INV_PLAYER_HEIGHT_INCHES"] = base_row.get(
                "__matched__PLAYER_HEIGHT_INCHES",
                "",
            )

        if weight is not None:
            base_row["INV_PLAYER_WEIGHT"] = -weight
            base_row["__matched__INV_PLAYER_WEIGHT"] = base_row.get(
                "__matched__PLAYER_WEIGHT",
                "",
            )

        if age is not None:
            base_row["INV_AGE"] = -age
            base_row["__matched__INV_AGE"] = base_row.get("__matched__AGE", "")

        if minutes is not None and games_played is not None:
            base_row["TOTAL_MIN"] = minutes * games_played
            base_row["__matched__TOTAL_MIN"] = "derived"

        if games_played is not None:
            base_row["Games Played"] = games_played
            base_row["__matched__Games Played"] = base_row.get("__matched__GP", "")

        if sprint is not None:
            base_row["INV_THREE_QUARTER_SPRINT"] = -sprint
            base_row["__matched__INV_THREE_QUARTER_SPRINT"] = base_row.get(
                "__matched__THREE_QUARTER_SPRINT",
                "",
            )

        if lane_agility is not None:
            base_row["INV_LANE_AGILITY_TIME"] = -lane_agility
            base_row["__matched__INV_LANE_AGILITY_TIME"] = base_row.get(
                "__matched__LANE_AGILITY_TIME",
                "",
            )

        if modified_lane_agility is not None:
            base_row["INV_MODIFIED_LANE_AGILITY_TIME"] = -modified_lane_agility
            base_row["__matched__INV_MODIFIED_LANE_AGILITY_TIME"] = base_row.get(
                "__matched__MODIFIED_LANE_AGILITY_TIME",
                "",
            )

        if not str(base_row.get("Avg Speed Offense", "")).strip():
            avg_speed_off = parse_float(base_row.get("AVG_SPEED_OFF", ""))
            if avg_speed_off is not None:
                base_row["Avg Speed Offense"] = avg_speed_off
                base_row["__matched__Avg Speed Offense"] = "derived-proxy"

        if not str(base_row.get("Avg Speed Defense", "")).strip():
            avg_speed_def = parse_float(base_row.get("AVG_SPEED_DEF", ""))
            if avg_speed_def is not None:
                base_row["Avg Speed Defense"] = avg_speed_def
                base_row["__matched__Avg Speed Defense"] = "derived-proxy"

        if not str(base_row.get("Usage Rate", "")).strip():
            usg_pct = parse_float(base_row.get("USG_PCT", ""))
            if usg_pct is not None:
                base_row["Usage Rate"] = usg_pct
                base_row["__matched__Usage Rate"] = "derived-proxy"


def build_metric_result(
    contexts,
    base_rows_by_key: Dict[Tuple[str, str], Dict[str, object]],
    metric_alias: str,
    current_season: str,
    current_season_min_threshold: float,
    standard_min_threshold: float,
) -> MetricResult:
    raw_values: List[Optional[float]] = []
    normalized_values: List[Optional[float]] = []
    matched_by_values: List[str] = []
    matched_numeric_values: List[float] = []

    for context in contexts:
        base_row = base_rows_by_key[get_row_key(context.universe_row)]
        raw_value = parse_float(base_row.get(metric_alias, ""))
        matched_by = str(base_row.get(f"__matched__{metric_alias}", "")).strip()
        raw_values.append(raw_value)
        matched_by_values.append(matched_by)
        if raw_value is not None:
            matched_numeric_values.append(raw_value)

    mean_value = None if not matched_numeric_values else statistics.mean(matched_numeric_values)
    stdev_value = (
        statistics.stdev(matched_numeric_values)
        if len(matched_numeric_values) >= 2
        else None
    )

    if metric_alias in DIRECT_SCORE_ALIASES:
        normalized_values = list(raw_values)
        return MetricResult(
            raw_values=raw_values,
            normalized_values=normalized_values,
            matched_by=matched_by_values,
            mean_value=mean_value,
            stdev_value=stdev_value,
            source_note=METRIC_SOURCE_NOTES.get(metric_alias, "merged physical sources"),
        )

    for index, context in enumerate(contexts):
        raw_value = raw_values[index]
        if raw_value is None or mean_value is None or stdev_value in (None, 0):
            normalized_values.append(None)
            continue

        raw_z = compute_capped_z_score(raw_value, mean_value, stdev_value)
        if metric_alias in RAW_Z_ONLY_ALIASES:
            normalized_values.append(raw_z)
        else:
            normalized_values.append(
                apply_cal_normalization(
                    season=context.universe_row.season,
                    rotation_role=context.universe_row.rotation_role,
                    minutes=context.effective_minutes,
                    raw_z=raw_z,
                    current_season=current_season,
                    current_season_min_threshold=current_season_min_threshold,
                    standard_min_threshold=standard_min_threshold,
                )
            )

    return MetricResult(
        raw_values=raw_values,
        normalized_values=normalized_values,
        matched_by=matched_by_values,
        mean_value=mean_value,
        stdev_value=stdev_value,
        source_note=METRIC_SOURCE_NOTES.get(metric_alias, "merged physical sources"),
    )


def average_numeric(values: Sequence[Optional[float]]) -> Optional[float]:
    numeric = [value for value in values if value is not None]
    if not numeric:
        return None
    return statistics.mean(numeric)


def compute_section_score(
    section: SectionSpec,
    component_values_by_alias: Dict[str, Optional[float]],
) -> float:
    score_total = 0.0
    for group in section.score_groups:
        group_average = average_numeric([component_values_by_alias.get(alias) for alias in group.aliases])
        if group_average is None:
            if group.optional_if_missing:
                continue
            return section.missing_score
        score_total += group_average * group.weight
    return score_total


def compute_piecewise_rating(
    value: float,
    population: Sequence[float],
    curve: Sequence[Tuple[float, float]],
    curve_start_rating: float,
) -> float:
    minimum = min(population)
    low_x = minimum
    low_y = curve_start_rating

    for percentile, high_y in curve:
        high_x = percentile_inc(population, percentile)
        if value <= high_x or percentile >= 1.0:
            return interpolate_rating(value, low_x, high_x, low_y, high_y)
        low_x = high_x
        low_y = high_y

    return curve[-1][1]


def build_section_outputs(
    section: SectionSpec,
    contexts,
    metric_results: Dict[str, MetricResult],
    extra_normalized_by_alias: Optional[Dict[str, List[Optional[float]]]] = None,
) -> Tuple[List[float], List[float], List[float], List[List[Optional[float]]]]:
    extra_normalized_by_alias = extra_normalized_by_alias or {}

    scores: List[float] = []
    component_rows: List[List[Optional[float]]] = []

    for index in range(len(contexts)):
        component_values_by_alias: Dict[str, Optional[float]] = {}
        component_row: List[Optional[float]] = []
        for slot in section.component_slots:
            if slot.metric_alias in extra_normalized_by_alias:
                value = extra_normalized_by_alias[slot.metric_alias][index]
            else:
                value = metric_results[slot.metric_alias].normalized_values[index]
            component_values_by_alias[slot.metric_alias] = value
            component_row.append(value)

        scores.append(compute_section_score(section, component_values_by_alias))
        component_rows.append(component_row)

    if len(scores) < 2:
        raise SystemExit(f"Not enough {section.name} scores to compute aggregate z-scores.")

    median_value = statistics.median(scores)
    stdev_value = statistics.stdev(scores)
    if stdev_value == 0:
        raise SystemExit(f"{section.name} scores have zero variance; cannot compute ratings.")

    aggregate_z_scores = [(value - median_value) / stdev_value for value in scores]
    ratings = [
        compute_piecewise_rating(
            value,
            aggregate_z_scores,
            section.curve,
            section.curve_start_rating,
        )
        for value in aggregate_z_scores
    ]
    return scores, aggregate_z_scores, ratings, component_rows


def main() -> None:
    args = parse_args()

    workbook_path = Path(args.workbook)
    universe_path = Path(args.universe_csv)
    details_path = resolve_details_csv_path(args.details_csv, universe_path)
    minutes_source_path = Path(args.minutes_source)
    bball_physical_source_path = Path(args.bball_physical_source)
    general_scoring_source_path = Path(args.general_scoring_source)
    general_advanced_source_path = Path(args.general_advanced_source)
    bios_source_path = Path(args.bios_source)
    tracking_speed_source_path = Path(args.tracking_speed_source)
    playtype_transition_source_path = Path(args.playtype_transition_source)
    playtype_offscreen_source_path = Path(args.playtype_offscreen_source)
    playtype_cut_source_path = Path(args.playtype_cut_source)
    hustle_source_path = Path(args.hustle_source)
    draft_source_path = Path(args.draft_source)
    overall_dunk_source_path = Path(args.overall_dunk_source)
    players_source_path = Path(args.players_source)
    post_control_audit_source_path = Path(args.post_control_audit_source)
    driving_dunk_audit_source_path = Path(args.driving_dunk_audit_source)
    standing_dunk_audit_source_path = Path(args.standing_dunk_audit_source)
    rebound_workbook_source_path = Path(args.rebound_workbook_source)
    defense_workbook_source_path = Path(args.defense_workbook_source)

    required_paths = (
        minutes_source_path,
        bball_physical_source_path,
        general_scoring_source_path,
        general_advanced_source_path,
        bios_source_path,
        tracking_speed_source_path,
        playtype_transition_source_path,
        playtype_offscreen_source_path,
        playtype_cut_source_path,
        hustle_source_path,
        draft_source_path,
        overall_dunk_source_path,
        players_source_path,
        post_control_audit_source_path,
        driving_dunk_audit_source_path,
        standing_dunk_audit_source_path,
        rebound_workbook_source_path,
        defense_workbook_source_path,
    )
    missing_paths = [str(path) for path in required_paths if not path.exists()]
    if missing_paths:
        raise SystemExit("Missing required source files:\n- " + "\n- ".join(missing_paths))
    if details_path and not details_path.exists():
        raise SystemExit(f"Details CSV not found: {details_path}")

    minutes_rows = read_history_csv(minutes_source_path)
    bball_physical_raw_rows = read_history_csv(bball_physical_source_path)
    bball_physical_rows = load_bball_index_rows(bball_physical_source_path)
    general_scoring_rows = standardize_rows(
        read_history_csv(general_scoring_source_path),
        season_column="Season",
        player_column="PLAYER_NAME",
    )
    general_advanced_rows = standardize_rows(
        read_history_csv(general_advanced_source_path),
        season_column="Season",
        player_column="PLAYER_NAME",
    )
    bios_rows = standardize_rows(
        read_history_csv(bios_source_path),
        season_column="Season",
        player_column="PLAYER_NAME",
    )
    tracking_speed_rows = standardize_rows(
        read_history_csv(tracking_speed_source_path),
        season_column="Season",
        player_column="PLAYER_NAME",
    )
    playtype_transition_rows = standardize_rows(
        read_history_csv(playtype_transition_source_path),
        season_column="Season",
        player_column="PLAYER_NAME",
    )
    playtype_offscreen_rows = standardize_rows(
        read_history_csv(playtype_offscreen_source_path),
        season_column="Season",
        player_column="PLAYER_NAME",
    )
    playtype_cut_rows = standardize_rows(
        read_history_csv(playtype_cut_source_path),
        season_column="Season",
        player_column="PLAYER_NAME",
    )
    hustle_rows = standardize_rows(
        read_history_csv(hustle_source_path),
        season_column="Season",
        player_column="PLAYER_NAME",
    )
    overall_dunk_rows = standardize_rows(
        read_history_csv(overall_dunk_source_path),
        season_column="Season",
        player_column="playerName",
    )
    players_rows = load_static_players_rows(players_source_path)
    draft_rows = load_static_draft_rows(draft_source_path)
    post_control_rows = load_score_csv_rows(
        post_control_audit_source_path,
        "Post Control",
        "Post Control",
    )
    driving_dunk_rows = load_score_csv_rows(
        driving_dunk_audit_source_path,
        "Driving Dunk",
        "Driving Dunk",
    )
    standing_dunk_rows = load_score_csv_rows(
        standing_dunk_audit_source_path,
        "Standing Dunk",
        "Standing Dunk",
    )
    rebound_score_rows = load_excel_score_rows(
        rebound_workbook_source_path,
        "F",
        "Offensive Rebound",
    )
    defense_block_rows = load_excel_score_rows(
        defense_workbook_source_path,
        "BL",
        "Block",
    )

    workbook_universe: List[CalUniverseRow] = []
    if workbook_path.exists():
        workbook_universe = load_cal_universe(workbook_path, sheet_name=args.sheet)

    if universe_path.exists():
        universe = load_universe_csv(universe_path)
    elif workbook_universe:
        universe = workbook_universe
    else:
        raise SystemExit(
            f"Universe CSV not found: {universe_path} and workbook not found: {workbook_path}"
        )

    if details_path:
        universe = enrich_universe_rows(universe, load_universe_csv(details_path))
    if bball_physical_raw_rows:
        universe = enrich_universe_rows(universe, load_bball_detail_rows(bball_physical_raw_rows))
    if workbook_universe:
        universe = enrich_universe_rows(universe, workbook_universe)

    current_season = detect_current_season(args.current_season, universe, minutes_rows)
    contexts = build_player_contexts(
        universe=universe,
        minutes_rows=minutes_rows,
        minutes_column=args.minutes_column,
        minutes_games_column=args.minutes_games_column,
        allow_id_fallback=args.allow_id_fallback,
    )

    base_rows_by_key: Dict[Tuple[str, str], Dict[str, object]] = {}
    for universe_row in universe:
        base_rows_by_key[get_row_key(universe_row)] = {
            "Season": universe_row.season,
            "PLAYER_NAME": universe_row.player,
            "PLAYER_ID": universe_row.nba_id,
            "NBA_ID": universe_row.nba_id,
            "TEAM_ABBREVIATION": "",
        }

    merge_metric_columns(
        universe,
        base_rows_by_key,
        bball_physical_rows,
        {
            "Transition Shot Creation": "Transition Shot Creation",
            "Avg Speed Offense": "Avg Speed Offense",
            "Avg Speed Defense": "Avg Speed Defense",
            "Movement Speed Rating": "Movement Speed Rating",
            "Offensive Transition Frequency Impact": "Offensive Transition Frequency Impact",
            "Transition Frequency Impact": "Transition Frequency Impact",
            "Perimeter Isolation Defense": "Perimeter Isolation Defense",
            "Stable And 1s Per 75": "Stable And 1s Per 75",
            "Post Defense": "Post Defense",
            "Screen Assists Per 75 Possessions": "Screen Assists Per 75 Possessions",
            "Finishing Talent": "Finishing Talent",
            "Screening Talent": "Screening Talent",
            "Usage Rate": "Usage Rate",
            "True Usage": "True Usage",
            "Stable True Usage%": "Stable True Usage%",
            "Defensive Playmaking": "Defensive Playmaking",
        },
        allow_id_fallback=args.allow_id_fallback,
    )
    merge_metric_columns(
        universe,
        base_rows_by_key,
        bios_rows,
        {
            "PLAYER_HEIGHT_INCHES": "PLAYER_HEIGHT_INCHES",
            "PLAYER_WEIGHT": "PLAYER_WEIGHT",
            "AGE": "AGE",
        },
        allow_id_fallback=args.allow_id_fallback,
    )
    merge_metric_columns(
        universe,
        base_rows_by_key,
        general_scoring_rows,
        {"PCT_PTS_FB": "PCT_PTS_FB"},
        allow_id_fallback=args.allow_id_fallback,
    )
    merge_metric_columns(
        universe,
        base_rows_by_key,
        general_advanced_rows,
        {"USG_PCT": "USG_PCT"},
        allow_id_fallback=args.allow_id_fallback,
    )
    merge_metric_columns(
        universe,
        base_rows_by_key,
        minutes_rows,
        {"MIN": "MIN", "GP": "GP"},
        allow_id_fallback=args.allow_id_fallback,
    )
    merge_metric_columns(
        universe,
        base_rows_by_key,
        tracking_speed_rows,
        {
            "AVG_SPEED": "AVG_SPEED",
            "AVG_SPEED_OFF": "AVG_SPEED_OFF",
            "AVG_SPEED_DEF": "AVG_SPEED_DEF",
        },
        allow_id_fallback=args.allow_id_fallback,
    )
    merge_metric_columns(
        universe,
        base_rows_by_key,
        playtype_transition_rows,
        {"TRANSITION_POSS": "POSS", "TRANSITION_POSS_PCT": "POSS_PCT"},
        allow_id_fallback=args.allow_id_fallback,
    )
    merge_metric_columns(
        universe,
        base_rows_by_key,
        playtype_offscreen_rows,
        {"OFFSCREEN_POSS": "POSS", "OFFSCREEN_POSS_PCT": "POSS_PCT"},
        allow_id_fallback=args.allow_id_fallback,
    )
    merge_metric_columns(
        universe,
        base_rows_by_key,
        playtype_cut_rows,
        {"CUT_POSS": "POSS", "CUT_POSS_PCT": "POSS_PCT"},
        allow_id_fallback=args.allow_id_fallback,
    )
    merge_metric_columns(
        universe,
        base_rows_by_key,
        hustle_rows,
        {
            "BOX_OUTS": "BOX_OUTS",
            "SCREEN_ASSISTS": "SCREEN_ASSISTS",
            "LOOSE_BALLS_RECOVERED": "LOOSE_BALLS_RECOVERED",
            "CONTESTED_SHOTS": "CONTESTED_SHOTS",
            "DEFLECTIONS": "DEFLECTIONS",
        },
        allow_id_fallback=args.allow_id_fallback,
    )
    merge_metric_columns(
        universe,
        base_rows_by_key,
        overall_dunk_rows,
        {
            "jumpSubscore_average": "jumpSubscore_average",
            "jumpSubscore_max": "jumpSubscore_max",
            "jumpSubscore_total": "jumpSubscore_total",
        },
        allow_id_fallback=args.allow_id_fallback,
    )
    merge_metric_columns(
        universe,
        base_rows_by_key,
        post_control_rows,
        {"Post Control": "Post Control"},
        allow_id_fallback=args.allow_id_fallback,
    )
    merge_metric_columns(
        universe,
        base_rows_by_key,
        driving_dunk_rows,
        {"Driving Dunk": "Driving Dunk"},
        allow_id_fallback=args.allow_id_fallback,
    )
    merge_metric_columns(
        universe,
        base_rows_by_key,
        standing_dunk_rows,
        {"Standing Dunk": "Standing Dunk"},
        allow_id_fallback=args.allow_id_fallback,
    )
    merge_metric_columns(
        universe,
        base_rows_by_key,
        rebound_score_rows,
        {"Offensive Rebound": "Offensive Rebound"},
        allow_id_fallback=args.allow_id_fallback,
    )
    merge_metric_columns(
        universe,
        base_rows_by_key,
        defense_block_rows,
        {"Block": "Block"},
        allow_id_fallback=args.allow_id_fallback,
    )
    merge_static_metric_columns(
        universe,
        base_rows_by_key,
        players_rows,
        {
            "attributes_Speed": "attributes_Speed",
            "attributes_Agility": "attributes_Agility",
            "attributes_Strength": "attributes_Strength",
            "attributes_Vertical": "attributes_Vertical",
            "attributes_Stamina": "attributes_Stamina",
            "attributes_Hustle": "attributes_Hustle",
        },
    )
    merge_static_metric_columns(
        universe,
        base_rows_by_key,
        draft_rows,
        {
            "THREE_QUARTER_SPRINT": "THREE_QUARTER_SPRINT",
            "LANE_AGILITY_TIME": "LANE_AGILITY_TIME",
            "MODIFIED_LANE_AGILITY_TIME": "MODIFIED_LANE_AGILITY_TIME",
            "VERTICAL_LEAP": "VERTICAL_LEAP",
        },
    )

    derive_physical_metrics(base_rows_by_key.values())

    all_metric_aliases = {
        slot.metric_alias
        for section in SECTIONS
        for slot in section.component_slots
        if slot.metric_alias != "SPEED"
    }
    metric_results: Dict[str, MetricResult] = {}
    for metric_alias in sorted(all_metric_aliases):
        metric_results[metric_alias] = build_metric_result(
            contexts=contexts,
            base_rows_by_key=base_rows_by_key,
            metric_alias=metric_alias,
            current_season=current_season,
            current_season_min_threshold=args.current_season_min_threshold,
            standard_min_threshold=args.standard_min_threshold,
        )

    speed_scores, speed_z, speed_ratings, speed_rows = build_section_outputs(
        section=SPEED_SECTION,
        contexts=contexts,
        metric_results=metric_results,
    )

    metric_results["SPEED"] = MetricResult(
        raw_values=speed_scores,
        normalized_values=speed_scores,
        matched_by=["derived"] * len(speed_scores),
        mean_value=statistics.mean(speed_scores) if speed_scores else None,
        stdev_value=statistics.stdev(speed_scores) if len(speed_scores) >= 2 else None,
        source_note=METRIC_SOURCE_NOTES["SPEED"],
    )

    agility_scores, agility_z, agility_ratings, agility_rows = build_section_outputs(
        section=AGILITY_SECTION,
        contexts=contexts,
        metric_results=metric_results,
        extra_normalized_by_alias={"SPEED": speed_scores},
    )
    strength_scores, strength_z, strength_ratings, strength_rows = build_section_outputs(
        section=STRENGTH_SECTION,
        contexts=contexts,
        metric_results=metric_results,
    )
    vertical_scores, vertical_z, vertical_ratings, vertical_rows = build_section_outputs(
        section=VERTICAL_SECTION,
        contexts=contexts,
        metric_results=metric_results,
    )
    stamina_scores, stamina_z, stamina_ratings, stamina_rows = build_section_outputs(
        section=STAMINA_SECTION,
        contexts=contexts,
        metric_results=metric_results,
    )
    hustle_scores, hustle_z, hustle_ratings, hustle_rows = build_section_outputs(
        section=HUSTLE_SECTION,
        contexts=contexts,
        metric_results=metric_results,
    )

    section_payload = (
        (SPEED_SECTION, speed_scores, speed_z, speed_ratings, speed_rows),
        (AGILITY_SECTION, agility_scores, agility_z, agility_ratings, agility_rows),
        (STRENGTH_SECTION, strength_scores, strength_z, strength_ratings, strength_rows),
        (VERTICAL_SECTION, vertical_scores, vertical_z, vertical_ratings, vertical_rows),
        (STAMINA_SECTION, stamina_scores, stamina_z, stamina_ratings, stamina_rows),
        (HUSTLE_SECTION, hustle_scores, hustle_z, hustle_ratings, hustle_rows),
    )

    sheet_rows: List[List[object]] = []
    rating_only_rows: List[Dict[str, object]] = []
    audit_rows: List[Dict[str, object]] = []
    unmatched_rows: List[Dict[str, object]] = []

    for index, context in enumerate(contexts):
        base_row = base_rows_by_key[get_row_key(context.universe_row)]
        team_value = str(base_row.get("TEAM_ABBREVIATION", "")).strip()

        sheet_rows.append(
            [
                context.universe_row.nba_id,
                context.universe_row.season,
                context.universe_row.player,
                speed_ratings[index],
                speed_z[index],
                speed_scores[index],
                *speed_rows[index],
                agility_ratings[index],
                agility_z[index],
                agility_scores[index],
                *agility_rows[index],
                strength_ratings[index],
                strength_z[index],
                strength_scores[index],
                *strength_rows[index],
                vertical_ratings[index],
                vertical_z[index],
                vertical_scores[index],
                *vertical_rows[index],
                stamina_ratings[index],
                stamina_z[index],
                stamina_scores[index],
                *stamina_rows[index],
                hustle_ratings[index],
                hustle_z[index],
                hustle_scores[index],
                *hustle_rows[index],
            ]
        )

        rating_only_rows.append(
            {
                "NBA_ID": context.universe_row.nba_id,
                "Season": context.universe_row.season,
                "Player": context.universe_row.player,
                "Speed Rating": speed_ratings[index],
                "Agility Rating": agility_ratings[index],
                "Strength Rating": strength_ratings[index],
                "Vertical Rating": vertical_ratings[index],
                "Stamina Rating": stamina_ratings[index],
                "Hustle Rating": hustle_ratings[index],
            }
        )

        audit_row: Dict[str, object] = {
            "NBA_ID": context.universe_row.nba_id,
            "Season": context.universe_row.season,
            "Player": context.universe_row.player,
            "Team(s)": team_value,
            "RotationRole": context.universe_row.rotation_role,
            "MIN": context.effective_minutes,
            "WorkbookMIN": context.workbook_minutes,
            "LiveMIN": context.live_minutes,
            "LiveMINPerGame": context.live_minutes_per_game,
            "LiveGP": context.live_gp,
            "MinutesMatchedBy": context.minutes_matched_by,
            "CurrentSeason": current_season,
        }

        for section, scores, aggregate_z, ratings, _rows in section_payload:
            audit_row[f"{section.name} Score"] = scores[index]
            audit_row[f"{section.name} AggregateZ"] = aggregate_z[index]
            audit_row[f"{section.name} Rating"] = ratings[index]

        missing_labels: List[str] = []
        for section in SECTIONS:
            missing_aliases = sorted(
                {
                    slot.metric_alias
                    for slot in section.component_slots
                    if metric_results.get(slot.metric_alias) is not None
                    and metric_results[slot.metric_alias].normalized_values[index] is None
                }
            )
            if missing_aliases:
                missing_labels.extend(f"{section.name}::{alias}" for alias in missing_aliases)
            audit_row[f"{section.name} MissingCount"] = len(missing_aliases)
            audit_row[f"{section.name} MissingMetrics"] = " | ".join(missing_aliases)

        for metric_alias in sorted(metric_results):
            result = metric_results[metric_alias]
            audit_row[f"{metric_alias} Raw"] = result.raw_values[index]
            audit_row[f"{metric_alias} Z"] = result.normalized_values[index]
            audit_row[f"{metric_alias} MatchedBy"] = result.matched_by[index]
            audit_row[f"{metric_alias} Mean"] = result.mean_value
            audit_row[f"{metric_alias} Stdev"] = result.stdev_value
            audit_row[f"{metric_alias} Source"] = result.source_note

        audit_rows.append(audit_row)

        if missing_labels:
            unmatched_rows.append(
                {
                    "NBA_ID": context.universe_row.nba_id,
                    "Season": context.universe_row.season,
                    "Player": context.universe_row.player,
                    "Team(s)": team_value,
                    "RotationRole": context.universe_row.rotation_role,
                    "MIN": context.effective_minutes,
                    "MissingCount": len(missing_labels),
                    "MissingMetrics": " | ".join(missing_labels),
                }
            )

    output_prefix = args.output_prefix.strip() or "physical_all"
    sheet_path = EXPORT_DIR / f"{output_prefix}_sheet.csv"
    rating_only_path = EXPORT_DIR / f"{output_prefix}_ratings.csv"
    audit_path = EXPORT_DIR / f"{output_prefix}_audit.csv"
    unmatched_path = EXPORT_DIR / f"{output_prefix}_unmatched.csv"

    write_matrix_csv(sheet_path, WORKBOOK_COLUMNS, sheet_rows)
    write_csv(rating_only_path, RATING_ONLY_HEADERS, rating_only_rows)
    write_csv(audit_path, list(audit_rows[0].keys()) if audit_rows else [], audit_rows)
    write_csv(
        unmatched_path,
        ["NBA_ID", "Season", "Player", "Team(s)", "RotationRole", "MIN", "MissingCount", "MissingMetrics"],
        unmatched_rows,
    )

    print(f"[OK] Built Physical export for {len(sheet_rows)} player-season rows")
    print(
        "[INFO] Exact source stack: playerlist.csv identity universe, bball_index_physical.csv for the "
        "movement/usage/screening metrics, general_traditional.csv for MIN and GP, general_scoring.csv "
        "for PCT_PTS_FB, bios.csv for height/weight/age, tracking_speed.csv for AVG_SPEED values, "
        "playtype transition/offscreen/cut tables for possession rates, hustle.csv for hustle box scores, "
        "draft.csv for combine agility and vertical inputs, Upgrade/players.csv for current database "
        "attributes, finishing audits for Post Control / Driving Dunk / Standing Dunk, 2k26_rebound.xlsx "
        "for Offensive Rebound score, and 2k26_defense.xlsx for Block score."
    )
    print(
        "[INFO] Workbook behavior mirrored here: Agility consumes the Physical Speed score directly, "
        "not the Speed aggregate z-score."
    )
    print(
        "[INFO] Older seasons can still show real gaps where the live bball-index catalog does not cover "
        "pre-2015 rows or where combine / dunk tracking data is absent."
    )
    print(f"[OUT] Sheet -> {sheet_path}")
    print(f"[OUT] Ratings -> {rating_only_path}")
    print(f"[OUT] Audit -> {audit_path}")
    print(f"[OUT] Unmatched -> {unmatched_path}")


if __name__ == "__main__":
    main()
