from __future__ import annotations

import argparse
import re
import unicodedata
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

STATS_DIR = Path(__file__).resolve().parents[1]
EXPORT_DIR = STATS_DIR / "exports"
HISTORY_DIR = STATS_DIR / "history"
MANUAL_DIR = STATS_DIR / "manual"

PLAYERLIST_PATH = MANUAL_DIR / "playerlist.csv"
TEMPLATE_PATH = MANUAL_DIR / "2k26_rating.xlsx"
POSITIONS_PATH = HISTORY_DIR / "bballref_position_estimate.csv"
OUTPUT_CSV_PATH = EXPORT_DIR / "rating_all_sheet.csv"
OUTPUT_XLSX_PATH = EXPORT_DIR / "rating_all_sheet.xlsx"
OUTPUT_AUDIT_PATH = EXPORT_DIR / "rating_all_audit.csv"
OUTPUT_UNMATCHED_PATH = EXPORT_DIR / "rating_all_unmatched.csv"

KEY_COLS = ["NBA_ID", "Season", "Player"]
SUFFIXES = {"JR", "JR.", "SR", "SR.", "II", "III", "IV", "V"}
BLANKISH_TEXT = {"", "0", "NAN", "NONE", "<NA>"}

RATING_SOURCES = {
    "finishing": (EXPORT_DIR / "finishing_all_ratings.csv", {
        "Layup Rating": "Layup",
        "Standing Dunk Rating": "ST   Dunk",
        "Driving Dunk Rating": "Dunk",
        "Post Hook Rating": "PHook",
        "Post Fade Rating": "PFade",
        "Post Control Rating": "PostC",
        "Draw Foul Rating": "Foul",
        "Hands Rating": "Hands",
    }),
    "shooting": (EXPORT_DIR / "shooting_all_ratings.csv", {
        "Close Shot Rating": "Close",
        "Mid-Range Shot Rating": "Mid",
        "3-Point Shot Rating": "3PT",
        "Free Throw Rating": "FT",
        "Shot IQ Rating": "SHOTIQ",
    }),
    "playmaking": (EXPORT_DIR / "playmaking_all_ratings.csv", {
        "Ball Handle Rating": "Ball",
        "Speed with Ball Rating": "SPD/BALL",
        "Pass Accuracy Rating": "Pass",
        "Pass IQ Rating": "Pass   IQ",
        "Pass Vision Rating": "Vision",
    }),
    "defense": (EXPORT_DIR / "defense_all_ratings.csv", {
        "Interior Defense Rating": "ID",
        "Perimeter Defense Rating": "PD",
        "Steal Rating": "STL",
        "Block Rating": "BLK",
        "Help Defense IQ Rating": "HelpDIQ",
        "Pass Perception Rating": "PSPER",
    }),
    "rebounding": (EXPORT_DIR / "rebounding_all_ratings.csv", {
        "Offensive Rebound Rating": "OREB",
        "Defensive Rebound Rating": "DREB",
    }),
    "physical": (EXPORT_DIR / "physical_all_ratings.csv", {
        "Speed Rating": "SPEED",
        "Agility Rating": "Agility",
        "Strength Rating": "STR",
        "Vertical Rating": "VERT",
        "Stamina Rating": "STAM",
        "Hustle Rating": "HSTL",
    }),
    "impact": (EXPORT_DIR / "impact_all_ratings.csv", {
        "Offensive Impact Rating": "OCNST",
        "Defensive Impact Rating": "DCNST",
        "Overall Impact Rating": "INTNGBL",
    }),
    "badges": (EXPORT_DIR / "badges_all_badges.csv", {
        "Aerial Wizard": "Aerial Wizard",
        "Ankle Assassin": "Ankle Assassin",
        "Bail Out": "Bail Out",
        "Boxout Beast": "Boxout Beast",
        "Break Starter": "Break Starter",
        "Brick Wall": "Brick Wall",
        "Challenger": "Challenger",
        "Deadeye": "Deadeye",
        "Dimer": "Dimer",
        "Float Game": "Float Game",
        "Glove": "Glove",
        "Handles for Days": "Handles for Days",
        "High Flying Denier": "High Flying Denier",
        "Hook Specialist": "Hook Specialist",
        "Immovable Enforcer": "Immovable Enforcer",
        "Interceptor": "Interceptor",
        "Layup Mixmaster": "Layup Mixmaster",
        "Lightning Launch": "Lightning Launch",
        "Limitless Range": "Limitless Range",
        "Mini Marksman": "Mini Marksman",
        "Off Ball Pest": "Off Ball Pest",
        "On-Ball Menace": "On-Ball Menace",
        "Paint Patroller": "Paint Patroller",
        "Paint Prodigy": "Paint Prodigy",
        "Physical Finisher": "Physical Finisher",
        "Pick Dodger": "Pick Dodger",
        "Pogo Stick": "Pogo Stick",
        "Posterizer": "Posterizer",
        "Post Fade Phenom": "Post Fade Phenom",
        "Post Lockdown": "Post Lockdown",
        "Post Powerhouse": "Post Powerhouse",
        "Post Up Poet": "Post Up Poet",
        "Rebound Chaser": "Rebound Chaser",
        "Rise Up": "Rise Up",
        "Set Shot Specialist": "Set Shot Specialist",
        "Shifty Shooter": "Shifty Shooter",
        "Slippery Off Ball": "Slippery Off Ball",
        "Strong Handle": "Strong Handle",
        "Unpluckable": "Unpluckable",
        "Versatile Visionary": "Versatile Visionary",
    }),
}

B_BALL_TEAM_PATHS = [
    HISTORY_DIR / "bball_index_playmaking.csv",
    HISTORY_DIR / "bball_index_physical.csv",
    HISTORY_DIR / "bball_index_impact.csv",
    HISTORY_DIR / "bball_index_defense.csv",
    HISTORY_DIR / "bball_index_rebounding.csv",
]


def read_template_headers(path: Path) -> tuple[str, list[str]]:
    workbook = load_workbook(path, read_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    clean_headers = [str(header) for header in headers if header is not None and str(header).strip()]
    return worksheet.title, clean_headers


def normalize_name(text: object, drop_suffix: bool = False) -> str:
    if pd.isna(text):
        return ""
    normalized = unicodedata.normalize("NFKD", str(text))
    normalized = "".join(char for char in normalized if not unicodedata.combining(char))
    normalized = normalized.replace("&", " and ")
    normalized = normalized.replace(".", " ").replace("-", " ").replace("'", "")
    normalized = re.sub(r"[^A-Za-z0-9 ]+", " ", normalized)
    tokens = re.sub(r"\s+", " ", normalized).strip().upper().split()
    collapsed: list[str] = []
    run: list[str] = []
    for token in tokens:
        if len(token) == 1 and token.isalpha():
            run.append(token)
            continue
        if run:
            collapsed.append("".join(run))
            run = []
        collapsed.append(token)
    if run:
        collapsed.append("".join(run))
    normalized = " ".join(collapsed)
    if drop_suffix and normalized:
        tokens = normalized.split()
        if tokens and tokens[-1] in SUFFIXES:
            normalized = " ".join(tokens[:-1]).strip()
    return normalized


def split_player_name(name: object) -> tuple[str, str]:
    if pd.isna(name):
        return "", ""
    text = str(name).strip()
    if not text:
        return "", ""
    tokens = text.split()
    if len(tokens) == 1:
        return "", tokens[0]
    last_span = 2 if tokens[-1].upper() in SUFFIXES and len(tokens) >= 3 else 1
    first = " ".join(tokens[:-last_span]).strip()
    last = " ".join(tokens[-last_span:]).strip()
    return first, last


def dedupe_by_keys(frame: pd.DataFrame) -> pd.DataFrame:
    return frame.drop_duplicates(subset=KEY_COLS, keep="first").copy()


def align_lookup_on_keys(base: pd.DataFrame, frame: pd.DataFrame, keys: list[str], value_columns: list[str]) -> pd.DataFrame:
    lookup = frame.drop_duplicates(subset=keys, keep="first")[keys + value_columns].set_index(keys)
    aligned = lookup.reindex(pd.MultiIndex.from_frame(base[keys]))
    aligned = aligned.reset_index(drop=True)
    return aligned


def align_lookup(base: pd.DataFrame, frame: pd.DataFrame, value_columns: list[str]) -> pd.DataFrame:
    return align_lookup_on_keys(base, dedupe_by_keys(frame), KEY_COLS, value_columns)


def first_nonblank(*labeled_series: tuple[str, pd.Series]) -> tuple[pd.Series, pd.Series]:
    if not labeled_series:
        empty = pd.Series(dtype="object")
        return empty, empty
    result = pd.Series(pd.NA, index=labeled_series[0][1].index, dtype="object")
    source = pd.Series(pd.NA, index=labeled_series[0][1].index, dtype="object")
    for label, series in labeled_series:
        text = series.astype("string")
        normalized = text.str.strip().str.upper()
        mask = result.isna() & text.notna() & ~normalized.isin(BLANKISH_TEXT)
        result.loc[mask] = series.loc[mask]
        source.loc[mask] = label
    return result, source


def load_workbook_fallbacks(path: Path) -> pd.DataFrame:
    frame = pd.read_excel(path)
    frame = frame.rename(columns={"NBA ID": "NBA_ID"})
    needed = KEY_COLS + [
        "Team(s)",
        "PLAYER_LAST_NAME",
        "PLAYER_FIRST_NAME",
        "Primary_Position",
        "Second_Position",
    ]
    return dedupe_by_keys(frame[needed])


def load_bball_team_history() -> pd.DataFrame:
    collected: list[pd.DataFrame] = []
    for path in B_BALL_TEAM_PATHS:
        if not path.exists():
            continue
        frame = pd.read_csv(path, usecols=["Season", "Player", "Team(s)"])
        frame["NBA_ID"] = pd.NA
        collected.append(frame)
    if not collected:
        return pd.DataFrame(columns=KEY_COLS + ["Team(s)"])
    combined = pd.concat(collected, ignore_index=True)
    combined["norm_name"] = combined["Player"].map(normalize_name)
    combined["norm_name_loose"] = combined["Player"].map(lambda value: normalize_name(value, drop_suffix=True))
    combined = combined.drop_duplicates(subset=["Season", "norm_name"], keep="first")
    return combined


def load_general_team_history() -> pd.DataFrame:
    path = HISTORY_DIR / "general_traditional.csv"
    if not path.exists():
        return pd.DataFrame(columns=["NBA_ID", "Season", "Team(s)"])
    frame = pd.read_csv(path, usecols=["Season", "PLAYER_ID", "TEAM_ABBREVIATION"])
    frame = frame.rename(columns={"PLAYER_ID": "NBA_ID", "TEAM_ABBREVIATION": "Team(s)"})
    return frame.drop_duplicates(subset=["NBA_ID", "Season"], keep="first")


def load_position_history(path: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    frame = pd.read_csv(path)
    frame["norm_name"] = frame["Player"].map(normalize_name)
    frame["norm_name_loose"] = frame["Player"].map(lambda value: normalize_name(value, drop_suffix=True))
    exact = frame.drop_duplicates(subset=["Season", "norm_name"], keep="first").set_index(["Season", "norm_name"])
    loose = frame.drop_duplicates(subset=["Season", "norm_name_loose"], keep="first").set_index(["Season", "norm_name_loose"])
    return exact, loose


def main() -> None:
    parser = argparse.ArgumentParser(description="Build the flattened ratings + badges sheet.")
    parser.add_argument("--playerlist", type=Path, default=PLAYERLIST_PATH)
    parser.add_argument("--template", type=Path, default=TEMPLATE_PATH)
    parser.add_argument("--positions", type=Path, default=POSITIONS_PATH)
    parser.add_argument("--output-csv", type=Path, default=OUTPUT_CSV_PATH)
    parser.add_argument("--output-xlsx", type=Path, default=OUTPUT_XLSX_PATH)
    parser.add_argument("--output-audit", type=Path, default=OUTPUT_AUDIT_PATH)
    parser.add_argument("--output-unmatched", type=Path, default=OUTPUT_UNMATCHED_PATH)
    parser.add_argument("--fill-overall", action="store_true", help="Fill the Overall column from Overall Impact Rating.")
    args = parser.parse_args()

    sheet_name, template_headers = read_template_headers(args.template)
    base = pd.read_csv(args.playerlist).copy()
    base["norm_name"] = base["Player"].map(normalize_name)
    base["norm_name_loose"] = base["Player"].map(lambda value: normalize_name(value, drop_suffix=True))
    parsed_names = base["Player"].map(split_player_name)
    base["PLAYER_FIRST_NAME"] = parsed_names.map(lambda value: value[0])
    base["PLAYER_LAST_NAME"] = parsed_names.map(lambda value: value[1])

    workbook_fallback = load_workbook_fallbacks(args.template)
    workbook_aligned = align_lookup(
        base,
        workbook_fallback,
        ["Team(s)", "PLAYER_LAST_NAME", "PLAYER_FIRST_NAME", "Primary_Position", "Second_Position"],
    )
    for column in ("PLAYER_FIRST_NAME", "PLAYER_LAST_NAME"):
        fallback = workbook_aligned[column].astype("string").fillna("").str.strip()
        mask = fallback.ne("")
        base.loc[mask, column] = fallback.loc[mask]

    final = pd.DataFrame(index=base.index)
    for header in template_headers:
        final[header] = pd.NA

    final["NBA ID"] = base["NBA_ID"]
    final["Season"] = base["Season"]
    final["Player"] = base["Player"]
    final["PLAYER_LAST_NAME"] = base["PLAYER_LAST_NAME"].astype("string").str.strip()
    final["PLAYER_FIRST_NAME"] = base["PLAYER_FIRST_NAME"].astype("string").str.strip()

    for _, (path, rename_map) in RATING_SOURCES.items():
        source = pd.read_csv(path)
        aligned = align_lookup(base, source, list(rename_map.keys()))
        for source_column, final_column in rename_map.items():
            final[final_column] = aligned[source_column].values

    if args.fill_overall:
        final["Overall"] = final["INTNGBL"]

    bball_team_history = load_bball_team_history()
    if not bball_team_history.empty:
        exact_bball = bball_team_history.set_index(["Season", "norm_name"])["Team(s)"]
        loose_bball = bball_team_history.drop_duplicates(subset=["Season", "norm_name_loose"], keep="first").set_index(["Season", "norm_name_loose"])["Team(s)"]
        bball_team_exact = exact_bball.reindex(pd.MultiIndex.from_frame(base[["Season", "norm_name"]])).reset_index(drop=True)
        bball_team_loose = loose_bball.reindex(pd.MultiIndex.from_frame(base[["Season", "norm_name_loose"]])).reset_index(drop=True)
    else:
        bball_team_exact = pd.Series(pd.NA, index=base.index, dtype="object")
        bball_team_loose = pd.Series(pd.NA, index=base.index, dtype="object")

    general_teams = load_general_team_history()
    general_team_aligned = (
        align_lookup_on_keys(base, general_teams, ["NBA_ID", "Season"], ["Team(s)"])
        if not general_teams.empty
        else pd.DataFrame({"Team(s)": pd.Series(pd.NA, index=base.index)})
    )

    positions_exact, positions_loose = load_position_history(args.positions)
    exact_position_aligned = positions_exact.reindex(pd.MultiIndex.from_frame(base[["Season", "norm_name"]])).reset_index(drop=True)
    loose_position_aligned = positions_loose.reindex(pd.MultiIndex.from_frame(base[["Season", "norm_name_loose"]])).reset_index(drop=True)

    bbr_team_exact = exact_position_aligned.get("Team", pd.Series(pd.NA, index=base.index))
    bbr_team_loose = loose_position_aligned.get("Team", pd.Series(pd.NA, index=base.index))

    team_values, team_sources = first_nonblank(
        ("workbook", workbook_aligned["Team(s)"]),
        ("bball_index_exact", bball_team_exact),
        ("bball_index_loose", bball_team_loose),
        ("general_traditional", general_team_aligned["Team(s)"]),
        ("bballref_exact", bbr_team_exact),
        ("bballref_loose", bbr_team_loose),
    )

    impact_source = pd.read_csv(RATING_SOURCES["impact"][0], usecols=KEY_COLS + ["Team(s)"])
    impact_team_aligned = align_lookup(base, impact_source, ["Team(s)"])
    impact_team = impact_team_aligned["Team(s)"]
    impact_mask = team_values.isna() & impact_team.astype("string").notna() & impact_team.astype("string").str.strip().ne("")
    team_values.loc[impact_mask] = impact_team.loc[impact_mask]
    team_sources.loc[impact_mask] = "impact"
    final["Team(s)"] = team_values

    primary_values, primary_sources = first_nonblank(
        ("bballref_exact", exact_position_aligned.get("Primary_Position", pd.Series(pd.NA, index=base.index))),
        ("bballref_loose", loose_position_aligned.get("Primary_Position", pd.Series(pd.NA, index=base.index))),
        ("workbook", workbook_aligned["Primary_Position"]),
    )
    second_values, second_sources = first_nonblank(
        ("bballref_exact", exact_position_aligned.get("Second_Position", pd.Series(pd.NA, index=base.index))),
        ("bballref_loose", loose_position_aligned.get("Second_Position", pd.Series(pd.NA, index=base.index))),
        ("workbook", workbook_aligned["Second_Position"]),
    )

    final["Primary_Position"] = primary_values
    final["Second_Position"] = second_values

    final = final.reindex(columns=template_headers)

    audit = pd.DataFrame(
        {
            "NBA ID": final["NBA ID"],
            "Season": final["Season"],
            "Player": final["Player"],
            "Team(s)": final["Team(s)"],
            "TeamSource": team_sources,
            "Primary_Position": final["Primary_Position"],
            "PrimaryPositionSource": primary_sources,
            "Second_Position": final["Second_Position"],
            "SecondPositionSource": second_sources,
        }
    )

    def is_blank(value: object) -> bool:
        if pd.isna(value):
            return True
        if isinstance(value, str):
            return value.strip().upper() in BLANKISH_TEXT
        return False

    optional_blank_columns = {"Overall", "Second_Position"}
    missing_columns = []
    for _, row in final.iterrows():
        missing: list[str] = []
        for column in template_headers:
            if column in optional_blank_columns:
                continue
            if is_blank(row[column]):
                missing.append(column)
        missing_columns.append(" | ".join(missing))

    unmatched = pd.DataFrame(
        {
            "NBA ID": final["NBA ID"],
            "Season": final["Season"],
            "Player": final["Player"],
            "Team(s)": final["Team(s)"],
            "Primary_Position": final["Primary_Position"],
            "Second_Position": final["Second_Position"],
            "MissingColumns": missing_columns,
        }
    )
    unmatched = unmatched[unmatched["MissingColumns"].astype(str).str.strip().ne("")]

    args.output_csv.parent.mkdir(parents=True, exist_ok=True)
    final.to_csv(args.output_csv, index=False, encoding="utf-8-sig")
    with pd.ExcelWriter(args.output_xlsx, engine="openpyxl") as writer:
        final.to_excel(writer, index=False, sheet_name=sheet_name[:31] or "Rating")
    audit.to_csv(args.output_audit, index=False, encoding="utf-8-sig")
    unmatched.to_csv(args.output_unmatched, index=False, encoding="utf-8-sig")

    print(f"[WROTE] {args.output_csv} ({len(final)} rows)")
    print(f"[WROTE] {args.output_xlsx}")
    print(f"[WROTE] {args.output_audit}")
    print(f"[WROTE] {args.output_unmatched} ({len(unmatched)} rows)")


if __name__ == "__main__":
    main()
